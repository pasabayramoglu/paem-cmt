# PAEM-CMT desktop application.

"""
PAEM-CMT is a desktop tool for purpose-aligned machine translation evaluation.

It loads source and translated documents for local preview, lets the user define
purpose, role expectations, and terminology, supports AI-assisted drafting for
selected context fields, runs repeated PAEM-CMT evaluations with stability
checks, and exports HTML and spreadsheet reports.

Document loading, parsing, previewing, and report generation run locally.
Network access is used only for OpenAI-backed features. The API key is kept in
memory for the current session and is not written to disk.

Core dependencies include PyQt5, the OpenAI SDK, python-docx, mammoth,
BeautifulSoup, markdown, chardet, colorama, and openpyxl.
""" 

import sys
import sys, traceback
from pathlib import Path
from PyQt5.QtWidgets import QMessageBox

def _excepthook(exc_type, exc, tb):
    txt = "".join(traceback.format_exception(exc_type, exc, tb))
    try:
        QMessageBox.critical(None, "Unhandled error", txt)
    except Exception:
        pass
    try:
        print(txt)
    except Exception:
        pass

sys.excepthook = _excepthook

def resource_path(name: str) -> str:
    return str((Path(__file__).resolve().parent / name).resolve())

import re
import json
import hashlib
import chardet
import docx
import mammoth
import markdown
import statistics
import warnings
try:
    from PyQt5 import sip
except Exception:
    import sip
from openai import OpenAI
from colorama import Fore, Style, init
init(autoreset=True)
from bs4 import BeautifulSoup
from docx import Document
from docx.shared import Pt, RGBColor

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QWidget, QVBoxLayout, QLabel,
    QPushButton, QTextEdit, QTableWidget, QTableWidgetItem, QComboBox,
    QMessageBox, QAction, QToolBar, QHeaderView, QSizePolicy, QGridLayout,
    QHBoxLayout, QSplashScreen, QGroupBox, QScrollArea, QSplitter, QDialog,
    QSpinBox, QLineEdit, QPlainTextEdit, QFrame, QDialogButtonBox, QCheckBox, QScroller, QProgressBar, QAbstractItemView
)
from PyQt5.QtWidgets import QGraphicsDropShadowEffect
from PyQt5.QtCore import (
    Qt, QEvent, QTimer, QEventLoop, QLoggingCategory, QDate,
    pyqtSignal, QObject, QThread, QUrl, QPropertyAnimation, QEasingCurve, QPoint, QRect, QSizeF
)
from PyQt5.QtGui import (
    QTextCharFormat, QFont, QFontMetrics, QTextCursor, QPixmap, QIcon,
    QGuiApplication, QColor, QDesktopServices, QCursor, QPainter, QPen, QPainterPath, QImage, QLinearGradient, QRadialGradient, QBrush
)
import time, math, unicodedata
USE_DETERMINISTIC_AUDIT = True
from difflib import SequenceMatcher
from collections import deque

# Robust statistics helpers.
def _mad(seq):
    m = statistics.median(seq)
    return statistics.median([abs(x - m) for x in seq]) or 1e-9

def _huber_mean(seq, c=1.25, it=10):
    mu = statistics.median(seq)
    s  = 1.4826 * _mad(seq)
    if s < 1e-6:
        mu = statistics.mean(seq)
        sd = statistics.stdev(seq) if len(seq) > 1 else 0.0
        return mu, sd
    for _ in range(it):
        w = []
        for x in seq:
            z = (x - mu) / s
            w.append(1.0 if abs(z) <= c else (c / (abs(z) + 1e-9)))
        mu = sum(w_i * x_i for w_i, x_i in zip(w, seq)) / (sum(w) or 1e-9)
    cap = c * s
    var = sum(min((x - mu) ** 2, cap ** 2) for x in seq) / max(len(seq) - 1, 1)
    if var <= 0.0 and len(seq) > 1:
        var = statistics.variance(seq)
    return mu, (var ** 0.5)

def _k2_split(xs, it=15):
    xs = sorted(float(x) for x in xs)
    if len(xs) < 4:
        return [xs], [statistics.mean(xs)]
    c1, c2 = xs[0], xs[-1]
    for _ in range(it):
        g1, g2 = [], []
        for x in xs:
            (g1 if abs(x - c1) <= abs(x - c2) else g2).append(x)
        if not g1 or not g2:
            break
        nc1 = sum(g1)/len(g1); nc2 = sum(g2)/len(g2)
        if abs(nc1 - c1) < 1e-6 and abs(nc2 - c2) < 1e-6:
            break
        c1, c2 = nc1, nc2
    groups = [g for g in (g1, g2) if g]
    cents  = [sum(g)/len(g) for g in groups]
    return groups, cents


try:
    from openpyxl import load_workbook, Workbook
except Exception:
    load_workbook = None
    Workbook = None


# Reduce noisy Qt platform warnings and sip deprecation warnings.
QLoggingCategory.setFilterRules("qt.qpa.*=false")
warnings.filterwarnings("ignore", message="sipPyTypeDict\\(\\) is deprecated", category=DeprecationWarning)

client = None
COMPLETION_CAP = 128000


_SESSION_API_KEY = ""

def get_api_key() -> str:
    return _SESSION_API_KEY

def set_api_key(k: str):
    global _SESSION_API_KEY
    _SESSION_API_KEY = (k or "").strip()

def new_client() -> OpenAI:
    k = get_api_key()
    if not k:
        raise RuntimeError("No OpenAI API key set.")
    return OpenAI(api_key=k)

def _gpt5_effort_none_kwargs(model: str | None) -> dict:
    """
    GPT-5.* only supports temperature/top_p when reasoning effort is set to "none".
    """
    m = (model or "").strip().lower()
    return {"reasoning_effort": "none"} if m.startswith("gpt-5") else {}


ANSI_RE = re.compile(r"\x1B\[[0-?]*[ -/]*[@-~]")

def strip_html_tags(text):
    """
    Convert HTML to plain text while preserving intended line breaks.
    """
    t = text or ""

    t = re.sub(r'(?i)<br\s*/?>', "\n", t)
    t = re.sub(r'(?i)</p\s*>', "\n", t)
    t = re.sub(r'(?i)</div\s*>', "\n", t)
    t = re.sub(r'(?i)</li\s*>', "\n", t)

    t = re.sub(r"<[^>]+>", "", t)

    t = t.replace("\r", "\n")
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n\s*\n+", "\n", t)
    t = re.sub(r" *\n *", "\n", t)
    return t.strip()

# ---------- Live console plumbing ----------
class EmittingStream(QObject):
    """Tee stdout/stderr to the original stream and a Qt signal."""
    textWritten = pyqtSignal(str)

    def __init__(self, tee=None):
        super().__init__()
        self._tee = tee

    def write(self, text):
        try:
            if self._tee:
                self._tee.write(text)
                self._tee.flush()
        except Exception:
            pass
        self.textWritten.emit(text)

    def flush(self):
        try:
            if self._tee:
                self._tee.flush()
        except Exception:
            pass
# -----------------------------------------

class LiveStats:
    """Rolling stats for latency, ETA, token usage, and remaining rate limits."""
    def __init__(self, maxlen: int = 20, on_update=None):
        self.times   = deque(maxlen=maxlen)
        self.tokens  = 0
        self.rpm_rem = None
        self.tpm_rem = None
        self.on_update = on_update

    def push(
        self,
        elapsed: float,
        usage: dict,
        headers: dict | None,
        run_idx: int,
        total_runs: int,
        score: float | None = None
    ):
        """Update metrics for a completed run and print a compact status line."""
        headers = headers or {}
        self.times.append(elapsed)
        self.tokens += usage.get("total_tokens", 0)

        self.rpm_rem = headers.get("x-ratelimit-remaining-requests", self.rpm_rem)
        self.tpm_rem = headers.get("x-ratelimit-remaining-tokens",   self.tpm_rem)

        avg = sum(self.times) / len(self.times)
        eta_sec = avg * (total_runs - run_idx - 1)
        m, s = divmod(int(round(eta_sec)), 60)
        eta_str = f"{m}m {s:02d}s"

        bar_done = Fore.CYAN + "#" * (run_idx + 1) + Style.RESET_ALL
        bar_todo = "-" * (total_runs - run_idx - 1)
        bar = f"[{bar_done}{bar_todo}]"

        score_part = f" | run score {Fore.WHITE}{score:.2f}{Style.RESET_ALL}" if score is not None else ""

        line = (
            f"{bar} run {run_idx+1}/{total_runs} "
            f"| {Fore.YELLOW}{elapsed:4.2f}s{Style.RESET_ALL} "
            f"(avg {avg:4.2f}s) "
            f"| ETA ≈ {Fore.YELLOW}{eta_str}{Style.RESET_ALL} "
            f"| used {Fore.GREEN}{self.tokens}{Style.RESET_ALL} tok "
            f"| RPM {Fore.CYAN}{self.rpm_rem}{Style.RESET_ALL}  "
            f"TPM {Fore.CYAN}{self.tpm_rem}{Style.RESET_ALL}"
            f"{score_part}"
        )
        print(line)
        if self.on_update:
            self.on_update(ANSI_RE.sub("", line))

class SplashScreen(QSplashScreen):
    def __init__(self):
        logo = QPixmap(resource_path("paemcmtlogo.png"))
        if not logo.isNull():
            logo = logo.scaled(270, 270, Qt.KeepAspectRatio, Qt.SmoothTransformation)

        canvas_w = max((logo.width() if not logo.isNull() else 0) + 80, 360)
        canvas_h = max((logo.height() if not logo.isNull() else 0) + 120, 360)

        canvas = QPixmap(canvas_w, canvas_h)
        canvas.fill(QColor("#0b1016"))

        painter = QPainter(canvas)
        try:
            painter.setRenderHint(QPainter.Antialiasing, True)
            painter.setRenderHint(QPainter.SmoothPixmapTransform, True)
            painter.setRenderHint(QPainter.TextAntialiasing, True)

            if not logo.isNull():
                x = (canvas.width() - logo.width()) // 2
                painter.drawPixmap(x, 20, logo)

            painter.setPen(QColor("#cfe2ff"))
            painter.setFont(QFont("Segoe UI", 11, QFont.DemiBold))
            text_rect = QRect(24, canvas.height() - 76, canvas.width() - 48, 56)
            painter.drawText(
                text_rect,
                Qt.AlignHCenter | Qt.AlignTop | Qt.TextWordWrap,
                "PAEM-CMT — Purpose-Aligned\nEvaluation for Customized MT"
            )
        finally:
            painter.end()

        super().__init__(canvas)
        self.setWindowFlag(Qt.FramelessWindowHint, True)
        self.setWindowOpacity(1.0)

    def fade_finish(self, widget, duration=520):
        self._fade_anim = QPropertyAnimation(self, b"windowOpacity", self)
        self._fade_anim.setDuration(duration)
        self._fade_anim.setStartValue(1.0)
        self._fade_anim.setEndValue(0.0)
        self._fade_anim.setEasingCurve(QEasingCurve.OutCubic)

        loop = QEventLoop(self)
        self._fade_anim.finished.connect(loop.quit)
        self._fade_anim.start()
        loop.exec_()

        self.finish(widget)

class ApiKeyDialog(QDialog):
    """Modal dialog for capturing the OpenAI API key for the current session."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔑 OpenAI API Key")
        self.setModal(True)

        lay = QVBoxLayout(self)
        title = QLabel("Paste your OpenAI API key:")
        title.setStyleSheet("font-weight:600;")
        lay.addWidget(title)

        self.edit = QLineEdit(self)
        self.edit.setPlaceholderText("sk-…")
        self.edit.setEchoMode(QLineEdit.Password)
        lay.addWidget(self.edit)

        row = QHBoxLayout()
        ok = QPushButton("✅ OK", self); ok.clicked.connect(self._on_ok); ok.setObjectName("primary")
        cancel = QPushButton("❌ Quit", self); cancel.clicked.connect(self.reject); cancel.setObjectName("quiet")
        row.addStretch(1); row.addWidget(ok); row.addWidget(cancel)
        lay.addLayout(row)

    def _on_ok(self):
        if not self.api_key():
            QMessageBox.warning(self, "Missing key", "Please enter your API key.")
            return
        self.accept()

    def api_key(self) -> str:
        return self.edit.text().strip()

# -------------------- Background eval worker --------------------
class EvalWorker(QObject):
    progress = pyqtSignal(str)
    finished = pyqtSignal(dict)
    error    = pyqtSignal(str)

    def __init__(self, tool, max_runs=40, min_runs=5, threshold=0.05):
        super().__init__()
        self.tool = tool
        self.max_runs = max_runs
        self.min_runs = min_runs
        self.threshold = threshold

    def run(self):
        try:
            data = self.tool._perform_evaluation_core(
                max_runs=self.max_runs,
                min_runs=self.min_runs,
                threshold=self.threshold,
                on_pulse=lambda msg: self.progress.emit(msg),
            )
            self.finished.emit(data)
        except Exception as e:
            self.error.emit(str(e))
# ---------------------------------------------------------------

# --- Drag-and-drop container for file inputs ---
class DropArea(QWidget):
    fileDropped = pyqtSignal(str)

    def __init__(self, exts: tuple[str, ...]):
        super().__init__()
        self._exts = tuple(e.lower() for e in exts)
        self.setAcceptDrops(True)

    def _ok(self, urls) -> str | None:
        if not urls:
            return None
        path = urls[0].toLocalFile()
        if not path:
            return None
        low = path.lower()
        return path if any(low.endswith(e) for e in self._exts) else None

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls() and self._ok(e.mimeData().urls()):
            e.acceptProposedAction()
        else:
            e.ignore()

    def dropEvent(self, e):
        p = self._ok(e.mimeData().urls())
        if p:
            self.fileDropped.emit(p)
            e.acceptProposedAction()
        else:
            e.ignore()

class ModernEditor(QTextEdit):
    """Editable text area with overlay scrollbar and kinetic scrolling."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(False)
        self.setMouseTracking(True)
        self.viewport().setMouseTracking(True)
        self.setViewportMargins(0, 0, 8, 0)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.verticalScrollBar().setStyleSheet("""
            QScrollBar:vertical {
                background: transparent;
                width: 10px;
                margin: 6px 4px 6px 4px;
            }
            QScrollBar::handle:vertical {
                background: rgba(180,200,230,0.45);
                border-radius: 6px;
                min-height: 24px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgba(180,200,230,0.75);
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px; width: 0px; border: 0; background: transparent;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: transparent;
            }
        """)
        if QScroller is not None:
            try:
                QScroller.grabGesture(self.viewport(), QScroller.LeftMouseButtonGesture)
            except Exception:
                pass
        self._reveal_ms = 1300
        self._hide_timer = QTimer(self); self._hide_timer.setSingleShot(True)
        self._hide_timer.timeout.connect(self._maybe_hide_scrollbar)
        self._anim = None

    def _reveal_scrollbar(self):
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._hide_timer.start(self._reveal_ms)

    def _maybe_hide_scrollbar(self):
        if not self.rect().contains(self.mapFromGlobal(QCursor.pos())) and not self._anim_running():
            self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

    def _anim_running(self):
        return (self._anim is not None) and (self._anim.state() == QPropertyAnimation.Running)

    def enterEvent(self, e):
        self._reveal_scrollbar(); super().enterEvent(e)
    def leaveEvent(self, e):
        self._hide_timer.start(300); super().leaveEvent(e)
    def mouseMoveEvent(self, e):
        self._reveal_scrollbar(); super().mouseMoveEvent(e)
    def keyPressEvent(self, e):
        self._reveal_scrollbar(); super().keyPressEvent(e)

    def wheelEvent(self, e):
        self._reveal_scrollbar()
        sb = self.verticalScrollBar(); cur = sb.value()
        dy = e.pixelDelta().y()
        if dy == 0: dy = e.angleDelta().y()
        page = sb.pageStep() or self.viewport().height() or 100
        factor = 0.10 if not e.pixelDelta().isNull() else 0.30
        delta = - (dy / 120.0) * (page * factor) if dy else 0
        tgt = int(max(sb.minimum(), min(sb.maximum(), cur + delta)))
        if self._anim is not None: self._anim.stop()
        self._anim = QPropertyAnimation(sb, b"value", self)
        self._anim.setDuration(180); self._anim.setEasingCurve(QEasingCurve.OutCubic)
        self._anim.setStartValue(cur); self._anim.setEndValue(tgt)
        self._anim.finished.connect(lambda: self._hide_timer.start(self._reveal_ms))
        self._anim.start(); e.accept()

class ModernPreview(QTextEdit):
    """Read-only preview with overlay scrollbar, smooth wheel animation, and optional kinetic drag."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
        self.setMouseTracking(True)
        self.viewport().setMouseTracking(True)

        self.setViewportMargins(0, 0, 8, 0)

        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        self.verticalScrollBar().setStyleSheet("""
            QScrollBar:vertical {
                background: transparent;
                width: 10px;
                margin: 6px 4px 6px 4px;
            }
            QScrollBar::handle:vertical {
                background: rgba(180,200,230,0.45);
                border-radius: 6px;
                min-height: 24px;
            }
            QScrollBar::handle:vertical:hover {
                background: rgba(180,200,230,0.75);
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px; width: 0px; border: 0; background: transparent;
            }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: transparent;
            }
        """)

        if QScroller is not None:
            try:
                QScroller.grabGesture(self.viewport(), QScroller.LeftMouseButtonGesture)
            except Exception:
                pass

        self._reveal_ms = 1300
        self._hide_timer = QTimer(self)
        self._hide_timer.setSingleShot(True)
        self._hide_timer.timeout.connect(self._maybe_hide_scrollbar)

        self._anim = None

    def _reveal_scrollbar(self):
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._hide_timer.start(self._reveal_ms)

    def _maybe_hide_scrollbar(self):
        if not self.rect().contains(self.mapFromGlobal(QCursor.pos())) and not self._anim_running():
            self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

    def _anim_running(self):
        return (self._anim is not None) and (self._anim.state() == QPropertyAnimation.Running)

    def enterEvent(self, e):
        self._reveal_scrollbar()
        super().enterEvent(e)

    def leaveEvent(self, e):
        self._hide_timer.start(300)
        super().leaveEvent(e)

    def mouseMoveEvent(self, e):
        self._reveal_scrollbar()
        super().mouseMoveEvent(e)

    def keyPressEvent(self, e):
        self._reveal_scrollbar()
        super().keyPressEvent(e)

    def wheelEvent(self, e):
        self._reveal_scrollbar()

        sb = self.verticalScrollBar()
        cur = sb.value()

        dy = e.pixelDelta().y()
        if dy == 0:
            dy = e.angleDelta().y()

        page = sb.pageStep() or self.viewport().height() or 100

        factor = 0.10 if not e.pixelDelta().isNull() else 0.30
        delta = - (dy / 120.0) * (page * factor) if dy else 0

        tgt = int(max(sb.minimum(), min(sb.maximum(), cur + delta)))

        if self._anim is not None:
            self._anim.stop()
        self._anim = QPropertyAnimation(sb, b"value", self)
        self._anim.setDuration(180)
        self._anim.setEasingCurve(QEasingCurve.OutCubic)
        self._anim.setStartValue(cur)
        self._anim.setEndValue(tgt)
        self._anim.finished.connect(lambda: self._hide_timer.start(self._reveal_ms))
        self._anim.start()

        e.accept()

# -------------------- Background workers: Quick Fill & Purpose --------------------
class QuickFillWorker(QObject):
    finished = pyqtSignal(dict)
    error    = pyqtSignal(str)

    def __init__(self, model: str, sl: str, tl: str, notes: str):
        super().__init__()
        self.model, self.sl, self.tl, self.notes = model, sl, tl, notes

    def run(self):
        try:
            prompt = f"""
            You will receive raw notes for four roles about translation constraints.

            TASK
            - Normalize and structure with MICRO-COMPRESSION (shorten without generalizing).
            - Preserve original constraints/wording; only micro-fix typos.
            - **DO NOT ANONYMIZE**: keep all proper nouns, organizations, places, geographies, culture markers, domains, and quoted strings verbatim.
            - Do NOT add ideas.

            OUTPUT (STRICT JSON)
            - Keys: target_audience, translator, source_owner, commissioner
            - Each value: a MULTILINE string; every line starts with "• ".
            - Max 5 bullets per role. If more, MERGE related points using ";" or " / " (do not drop constraints).
            - Each bullet ≤ 22 words. Delete filler/hedges/redundancy; KEEP negations, named entities, locales, institutions, domain terms.

            CONVERSION RULES
            - Map labels: “Target audience”→target_audience; “Translator”→translator;
              “Source text owner/Author”→source_owner; “Commissioner/Publisher”→commissioner.
            - Keep original order per role. Prefer imperative mood (“Use…”, “Avoid…”, “Do not…”).
            - If notes include a “Topic/Context/Scenario” line, **carry its concrete specifics verbatim** (who/where/what domain/audience) into the most relevant role bullets:
              • Always include key specifics in target_audience first bullet.  
              • Also include goal-relevant specifics in commissioner if applicable.  
              (Copy-only; no re-interpretation.)

            Source language: {self.sl}
            Target language: {self.tl}

            NOTES
            {self.notes}
            """.strip()

            print(f"[QF] model={self.model}  sl={self.sl} tl={self.tl}")
            raw = client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "You are a precise information conserver. Output VALID JSON only (no code fences). Never anonymize or paraphrase named entities, places, institutions, or culturally specific terms—copy them verbatim. Shorten by deleting filler, not content nouns/modifiers. Preserve constraints exactly."},
                    {"role": "user",   "content": prompt}
                ],
                max_completion_tokens=128000,
                temperature=0.0,
                **_gpt5_effort_none_kwargs(self.model),
            ).choices[0].message.content.strip()

            if raw.startswith("```"):
                raw = re.sub(r"^```(?:json)?|```$", "", raw, flags=re.S).strip()
            match = re.search(r"\{.*\}", raw, re.S)
            if not match:
                raise RuntimeError("No JSON object in response.")
            data = json.loads(match.group())
            print("[QF] received keys:", ", ".join(sorted(data.keys())))
            self.finished.emit({
                "target_audience": data.get("target_audience","").strip(),
                "translator":      data.get("translator","").strip(),
                "source_owner":    data.get("source_owner","").strip(),
                "commissioner":    data.get("commissioner","").strip(),
            })
        except Exception as e:
            self.error.emit(str(e))


class PurposeWorker(QObject):
    finished = pyqtSignal(str)
    error    = pyqtSignal(str)

    def __init__(self, model: str, sl: str, tl: str, roles: dict):
        super().__init__()
        self.model, self.sl, self.tl, self.roles = model, sl, tl, roles

    def run(self):
        try:
            roles_str = "\n".join(f"- {k}: {v or '(not specified)'}" for k, v in self.roles.items())

            prompt = f"""
            Write a clear content that states the intended purpose (skopos) of the translation,
            grounded in the role expectations below.

            Rules
            - MICRO-COMPRESSION: shorten by deleting filler/redundancy; do NOT paraphrase or generalize.
            - DO NOT ANONYMIZE: keep proper nouns, places, institutions, culture markers, and domain terms verbatim.
            - Include salient specifics (who/where/what domain) when present.
            - Neutral, inclusive wording. No invented goals.

            Source language: {self.sl or '—'}
            Target language: {self.tl or '—'}

            Role expectations:
            {roles_str}

            Return only the sentence (no quotes, no preface).
            """.strip()


            print(f"[PURPOSE] model={self.model}  sl={self.sl} tl={self.tl}")
            rsp = client.chat.completions.create(
                model=self.model,
                max_completion_tokens=128000,
                messages=[
                    {"role": "system", "content": "Never anonymize or paraphrase named entities, places, institutions, or domain terms—copy them verbatim. Use micro-compression by deleting filler only. No preface, no JSON."},
                    {"role": "user",   "content": prompt},
                ],
                temperature=0.0,
                **_gpt5_effort_none_kwargs(self.model),
            )
            print("[PURPOSE] response received.")
            sent = (rsp.choices[0].message.content or "").strip().strip('"')
            if not sent:
                raise RuntimeError("Empty response.")
            print("[PURPOSE] draft generated.")
            self.finished.emit(sent)
        except Exception as e:
            self.error.emit(str(e))
# -------------------------------------------------------------------------------

class LiveChart(QWidget):
    """
    Live score chart.
      • Fixed Y range: [0.00, 5.00]
      • Score sparkline and dots
      • Per-point labels with collision handling
      • Distinct plot background for readability
    """
    def __init__(self, parent=None, threshold: float = 0.05, active_after: int = 10):
        super().__init__(parent)
        self.scores: list[float] = []
        self.mu:     list[float] = []
        self.ci:     list[float] = []
        self.threshold   = float(threshold)
        self.active_after = int(active_after)
        self.setMinimumSize(420, 260)

        self.bg         = QColor("#0b1016")
        self.plot_top   = QColor("#0f2239")
        self.plot_bot   = QColor("#0b1627")
        self.vignette   = QColor(5, 10, 18, 115)
        self.border     = QColor("#284263")
        self.text       = QColor("#e1ecff")
        self.muted      = QColor("#a9bfdc")
        self.grid       = QColor(28, 54, 84, 155)
        self.grid_minor = QColor(22, 40, 66, 135)

        self.line_col   = QColor("#a9c7ff")
        self.dot_fill   = QColor("#d7e6ff")
        self.dot_pulse  = QColor(30, 120, 255, 38)
        self.label_bg   = QColor(18, 26, 40, 210)
        self.label_bd   = QColor(55, 93, 150, 180)
        self._seen_keys: set[str] = set()
        self._seen_mu_keys: set[str] = set()
        self._last_appended_val = None
        self._last_append_ts = 0.0

    def set_threshold(self, t: float):
        self.threshold = float(t); self.update()

    def push_score(self, score: float, key: str | None = None):
        try:
            s4 = round(float(score), 4)
            now = time.monotonic()

            if key is not None and key in self._seen_keys:
                return

            if self._last_appended_val == s4 and (now - self._last_append_ts) < 1.5:
                return

            if key is not None:
                self._seen_keys.add(key)

            self.scores.append(max(0.0, min(5.0, s4)))
            self._last_appended_val = s4
            self._last_append_ts = now
            self.update()
        except Exception:
            pass

    def push_mu_ci(self, mu: float, ci: float, key: str | None = None):
        try:
            if key and key in self._seen_mu_keys:
                return
            if key:
                self._seen_mu_keys.add(key)
            self.mu.append(max(0.0, min(5.0, float(mu))))
            self.ci.append(max(0.0, float(ci)))
            self.update()
        except Exception:
            pass

    def sizeHint(self): 
        return self.minimumSize()

    def _map(self, v, vmin, vmax, a, b):
        if vmax <= vmin: return a
        t = (v - vmin) / (vmax - vmin)
        return a + t * (b - a)

    def _auto_label_step_and_font(self, xs, width_px):
        """Return (step, font_px) based on label density."""
        N = max(1, len(xs))
        if N < 2:
            return 1, 11
        min_dx = min(xs[i] - xs[i-1] for i in range(1, N))
        font_px = 11
        if N > 10: font_px = 10
        if N > 16: font_px = 9
        if N > 22: font_px = 8
        if N > 28: font_px = 7

        if min_dx < 26: font_px -= 1
        if min_dx < 18: font_px -= 1
        font_px = max(7, font_px)

        step = 1
        if min_dx < 24 or N > 22: step = 2
        if min_dx < 16 or N > 34: step = 3
        return step, font_px

    def paintEvent(self, _):
        p = QPainter(self)
        try:
            p.setRenderHint(QPainter.Antialiasing, True)
            full = self.rect()

            p.fillRect(full, self.bg)

            r = full.adjusted(16, 20, -16, -24)

            grad = QLinearGradient(r.topLeft(), r.bottomLeft())
            grad.setColorAt(0.0, self.plot_top)
            grad.setColorAt(1.0, self.plot_bot)
            p.fillRect(r, QBrush(grad))

            vg = QRadialGradient(r.center(), max(r.width(), r.height()))
            vg.setColorAt(0.75, QColor(0,0,0,0))
            vg.setColorAt(1.00, self.vignette)
            p.fillRect(r, QBrush(vg))

            p.setPen(QPen(self.border, 1))
            p.drawRoundedRect(r, 10, 10)

            lo, hi = 0.0, 5.0
            xpad = max(10, int(r.width() * 0.02))
            x0, x1 = r.left()+48 + xpad,  r.right()-14 - xpad
            y0, y1 = r.bottom()-28,       r.top()+12

            for v in [i/2 for i in range(0, 11)]:
                y = self._map(v, lo, hi, y0, y1)
                col = self.grid if abs(v - round(v)) < 1e-6 else self.grid_minor
                p.setPen(QPen(col, 1))
                p.drawLine(int(x0), int(round(y)), int(x1), int(round(y)))
                if abs(v - round(v)) < 1e-6:
                    p.setPen(self.muted)
                    f = QFont("Segoe UI", 9)
                    p.setFont(f)
                    p.drawText(QRect(r.left()+6, int(round(y))-10, 40, 20),
                               Qt.AlignRight | Qt.AlignVCenter, f"{v:.2f}")

            p.setPen(QPen(self.border, 1))
            p.drawLine(int(x0), int(y0), int(x1), int(y0))

            N = len(self.scores)
            if N == 0:
                p.setPen(self.muted); p.setFont(QFont("Segoe UI", 10))
                p.drawText(r, Qt.AlignCenter, "Awaiting first scored run…")
                return

            xs = [self._map(i, 0, max(1, N-1), x0, x1) for i in range(N)]
            ys = [self._map(s, lo, hi, y0, y1) for s in self.scores]

            p.setPen(QPen(self.line_col, 2))
            path = QPainterPath()
            path.moveTo(xs[0], ys[0])
            for i in range(1, N):
                path.lineTo(xs[i], ys[i])
            p.drawPath(path)

            p.setBrush(self.dot_fill)
            p.setPen(Qt.NoPen)
            for i in range(N):
                p.drawEllipse(QPoint(int(round(xs[i])), int(round(ys[i]))), 3, 3)

            lx, ly = int(round(xs[-1])), int(round(ys[-1]))
            p.setBrush(self.dot_pulse)
            p.drawEllipse(QPoint(lx, ly), 8, 8)

            step, font_px = self._auto_label_step_and_font(xs, r.width())
            f = QFont("Segoe UI", font_px, QFont.DemiBold)
            p.setFont(f)

            placed = []
            for i in range(0, N, step):
                txt = f"{self.scores[i]:.2f}"
                fm  = p.fontMetrics()
                tw  = fm.horizontalAdvance(txt)
                th  = fm.height()

                cx  = int(round(xs[i]))
                cy  = int(round(ys[i]))

                for dy in (-14, -26, -38, 8):
                    rx = cx - (tw//2) - 6
                    ry = cy + dy - th
                    rect = QRect(rx, ry, tw + 12, th + 6)
                    if not any(rect.intersects(other) for other in placed):
                        p.setPen(QPen(self.label_bd, 1))
                        p.setBrush(self.label_bg)
                        p.drawRoundedRect(rect, 6, 6)
                        p.setPen(self.text)
                        p.drawText(rect, Qt.AlignCenter, txt)
                        placed.append(rect)
                        break

            p.setPen(self.muted); p.setFont(QFont("Segoe UI", 9))
            p.drawText(QRect(x0, y0+6, x1-x0, 20), Qt.AlignHCenter | Qt.AlignTop, "runs →")

        except Exception as e:
            print("[LiveChart] paint error:", e)
            import traceback; traceback.print_exc()
        finally:
            if p.isActive():
                p.end()

# --- Bullet labeling worker -------------------------------------
class BulletLabelWorker(QObject):
    finished = pyqtSignal(dict, str)
    error    = pyqtSignal(str)

    def __init__(self, tool, data):
        super().__init__()
        self.tool = tool
        self.data = data

    def run(self):
        try:
            labels = {}
            self.tool._bullet_labels = labels
            html = self.tool._build_report_html(self.data)
            self.finished.emit(labels, html)
        except Exception as e:
            self.error.emit(str(e))
# ----------------------------------------------------------------
# --- Report build worker ----------------------------------------
class ReportBuildWorker(QObject):
    finished = pyqtSignal(dict, str)
    error    = pyqtSignal(str)

    def __init__(self, tool, data):
        super().__init__()
        self.tool = tool
        self.data = data

    def run(self):
        try:
            labels = {}
            self.tool._bullet_labels = labels
            html = self.tool._build_report_html(self.data)
            self.finished.emit(labels, html)
        except Exception as e:
            self.error.emit(str(e))

# ----------------------------------------------------------------

class TranslationTool(QMainWindow):
    """Main window for PAEM-CMT."""

    STANDARD_HEADINGS = [
        "Intended Purpose",
        "Target Audience",
        "Translator",
        "Source Owner",
        "Commissioner",
        "Terminology Adherence"
    ]

    ADDITIONAL_INSTRUCTION_TEMPLATE = """
    ### Intended Purpose / Function of the Translation:

    ### Roles and Their Expectations in the Translation Process
    - **Target Audience:**  
    - **Translator:**  
    - **Source Text Owner:**  
    - **Commissioner:**  
    """

    def __init__(self):
        """Initialize window state without starting I/O or network work."""
        super().__init__()
        self.audit_claims = False
        self.setWindowIcon(QIcon(resource_path("paemcmtlogo.ico")))
        self.source_language = ""
        self.target_language = ""
        self.translation_instructions = ""
        self._stdout_orig = sys.stdout
        self._stderr_orig = sys.stderr
        self._log_buffer = []
        self.role_weight = 0.5
        self.term_weight = 0.5
        self.model_name = "gpt-5.4-2026-03-05"
        import os
        self._seed_anchor = (os.getenv("PAEM_SEED_ANCHOR", "1") != "0")
        self.term_pairs: list[tuple[str, str]] = []
        self._context_state = {
            "purpose": "",
            "audience": "",
            "translator": "",
            "owner": "",
            "commissioner": "",
            "termpairs": [],
        }
        self._term_obligation_cache = {}
        self.term_logic_version = "obligation_v3_relaxed_matcher"
        self.eval_parse_fallback_enabled = False
        self.audit_major_drop_threshold = 1

        self.toast = QLabel("", self)
        self.toast.setObjectName("toast")
        self.toast.hide()
        self._toast_timer = QTimer(self)
        self._toast_timer.setSingleShot(True)
        self._toast_timer.timeout.connect(lambda: self.toast.hide())

        self.setWindowFlag(Qt.FramelessWindowHint, True)
        self._title_h = 30
        self._drag_pos = None
        self._build_title_bar()

        self.initUI()
        self.apply_dark_theme()
        self._install_log_capture()

    def apply_dark_theme(self):
        """Apply the dark application theme."""
        T = {
            "bg":           "#0b1016",
            "panel":        "#141b25",
            "card":         "#151b24",
            "card_border":  "#223148",
            "text":         "#e8eefc",
            "muted":        "#a9bfdc",
            "accent":       "#6ea8ff",
            "accent_hi":    "#8ebcff",
            "indigo":       "#3a4f7a",
            "indigo_hi":    "#4b69a6",
            "success_bg":   "#183124",
            "success_bd":   "#2f6d45",
            "danger_bg":    "#2a161a",
            "danger_bd":    "#a23b4c",
            "chip_bg":      "#1b2340",
            "chip_bd":      "#405a92",
            "chip_hover":   "#23305a",
            "dash":         "#3a4f7a",
            "overlay":      "rgba(17,25,37,0.72)"
        }

        self.setStyleSheet(f"""
        * {{ font-family: 'Segoe UI','Inter','Helvetica Neue',Arial,sans-serif; }}
        QWidget {{ background:{T['bg']}; color:{T['text']}; font-size:17px; }}

        QLabel#formLabel {{ color:#bcd0ff; font-weight:600; }}

        /* Cards */
        QGroupBox {{
            background:{T['card']};
            border:1px solid {T['card_border']};
            border-radius:16px;
            margin-top:12px;
            padding:14px;
            font-weight:600;
        }}
        QGroupBox::title {{
            subcontrol-origin: margin;
            left:14px; top:-4px;
            padding:0 6px;
            color:#9eb7ff;
        }}
        QGroupBox:hover {{ border-color:{T['indigo_hi']}; }}

        /* Buttons */
        QPushButton {{
            border:0;
            border-radius:18px;
            padding:12px 18px;
            font-size:17px;
            font-weight:800;
            color:{T['text']};
            background:#1a2130;
        }}
        QPushButton:hover  {{ background:#202a41; }}
        QPushButton:pressed{{ padding-top:13px; padding-bottom:11px; }}
        QPushButton:disabled {{ color:#9aa4b8; background:#141a26; border:1px solid #26344d; }}

        /* Primary CTA */
        QPushButton#primary {{
            background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {T['accent']}, stop:1 {T['accent_hi']});
            color:#051428; font-weight:900;
        }}
        QPushButton#primary:hover {{
            background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {T['accent_hi']}, stop:1 #9dc8ff);
        }}

        /* Quiet buttons */
        QPushButton#quiet, QPushButton#quietBand {{
            background:transparent;
            border:1px solid {T['indigo']};
            color:{T['muted']};
            font-weight:700;
            border-radius:14px;
            padding:8px 12px;
        }}
        QPushButton#quiet:hover, QPushButton#quietBand:hover {{
            background:#151c29; border-color:{T['indigo_hi']}; color:{T['text']};
        }}

        /* Mini corner chip */
        QPushButton#miniCorner {{
                background:rgba(18,26,40,0.65);
                border:1px solid {T['chip_bd']};
                color:{T['text']};
                font-weight:800;
                padding:6px 10px;
                border-radius:14px;
        }}
        QPushButton#miniCorner:hover {{ background:rgba(22,32,52,0.8); }}

        /* Empty-state cover */
        QPushButton#emptyCover {{
            background:{T['overlay']};
            border:2px dashed {T['dash']};
            border-radius:12px;
            font-weight:900;
        }}
        QPushButton#emptyCover:hover {{ background:rgba(20,30,48,0.80); border-color:{T['indigo_hi']}; }}

        /* API key states */
        QPushButton#recordBtn {{ background:{T['success_bg']}; border:2px solid {T['success_bd']}; color:#e8fff4; font-weight:900; }}
        QPushButton#recordBtn:hover {{ background:#1d3a26; }}
        QPushButton#dangerBtn {{ background:{T['danger_bg']}; border:1px solid {T['danger_bd']}; color:#ffdadf; font-weight:900; }}
        QPushButton#dangerBtn:hover {{ background:#34191f; }}

        /* Inputs */
        QLineEdit, QTextEdit, QPlainTextEdit, QComboBox {{
            background:#0f1522;
            border:1.5px solid {T['indigo']};
            border-radius:12px;
            padding:10px 12px;
            font-size:17px;
            selection-background-color:#2d4373;
            selection-color:#ffffff;
        }}
        QTextEdit[readOnly="true"] {{ background:#0f1522; }}

        /* Pills */
        #chip {{
            background:{T['chip_bg']};
            border:1px solid {T['chip_bd']};
            border-radius:16px;
        }}

        #chip * {{
            background:transparent;
            background-color:transparent;
            qproperty-autoFillBackground: 0;
        }}

        #chip QLabel {{
            color:#cfe2ff;
            font-weight:800;
            padding:6px 10px;
        }}

        #chip QLineEdit#chipEdit {{
            background:transparent;
            background-color:transparent;
            border:0;
            padding:8px 10px;
            color:#eaf0ff;
        }}

        #chip QLineEdit#chipEdit:focus {{
            background:transparent;
            background-color:transparent;
            border:0;
        }}

        /* Hide clear button */
        QLineEdit#chipEdit::clear-button {{
            width:0px; height:0px; padding:0; image:none;
        }}

        /* Error state outlines */
        QFrame#chip[error="true"] {{
            border: 1.5px solid #a95060;
            box-shadow: 0 0 0 1px rgba(210,120,130,0.20) inset;
        }}
        QGroupBox[error="true"] {{
            border: 1.5px solid #a95060;
        }}

        /* Logs console */
        QPlainTextEdit#console {{
            font-family:'Cascadia Mono','JetBrains Mono',Consolas,monospace;
            font-size:15px;
            background:#0a0f18;
            border:2px solid #6b4dff;
        }}

        /* Tables */
        QTableWidget {{
            gridline-color:#2b3d5c;
            background:#0e1420;
            border:1px solid #21304a;
            border-radius:12px;
            alternate-background-color:#0f1522;
        }}
        QHeaderView::section {{
            background:#151a23;
            color:#bcd0ff;
            border:0;
            padding:10px;
            font-weight:700;
        }}
        QTableWidget::item {{
            background:#0e1420;
            border-bottom:1px solid #21304a;
            padding:4px 6px;
        }}
        QTableView::item:selected {{
            background:#1b2744;
            color:#e8eefc;
        }}
        QTableView::item:hover {{
            background:#152039;
        }}

        /* Scrollers */
        QScrollBar:vertical {{ background:{T['bg']}; width:12px; margin:0; }}
        QScrollBar::handle:vertical {{ background:#2b3d5c; min-height:24px; border-radius:6px; }}
        QScrollBar::handle:vertical:hover {{ background:{T['indigo_hi']}; }}

        QSplitter::handle {{ background:{T['bg']}; }}

        /* Tooltip */
        QToolTip {{
            background-color:#141b2b; color:{T['text']};
            border:1px solid #2b3d5c; padding:6px; border-radius:8px;
        }}

        /* Small status chip */
        QLabel#chip {{
            background:#14233e; border:1px solid #2c58a0; border-radius:999px;
            padding:6px 10px; color:#cfe2ff; font-size:12px; max-width: 360px;
        }}

        /* Divider */
        QFrame#line {{ background:#233249; max-height:1px; min-height:1px; }}

        /* Toast */
        #toast {{
            background:rgba(20,27,43,0.94);
            border:1px solid #2b3d5c;
            border-radius:10px;
            padding:10px 14px;
            color:{T['text']};
            font-weight:700;
        }}

        /* Context stage chip */
        #stageChip {{
            background:#203241; border:1px solid #2b4152; color:#dbe8f5;
            font-weight:700; font-size:14px; padding:6px 12px; border-radius:10px;
            qproperty-alignment: AlignCenter;
        }}
        #stageChip[state="idle"]  {{ background:#211a33; border-color:#5f4b8b; color:#efeaff; }}
        #stageChip[state="busy"]  {{ background:#2a2147; border-color:#b896ff; color:#f2ebff; }}
        #stageChip[pulse="on"][state="busy"] {{ background:#352a5d; border-color:#c9b3ff; }}
        #stageChip[state="ok"]    {{ background:#1e2f39; border-color:#35e0a7; }}

        /* Frameless title bar */
        QFrame#titleBar {{
            background:{T['bg']};
            border-bottom:1px solid {T['card_border']};
        }}
        QPushButton#btnMin, QPushButton#btnClose {{
            border:1px solid transparent;
            border-radius:6px;
            padding:0;
            font-size:14px;
            min-width:26px;
            min-height:22px;
            background:transparent;
            color:{T['text']};
        }}
        QPushButton#btnMin:hover {{
            background:{T['success_bg']};
            border-color:{T['success_bd']};
        }}
        QPushButton#btnClose:hover {{
            background:{T['danger_bg']};
            border-color:{T['danger_bd']};
        }}
        """)

    def _build_terminology_card(self) -> QGroupBox:
        container = QWidget(self)
        v = QVBoxLayout(container); v.setContentsMargins(0,0,0,0); v.setSpacing(8)

        hdr = QWidget(self); h = QHBoxLayout(hdr); h.setContentsMargins(0,0,0,0)
        h.addStretch(1)
        btn_import = QPushButton("📥 Import", self); btn_import.setObjectName("quiet")
        btn_import.setToolTip("Import a two-column .xlsx (Source | Target)")
        btn_import.clicked.connect(self.import_term_xlsx)
        btn_add = QPushButton("➕ Add", self); btn_add.setObjectName("quiet"); btn_add.clicked.connect(self.add_term_row)
        btn_del = QPushButton("🗑️ Delete", self); btn_del.setObjectName("quiet"); btn_del.clicked.connect(self.delete_term_row)
        for b in (btn_import, btn_add, btn_del): h.addWidget(b, 0)
        v.addWidget(hdr)

        self.term_table = QTableWidget(0, 2, self)
        self.term_table.verticalHeader().setVisible(False)
        self.term_table.setHorizontalHeaderLabels(["Source", "Target"])
        self.term_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.term_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.term_table.verticalHeader().setDefaultSectionSize(44)

        v.addWidget(self.term_table, 1)
        try: QScroller.grabGesture(self.term_table.viewport(), QScroller.LeftMouseButtonGesture)
        except Exception: pass

        return self._make_card("📚 Terminology", container)

    def add_term_row(self):
        r = self.term_table.rowCount()
        self.term_table.insertRow(r)
        for c in (0,1):
            self.term_table.setItem(r, c, QTableWidgetItem(""))
        self.term_table.setRowHeight(r, 44)

    def delete_term_row(self):
        r = self.term_table.currentRow()
        if r >= 0:
            self.term_table.removeRow(r)

    def import_term_xlsx(self):
        if load_workbook is None:
            QMessageBox.warning(self, "Missing dependency", "Install openpyxl to import .xlsx:  pip install openpyxl")
            return
        fn, _ = QFileDialog.getOpenFileName(self, "Import Terminology (.xlsx)", "", "Excel Files (*.xlsx)")
        if not fn:
            return
        try:
            wb = load_workbook(fn)
            sh = wb.active
            rows = []
            for row in sh.iter_rows(values_only=True):
                if not row:
                    continue
                src = (row[0] or "").strip()
                tgt = (row[1] or "").strip() if len(row) > 1 else ""
                if src:
                    rows.append((src, tgt))
            self.term_table.setRowCount(0)
            for src, tgt in rows:
                r = self.term_table.rowCount()
                self.term_table.insertRow(r)
                self.term_table.setItem(r, 0, QTableWidgetItem(src))
                self.term_table.setItem(r, 1, QTableWidgetItem(tgt))
            self._show_toast(f"Imported {len(rows)} term(s).")
        except Exception as e:
            QMessageBox.critical(self, "Import failed", str(e))

    def _has_widget(self, name):
        w = getattr(self, name, None)
        try:
            from PyQt5 import sip
        except Exception:
            import sip
        return (w is not None) and (not sip.isdeleted(w))

    def _current_languages(self):
        src = (self.source_language or
               (self.src_edit.text().strip() if self._has_widget("src_edit") else ""))
        tgt = (self.target_language or
               (self.tgt_edit.text().strip() if self._has_widget("tgt_edit") else ""))
        mdl = (self.model_name or
               (self.model_edit.text().strip() if self._has_widget("model_edit") else ""))
        return src, tgt, mdl

    def _collect_term_pairs(self) -> list[tuple[str,str]]:
        pairs = []
        for r in range(self.term_table.rowCount()):
            src = (self.term_table.item(r,0).text() if self.term_table.item(r,0) else "").strip()
            tgt = (self.term_table.item(r,1).text() if self.term_table.item(r,1) else "").strip()
            if src:
                pairs.append((src, tgt))
        seen = set(); out = []
        for s,t in pairs:
            if s.lower() in seen: continue
            seen.add(s.lower()); out.append((s,t))
        return out

    def _build_purpose_card(self) -> QGroupBox:
        wrap = QWidget(self)
        v = QVBoxLayout(wrap); v.setContentsMargins(0,0,0,0); v.setSpacing(8)

        hdr = QWidget(self); h = QHBoxLayout(hdr); h.setContentsMargins(0,0,0,0)
        lab = QLabel("🧭 Purpose", self)
        lab.setObjectName("formLabel")
        h.addWidget(lab, 0)

        self.btn_create_purpose = QPushButton("✨ Create Purpose", self)
        self.btn_create_purpose.setObjectName("quiet")
        self.btn_create_purpose.setToolTip("Summarize the role expectations into one clear sentence.")
        self.btn_create_purpose.clicked.connect(self.create_purpose_from_roles)
        h.addStretch(1); h.addWidget(self.btn_create_purpose)
        v.addWidget(hdr)

        self.purpose_edit = ModernEditor(self)
        self.purpose_edit.setAcceptRichText(False)
        self.purpose_edit.setPlaceholderText("One-sentence aim and outcome of the translation.")
        self.purpose_edit.setMinimumHeight(96)
        v.addWidget(self.purpose_edit)
        QTimer.singleShot(0, self.purpose_edit._reveal_scrollbar)

        return self._make_card("🧾 Translation Brief", wrap)
        
    def _build_expectations_card(self) -> QGroupBox:
        wrap = QWidget(self); v = QVBoxLayout(wrap); v.setContentsMargins(0,0,0,0); v.setSpacing(10)

        def row(title: str, attr: str, placeholder: str):
            box = QWidget(self); vb = QVBoxLayout(box); vb.setContentsMargins(0,0,0,0); vb.setSpacing(6)
            lab = QLabel(title, self); lab.setObjectName("formLabel")
            vb.addWidget(lab, 0)
            ed = ModernEditor(self); ed.setAcceptRichText(False)
            ed.setPlaceholderText(placeholder)
            ed.setMinimumHeight(64)
            setattr(self, attr, ed)
            vb.addWidget(ed)
            QTimer.singleShot(0, ed._reveal_scrollbar)
            return box

        v.addWidget(row("👥 Target Audience", "audience_edit",
            "What readers expect: clarity level, tone, examples, cultural fit, accessibility."))
        v.addWidget(row("🖊️ Translator", "translator_edit",
            "What the translator’s work should prioritize: approach, register, checks, constraints."))
        v.addWidget(row("🏷️ Source Owner", "owner_edit",
            "What the owner expects to carry over: intent, tone, official names, legal/policy notes."))
        v.addWidget(row("💼 Commissioner", "commissioner_edit",
            "What the commissioner requires: success criteria, risk limits, format/delivery, approval steps."))

        row_btns = QWidget(self); h = QHBoxLayout(row_btns); h.setContentsMargins(0,0,0,0)
        h.addStretch(1)
        btn_quick = QPushButton("⚡ Quick Fill", self); btn_quick.setObjectName("quiet")
        btn_quick.clicked.connect(self.open_quick_fill_dialog)
        h.addWidget(btn_quick)
        v.addWidget(row_btns, 0, Qt.AlignRight)

        return self._make_card("🎯 Expectations", wrap)

    def _open_busy_popup(self, text: str = "Preparing report…"):
        scrim = QWidget(self)
        scrim.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        scrim.setStyleSheet("background: rgba(10, 14, 22, 0.58);")
        scrim.setGeometry(self.rect())
        scrim.show(); scrim.raise_()

        dlg = QDialog(self)
        dlg.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        dlg.setModal(False)
        dlg.resize(520, 180)
        dlg.setStyleSheet("""
            QDialog { background:#111824; border:1px solid #2b3d5c; border-radius:16px; }
            QLabel  { color:#cfe2ff; font-weight:800; }
            QProgressBar { background:#141a26; border:1px solid #2b3d5c; border-radius:8px; height:12px; }
            QProgressBar::chunk { background:#6ea8ff; border-radius:8px; }
        """)

        lay = QVBoxLayout(dlg); lay.setContentsMargins(16,16,16,16); lay.setSpacing(12)
        title = QLabel("📄 Preparing report", dlg)
        msg   = QLabel(text, dlg); msg.setWordWrap(True)
        bar   = QProgressBar(dlg); bar.setRange(0, 0)
        lay.addWidget(title); lay.addWidget(msg); lay.addWidget(bar)

        geo = self.geometry()
        dlg.move(geo.center() - dlg.rect().center())

        timer = QTimer(dlg)
        dots = ["", ".", "..", "..."]
        state = {"i": 0}
        def tick():
            state["i"] = (state["i"] + 1) % len(dots)
            msg.setText(f"{text.rstrip('. ')}{dots[state['i']]}")
        timer.timeout.connect(tick); timer.start(550)

        self._busy_scrim = scrim
        self._busy_dlg   = dlg
        self._busy_timer = timer

        print("[Report] progress dialog opened.")
        dlg.show(); dlg.raise_()

    def _update_busy_popup(self, text: str):
        dlg = getattr(self, "_busy_dlg", None)
        if not dlg: return
        try:
            lbl = dlg.layout().itemAt(1).widget()
            lbl.setText(text)
        except Exception:
            pass

    def _close_busy_popup(self):
        try:
            if getattr(self, "_busy_timer", None):
                self._busy_timer.stop()
                self._busy_timer.deleteLater()
        except Exception:
            pass
        for attr in ("_busy_dlg", "_busy_scrim"):
            w = getattr(self, attr, None)
            if w:
                try: w.close(); w.deleteLater()
                except Exception: pass
            setattr(self, attr, None)
        print("[Report] progress dialog closed.")

    def _add_shadow(self, w: QWidget, blur: int = 24, y: int = 12, a: int = 160):
        """Apply a drop shadow to a widget."""
        eff = QGraphicsDropShadowEffect(self)
        eff.setBlurRadius(blur)
        eff.setXOffset(0)
        eff.setYOffset(y)
        eff.setColor(QColor(0, 0, 0, a))
        w.setGraphicsEffect(eff)

    def _make_card(self, title: str, child: QWidget) -> QGroupBox:
        """Wrap a widget in a titled group box."""
        gb = QGroupBox(title, self)
        lay = QVBoxLayout(gb)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(8)
        lay.addWidget(child)
        self._add_shadow(gb)
        return gb

    def _divider(self):
        line = QFrame()
        line.setObjectName("line")
        line.setFrameShape(QFrame.NoFrame)
        return line

    def _safe_set_chip(self, lbl, text):
        try:
            if lbl is not None and not sip.isdeleted(lbl):
                lbl.setText(text)
        except Exception:
            pass

    def _set_progress_label(self, text: str, maxlen: int = 48):
        """
        Update status chips without overwriting the elapsed-time clock.
        """
        if getattr(self, "_elapsed_only_mode", False):
            return

        text = text or ""
        short = text if len(text) <= maxlen else (text[:maxlen - 1] + "…")

        if hasattr(self, "progress_chip") and self.progress_chip:
            try:
                self.progress_chip.setText(short)
                self.progress_chip.setToolTip(text)
            except Exception:
                pass

        if hasattr(self, "eval_chip") and self.eval_chip:
            try:
                self.eval_chip.setText(short)
                self.eval_chip.setToolTip(text)
            except Exception:
                pass

    def _show_toast(self, msg: str, ms: int = 2200):
        self.toast.setText(msg)
        self.toast.adjustSize()
        self.toast.move(self.width() - self.toast.width() - 24, 18)
        self.toast.show()
        self._toast_timer.start(ms)

    def _format_dur(self, seconds: float) -> str:
        seconds = int(max(0, seconds))
        m, s = divmod(seconds, 60)
        h, m = divmod(m, 60)
        return f"{h}h {m:02d}m {s:02d}s" if h else f"{m}m {s:02d}s"

    def _start_elapsed_clock(self):
        self._eval_start_ts = time.monotonic()
        self._elapsed_only_mode = True
        if not hasattr(self, "_elapsed_timer"):
            self._elapsed_timer = QTimer(self)
            self._elapsed_timer.timeout.connect(self._tick_elapsed_clock)
        self._elapsed_timer.start(1000)
        self._tick_elapsed_clock()

    def _tick_elapsed_clock(self):
        if hasattr(self, "_eval_start_ts"):
            secs = time.monotonic() - self._eval_start_ts
            self._safe_set_chip(self.progress_chip, f"⏱ {self._format_dur(secs)}")

    def _stop_elapsed_clock(self):
        if hasattr(self, "_elapsed_timer"):
            self._elapsed_timer.stop()

    def _update_stage_chip(self, state: str, text: str):
        if hasattr(self, "stage_chip") and self.stage_chip:
            self.stage_chip.setText(text or "")
            self.stage_chip.setToolTip(text or "")
            self.stage_chip.setProperty("state", state or "idle")
            self.stage_chip.style().unpolish(self.stage_chip)
            self.stage_chip.style().polish(self.stage_chip)

    def _start_chip_pulse(self, interval_ms: int = 520):
        if not hasattr(self, "_chip_pulse_timer"):
            self._chip_pulse_timer = QTimer(self)
            self._chip_pulse_timer.timeout.connect(self._flip_chip_pulse)
        self._chip_pulse_on = False
        self._chip_pulse_timer.start(interval_ms)

    def _stop_chip_pulse(self):
        if hasattr(self, "_chip_pulse_timer"):
            self._chip_pulse_timer.stop()
        self._set_chip_pulse(False)

    def _flip_chip_pulse(self):
        self._chip_pulse_on = not getattr(self, "_chip_pulse_on", False)
        self._set_chip_pulse(self._chip_pulse_on)

    def _set_chip_pulse(self, on: bool):
        if hasattr(self, "stage_chip") and self.stage_chip:
            self.stage_chip.setProperty("pulse", "on" if on else "off")
            self.stage_chip.style().unpolish(self.stage_chip)
            self.stage_chip.style().polish(self.stage_chip)

    def stage(self, key: str, detail: str = ""):
        em = {"PURP":"✨", "QF":"⚡", "OK":"✅", "ERR":"❌", "RDY":"🧭"}
        defaults = {"PURP":"Create Purpose…", "QF":"Quick Fill…", "OK":"Done", "ERR":"Error", "RDY":"Ready"}
        state = {"PURP":"busy", "QF":"busy", "OK":"ok", "ERR":"idle", "RDY":"idle"}.get(key, "idle")
        prefix = "" if (detail.strip().startswith(tuple(em.values()))) else em.get(key, "")
        label  = f"{prefix} {detail or defaults.get(key,'')}".strip()
        self._update_stage_chip(state, label)
        if state == "busy": self._start_chip_pulse(520)
        else:               self._stop_chip_pulse()

    def _snapshot_context_ui(self) -> dict:
        """Read current context widgets into a dict."""
        def _txt(name):
            w = getattr(self, name, None)
            return w.toPlainText().strip() if w else ""
        try:
            pairs = self._collect_term_pairs()
        except Exception:
            pairs = []
        return {
            "purpose":      _txt("purpose_edit"),
            "audience":     _txt("audience_edit"),
            "translator":   _txt("translator_edit"),
            "owner":        _txt("owner_edit"),
            "commissioner": _txt("commissioner_edit"),
            "termpairs":    pairs,
        }

    def _restore_context_ui(self, data: dict | None = None):
        """Restore saved values into current widgets."""
        data = data or getattr(self, "_context_state", {}) or {}
        def _set(name, val):
            w = getattr(self, name, None)
            if w is not None:
                w.setPlainText(val or "")
        _set("purpose_edit",      data.get("purpose", ""))
        _set("audience_edit",     data.get("audience", ""))
        _set("translator_edit",   data.get("translator", ""))
        _set("owner_edit",        data.get("owner", ""))
        _set("commissioner_edit", data.get("commissioner", ""))

        if hasattr(self, "term_table") and self.term_table is not None:
            self.term_table.blockSignals(True)
            self.term_table.setRowCount(0)
            for src, tgt in data.get("termpairs", []) or []:
                r = self.term_table.rowCount()
                self.term_table.insertRow(r)
                self.term_table.setItem(r, 0, QTableWidgetItem(src))
                self.term_table.setItem(r, 1, QTableWidgetItem(tgt))
            self.term_table.blockSignals(False)

    def _lazy_sync_context_state(self):
        """Refresh the cached context snapshot from the UI."""
        try:
            self._context_state = self._snapshot_context_ui()
        except Exception:
            pass

    def _set_chip_state(self, wrap: QWidget, **props):
        for k, v in props.items():
            wrap.setProperty(k, v)
        wrap.style().unpolish(wrap)
        wrap.style().polish(wrap)
        wrap.update()

    def _refresh_api_key_button_style(self):
        if not hasattr(self, "set_key_btn") or not self.set_key_btn:
            return
        has = bool(get_api_key())
        self.set_key_btn.setObjectName("recordBtn" if has else "dangerBtn")
        self.set_key_btn.style().unpolish(self.set_key_btn)
        self.set_key_btn.style().polish(self.set_key_btn)

    def _refresh_previews(self):
        if hasattr(self, "src_preview") and self.src_preview:
            self.src_preview.setHtml(getattr(self, "docx_content", ""))
        if hasattr(self, "tgt_preview") and self.tgt_preview:
            self.tgt_preview.setHtml(getattr(self, "_last_translation_html", ""))

        if hasattr(self, "_sync_src_overlay"): self._sync_src_overlay()
        if hasattr(self, "_sync_tgt_overlay"): self._sync_tgt_overlay()
        if hasattr(self, "_sync_change_band"): self._sync_change_band()

    def set_api_key_action(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("OpenAI API Key")
        dlg.setModal(True)
        dlg.setWindowFlag(Qt.WindowContextHelpButtonHint, False)

        lay = QVBoxLayout(dlg); lay.setSpacing(10)

        title = QLabel("Enter your OpenAI API key")
        title.setStyleSheet("font-weight:800;")
        lay.addWidget(title)

        hint = QLabel("🔒 Stored in memory only until you quit (not saved to disk).")
        hint.setWordWrap(True); hint.setStyleSheet("color:#a6c3d9;")
        lay.addWidget(hint)

        edit = QLineEdit(); edit.setEchoMode(QLineEdit.Password)
        edit.setPlaceholderText("sk-...")
        edit.setMinimumWidth(520)
        lay.addWidget(edit)

        show_cb = QCheckBox("Show key")
        show_cb.toggled.connect(lambda v: edit.setEchoMode(QLineEdit.Normal if v else QLineEdit.Password))
        lay.addWidget(show_cb)

        qa = QHBoxLayout()
        btn_paste = QPushButton("Paste");   btn_paste.clicked.connect(lambda: edit.setText(QGuiApplication.clipboard().text().strip()))
        btn_get   = QPushButton("Get a key"); btn_get.clicked.connect(lambda: QDesktopServices.openUrl(QUrl("https://platform.openai.com/account/api-keys")))
        qa.addStretch(1); qa.addWidget(btn_paste); qa.addWidget(btn_get)
        lay.addLayout(qa)

        err = QLabel(""); err.setStyleSheet("color:#ff6b6b; font-weight:700;"); err.hide()
        lay.addWidget(err)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(btns)

        def _ok():
            k = (edit.text() or "")
            k = re.sub(r"\s+", "", k).replace("\u200b", "")
            if not (k.startswith("sk-") and len(k) >= 24):
                err.setText("That doesn’t look like an OpenAI API key (should start with “sk-”).")
                err.show(); return
            set_api_key(k)
            global client
            client = new_client()
            self._refresh_api_key_button_style()
            self._show_toast("API key set for this session.")
            dlg.accept()

        btns.accepted.connect(_ok)
        btns.rejected.connect(dlg.reject)
        dlg.exec_()

    def load_source_file(self, fn: str):
        """Load a source .docx file and refresh the left preview."""
        if not fn:
            return
        self.file_path = fn
        self.docx_content = mammoth.convert_to_html(open(fn, "rb")).value
        print(Fore.CYAN + f"[FILE] source loaded: {fn}")
        self._refresh_previews()
        self._show_toast("Source uploaded.")

    def load_translation_file(self, fn: str):
        """Load a translation file and refresh the right preview."""
        if not fn:
            return
        if fn.lower().endswith(".docx"):
            html = mammoth.convert_to_html(open(fn, "rb")).value
        else:
            try:
                with open(fn, "r", encoding="utf-8") as f:
                    html = f.read()
            except UnicodeDecodeError:
                with open(fn, "rb") as f:
                    raw = f.read()
                enc = (chardet.detect(raw).get("encoding") or "utf-8").strip() or "utf-8"
                html = raw.decode(enc, errors="replace")
        self._last_translation_html = html
        self._skip_translation = True
        print(Fore.CYAN + f"[FILE] translation loaded: {fn}")
        self._refresh_previews()
        self._show_toast("Translation loaded.")

    def eventFilter(self, obj, ev):
        return super().eventFilter(obj, ev)

    def _build_drop_preview(self, which: str):
        """
        Preview card with drag-and-drop, an empty-state cover, and a corner
        change button shown on hover when content exists.
        """
        assert which in ("source", "target")
        is_source = which == "source"

        pane = QWidget(self)
        v = QVBoxLayout(pane); v.setContentsMargins(0,0,0,0); v.setSpacing(0)

        preview = ModernPreview(self)
        preview.setLineWrapMode(QTextEdit.WidgetWidth)
        preview.setTextInteractionFlags(Qt.TextSelectableByMouse | Qt.TextSelectableByKeyboard)
        v.addWidget(preview)

        if is_source: self.src_preview = preview
        else:         self.tgt_preview = preview

        drop = DropArea((".docx",) if is_source else (".docx", ".html"))
        drop.setParent(pane)
        drop.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        drop.setStyleSheet("background:transparent;")

        def on_drop(path: str):
            if is_source: self.load_source_file(path)
            else:         self.load_translation_file(path)
            QTimer.singleShot(0, self._live_clear_setup_errors)
            QTimer.singleShot(0, sync_controls)

        drop.fileDropped.connect(on_drop)

        empty_cover = QPushButton(pane)
        empty_cover.setObjectName("emptyCover")
        empty_cover.setCursor(Qt.PointingHandCursor)
        empty_cover.setFlat(True)
        empty_cover.setText("📄  Upload Source (.docx)" if is_source else "📥  Upload Translation (.docx/.html)")
        empty_cover.clicked.connect(lambda: self.open_file_dialog() if is_source else self.open_translation_dialog())

        corner = QPushButton(pane)
        corner.setObjectName("miniCorner")
        corner.setCursor(Qt.PointingHandCursor)
        corner.setText("🗁 Change Source" if is_source else "🗁 Change Translation")
        corner.clicked.connect(lambda: self.open_file_dialog() if is_source else self.open_translation_dialog())
        corner.hide()

        def _resize(_=None):
            drop.setGeometry(pane.rect())
            r = pane.rect().adjusted(8, 8, -8, -8)
            empty_cover.setGeometry(r)
            sz = corner.sizeHint()
            corner.setGeometry(pane.width() - sz.width() - 16, 12, sz.width(), sz.height())
        pane.resizeEvent = _resize
        _resize()

        def has_content() -> bool:
            return bool(getattr(self, "docx_content" if is_source else "_last_translation_html", ""))

        def sync_controls():
            loaded = has_content()
            empty_cover.setVisible(not loaded)
            if loaded:
                corner.hide()
            else:
                corner.hide()

            self._sync_change_band()

        def enterEvent(_):
            if has_content():
                corner.show()
        def leaveEvent(_):
            if has_content():
                corner.hide()

        pane.enterEvent = enterEvent
        pane.leaveEvent = leaveEvent
        sync_controls()

        if is_source: self._sync_src_overlay = sync_controls
        else:         self._sync_tgt_overlay = sync_controls

        title = "📄 Source Preview" if is_source else "📝 Target Preview"
        return self._make_card(title, pane), preview

    def _any_context_filled(self) -> bool:
        edits = [getattr(self, x, None) for x in (
            "purpose_edit","audience_edit","translator_edit","owner_edit","commissioner_edit"
        )]
        return any(e and e.toPlainText().strip() for e in edits)

    def _ask_fill_context_popup(self):
        dlg = QDialog(self); dlg.setWindowTitle("Add a little context")
        lay = QVBoxLayout(dlg)
        lab = QLabel("Please enter at least Purpose or one role’s expectation before generating notes.")
        lab.setWordWrap(True); lay.addWidget(lab)
        lay.addWidget(QDialogButtonBox(QDialogButtonBox.Ok))
        dlg.exec_()

    def _install_log_capture(self):
        """Route print() output to the in-app console and the original streams."""
        self._stdout_stream = EmittingStream(tee=self._stdout_orig)
        self._stderr_stream = EmittingStream(tee=self._stderr_orig)
        self._stdout_stream.textWritten.connect(self._append_log)
        self._stderr_stream.textWritten.connect(self._append_log)
        sys.stdout = self._stdout_stream
        sys.stderr = self._stderr_stream

    def _append_log(self, text: str):
        clean = ANSI_RE.sub("", text)
        if not hasattr(self, "console") or self.console is None:
            self._log_buffer.append(clean)
            return
        if self._log_buffer:
            self.console.moveCursor(QTextCursor.End)
            self.console.insertPlainText("".join(self._log_buffer))
            self._log_buffer.clear()
        self.console.moveCursor(QTextCursor.End)
        self.console.insertPlainText(clean)
        self.console.ensureCursorVisible()

    def _build_console_card(self) -> QGroupBox:
        wrapper = QWidget(self)
        v = QVBoxLayout(wrapper); v.setContentsMargins(0,0,0,0); v.setSpacing(8)

        bar = QWidget(self); h = QHBoxLayout(bar); h.setContentsMargins(0,0,0,0); h.setSpacing(8)
        clear_btn = QPushButton("🧹 Clear"); clear_btn.setObjectName("quiet")
        copy_btn  = QPushButton("📋 Copy");  copy_btn.setObjectName("quiet")
        h.addStretch(1); h.addWidget(clear_btn); h.addWidget(copy_btn)
        v.addWidget(bar)

        self.console = QPlainTextEdit(self)
        self.console.setObjectName("console")
        self.console.setReadOnly(True)
        self.console.setPlaceholderText("Live logs will appear here…")
        v.addWidget(self.console)

        def clear(): self.console.clear()
        def copy():
            QGuiApplication.clipboard().setText(self.console.toPlainText())
        clear_btn.clicked.connect(clear)
        copy_btn.clicked.connect(copy)

        if self._log_buffer:
            self.console.appendPlainText("".join(self._log_buffer))
            self._log_buffer.clear()

        return self._make_card("🖨 Live Logs", wrapper)

    def _init_runs_table(self):
        """Create the per-run table."""
        self.runs_table = QTableWidget(0, 9, self)
        self.runs_table.setHorizontalHeaderLabels(
            ["Run", "Status", "Score", "Tokens", "Elapsed", "Temp", "Top-p", "Average", "Stability"]
        )

        hv = self.runs_table.horizontalHeader()
        hv.setMinimumSectionSize(72)
        hv.setDefaultSectionSize(120)
        hv.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hv.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hv.setSectionResizeMode(2, QHeaderView.Stretch)
        hv.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hv.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        hv.setSectionResizeMode(5, QHeaderView.ResizeToContents)
        hv.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        hv.setSectionResizeMode(7, QHeaderView.Stretch)
        hv.setSectionResizeMode(8, QHeaderView.Stretch)

        self.runs_table.verticalHeader().setVisible(False)
        self.runs_table.setSelectionMode(QTableWidget.NoSelection)
        self.runs_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.runs_table.setAlternatingRowColors(True)
        self.runs_table.setWordWrap(True)
        self.runs_table.verticalHeader().setDefaultSectionSize(44)
        self.runs_table.setMouseTracking(True)
        self.runs_table.cellClicked.connect(self._on_runs_cell_clicked)
        self.runs_table.cellDoubleClicked.connect(self._on_runs_cell_dblclicked)
        self.runs_table.setSortingEnabled(False)
        self.runs_table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.runs_table.verticalScrollBar().setSingleStep(18)

        self._run_rows = {}
        self._last_seen_run = 0
        self._current_run_row = None
        self._pulse_phase = 0

        self._attempt_to_run = {}
        self._success_count  = 0

        self._audit_by_row = {}
        self._discarded_attempts_shown = set()
        self._run_export = {}
        self._last_live_stability_signature = None

    def _set_current_run_row(self, run_no: int):
        r = self._ensure_run_row(run_no)
        prev = getattr(self, "_current_run_row", None)
        if prev is not None and prev != r:
            for c in range(self.runs_table.columnCount()):
                it = self.runs_table.item(prev, c)
                if it: it.setBackground(QColor(0, 0, 0, 0))

        self._current_run_row = r

        timer = getattr(self, "_row_pulse_timer", None)
        if timer is None:
            self._row_pulse_timer = QTimer(self)
            self._row_pulse_timer.timeout.connect(self._pulse_current_row)
            timer = self._row_pulse_timer
        if not timer.isActive():
            timer.start(320)

        self._pulse_phase = 0

    def _pulse_current_row(self):
        r = getattr(self, "_current_run_row", None)

        if r is None or not self._has_widget("runs_table"):
            if hasattr(self, "_row_pulse_timer") and self._row_pulse_timer:
                try: self._row_pulse_timer.stop()
                except Exception: pass
            return

        try:
            if r < 0 or r >= self.runs_table.rowCount():
                return
            self._pulse_phase = (getattr(self, "_pulse_phase", 0) + 1) % 6
            alpha = 42 + self._pulse_phase * 18
            col = QColor(110, 168, 255, alpha)
            for c in range(self.runs_table.columnCount()):
                it = self.runs_table.item(r, c)
                if it:
                    it.setBackground(col)
        except Exception:
            try:
                self._row_pulse_timer.stop()
            except Exception:
                pass

    def _add_discard_row(self, items, attempt_no=None):
        """Append an unnumbered discarded row and bind popup data to it."""
        r = self.runs_table.rowCount()
        self.runs_table.insertRow(r)
        for c in range(self.runs_table.columnCount()):
            self.runs_table.setItem(r, c, QTableWidgetItem(""))

        self.runs_table.item(r, 0).setText("—")
        self.runs_table.item(r, 1).setText("discarded")
        self.runs_table.item(r, 1).setForeground(QColor("#ff9aa8"))
        for cc in range(self.runs_table.columnCount()):
            self.runs_table.item(r, cc).setToolTip("Click to review excluded audit items for this attempt")

        self._audit_by_row[r] = items or []
        self._auto_scroll_runs_bottom()

        return r

    def _highlight_mu_ci(self, run_no: int):
        """Highlight average and CI cells after they are updated."""
        r = self._run_rows.get(run_no)
        if r is None:
            return
        for col in (7, 8):
            it = self.runs_table.item(r, col)
            if not it:
                continue
            f = it.font(); f.setBold(True); it.setFont(f)
            it.setForeground(QColor("#e8eefc"))
            it.setBackground(QColor(20, 60, 110, 90))

    def _auto_scroll_runs_bottom(self):
        try:
            QTimer.singleShot(0, lambda: self.runs_table.scrollToBottom())
        except Exception:
            pass

    def _ensure_status_kept(self, run_no: int):
        r = self._run_rows.get(run_no)
        if r is None: return
        cur = (self.runs_table.item(r, 1).text() or "").lower()
        if "anchor" in cur:
            return
        if "kept" not in cur and "discard" not in cur:
            self._set_run_cell(run_no, "Status", "kept")

    def _on_runs_cell_clicked(self, row: int, col: int):
        try:
            status = (self.runs_table.item(row, 1).text() or "").lower()
        except Exception:
            return

        if "discard" not in status:
            return

        run_text = (self.runs_table.item(row, 0).text() or "").strip()
        items = None
        title = "🧯 Discarded bullets"

        if run_text.isdigit():
            run_no = int(run_text)
            items = self._audit_by_run.get(run_no, [])
            title = f"🧯 Discarded bullets — run {run_no}"
        else:
            items = self._audit_by_row.get(row, [])
            title = "🧯 Discarded bullets — attempt"

        if items:
            self._open_discard_popup(items=items, title=title)

    def _on_runs_cell_dblclicked(self, row: int, col: int):
        self._on_runs_cell_clicked(row, col)

    def _open_discard_popup(self, run_no: int = None, items: list = None, title: str = None):
        if items is None and run_no is not None:
            items = self._audit_by_run.get(int(run_no), [])
        items = items or []

        scrim = QWidget(self)
        scrim.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        scrim.setStyleSheet("background: rgba(10, 14, 22, 0.58);")
        scrim.setGeometry(self.rect())
        scrim.show(); scrim.raise_()

        dlg = QDialog(self)
        dlg.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        dlg.setModal(True)
        dlg.resize(1000, 560)

        dlg.setStyleSheet("""
            QDialog { background:#111824; border:1px solid #2b3d5c; border-radius:14px; }
            QLabel  { color:#cfe2ff; font-weight:700; }
            QTableWidget { background:#0f1522; border:1.5px solid #3b4f77; border-radius:12px; }
            QHeaderView::section { background:#151a23; color:#bcd0ff; border:0; padding:10px; font-weight:700; }
        """)

        lay = QVBoxLayout(dlg); lay.setContentsMargins(14,14,14,14); lay.setSpacing(10)
        hdr = QLabel(title or f"🧯 Discarded bullets — run {run_no or '—'}"); lay.addWidget(hdr)

        tbl = QTableWidget(len(items), 3, dlg)
        tbl.setHorizontalHeaderLabels(["Dimension", "Bullet", "Why"])
        tbl.verticalHeader().setVisible(False)
        tbl.setWordWrap(True)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.setSelectionMode(QAbstractItemView.NoSelection)
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)

        for r, it in enumerate(items):
            d = it.get("dimension","?")
            b = it.get("justification","")
            y = it.get("why","")
            for c, txt in enumerate((d, b, y)):
                cell = QTableWidgetItem(txt)
                cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
                cell.setTextAlignment(Qt.AlignLeft | Qt.AlignTop)
                tbl.setItem(r, c, cell)
        tbl.resizeRowsToContents()
        lay.addWidget(tbl, 1)

        row = QHBoxLayout(); row.addStretch(1)
        btn_copy = QPushButton("📋 Copy rows"); btn_copy.setObjectName("quiet")
        btn_close = QPushButton("Close"); btn_close.setObjectName("primary")
        def _copy():
            lines = [f"[{it.get('dimension','?')}] {it.get('justification','')} — {it.get('why','')}" for it in items]
            QGuiApplication.clipboard().setText("\n".join(lines))
            self._show_toast("Copied.")
        btn_copy.clicked.connect(_copy)
        btn_close.clicked.connect(dlg.accept)
        row.addWidget(btn_copy); row.addWidget(btn_close)
        lay.addLayout(row)

        dlg.exec_()
        scrim.deleteLater()

    def _ensure_run_row(self, run_no: int) -> int:
        """Ensure a row exists for run_no and return its index."""
        if run_no in self._run_rows:
            return self._run_rows[run_no]
        r = self.runs_table.rowCount()
        self.runs_table.insertRow(r)
        for c in range(self.runs_table.columnCount()):
            self.runs_table.setItem(r, c, QTableWidgetItem(""))
        self.runs_table.item(r, 0).setText(str(run_no))
        self._run_rows[run_no] = r
        self._auto_scroll_runs_bottom()
        return r

    def _set_run_cell(self, run_no: int, col_name: str, text: str):
        cols = {"Run":0,"Status":1,"Score":2,"tok":3,"Elapsed":4,"temp":5,"top_p":6,"μ":7,"CI±":8}
        r = self._ensure_run_row(run_no)
        c = cols[col_name]
        item = self.runs_table.item(r, c)
        item.setText(text)

        if col_name in ("Run","Score","tok","Elapsed","μ","CI±"):
            item.setTextAlignment(Qt.AlignCenter)

        if col_name == "Status":
            t = text.lower()
            existing = (self.runs_table.item(r, c).text() or "").lower()
            if "anchor" in existing and "anchor" not in t:
                return
            if "discard" in t:
                item.setForeground(QColor("#ff9aa8"))
                for cc in range(self.runs_table.columnCount()):
                    self.runs_table.item(r, cc).setToolTip("Click to review excluded audit items for this run")
            elif "anchor" in t or "kept" in t:
                item.setForeground(QColor("#9cffd0"))

    def _should_show_live_stability(self, guard: bool, basis_size: int, ci: float) -> bool:
        try:
            if guard:
                return False

            basis_size = int(basis_size or 0)
            ci = float(ci)

            if basis_size <= 0 or not math.isfinite(ci):
                return False

            sig = (basis_size, round(ci, 6))
            prev = getattr(self, "_last_live_stability_signature", None)

            if sig == prev:
                return False

            self._last_live_stability_signature = sig
            return True
        except Exception:
            return False

    def _set_run_export(self, run_no: int, **kwargs):
        if not hasattr(self, "_run_export") or not isinstance(self._run_export, dict):
            self._run_export = {}
        d = self._run_export.setdefault(int(run_no), {})
        for k, v in kwargs.items():
            if v is not None:
                d[k] = v

    def _log_event(self, text: str):
        return

    def _sync_change_band(self):
        return

    def _mark_group_error(self, w: QWidget | None, on: bool):
        if not w: return
        w.setProperty("error", bool(on))
        w.style().unpolish(w); w.style().polish(w); w.update()

    def _validate_setup_and_highlight(self, silent: bool = False) -> bool:
        """Check required fields and files; highlight missing ones."""
        src_ok = bool(self.src_edit.text().strip())
        tgt_ok = bool(self.tgt_edit.text().strip())
        mdl_ok = bool(self.model_edit.text().strip())
        src_file_ok = bool(getattr(self, "docx_content", ""))
        tgt_file_ok = bool(getattr(self, "_last_translation_html", ""))

        self._set_chip_state(self.src_pill, error=not src_ok)
        self._set_chip_state(self.tgt_pill, error=not tgt_ok)
        self._set_chip_state(self.mdl_pill, error=not mdl_ok)

        self._mark_group_error(getattr(self, "src_card", None), not src_file_ok)
        self._mark_group_error(getattr(self, "tgt_card", None), not tgt_file_ok)

        missing = []
        if not src_ok:      missing.append("Source language")
        if not tgt_ok:      missing.append("Target language")
        if not mdl_ok:      missing.append("GPT model")
        if not src_file_ok: missing.append("Source file")
        if not tgt_file_ok: missing.append("Target file")

        if missing and not silent:
            self._show_toast("Please complete: " + ", ".join(missing))
        return not missing

    def _live_clear_setup_errors(self):
        """Keep setup error highlighting in sync after the first warning."""
        if getattr(self, "_setup_warned", False):
            self._validate_setup_and_highlight(silent=True)

    def _chip_input(self, label_text: str, edit: QLineEdit, width: int = 260) -> QWidget:
        wrap = QFrame(self)
        wrap.setObjectName("chip")
        h = QHBoxLayout(wrap); h.setContentsMargins(10, 4, 10, 4); h.setSpacing(4)

        lab = QLabel(label_text, wrap)
        lab.setAutoFillBackground(False)
        h.addWidget(lab, 0)

        edit.setObjectName("chipEdit")
        edit.setAutoFillBackground(False)
        edit.setFrame(False)
        edit.setClearButtonEnabled(False)
        h.addWidget(edit, 1)

        wrap.setFixedWidth(width)

        eff = QGraphicsDropShadowEffect(self)
        eff.setBlurRadius(14)
        eff.setXOffset(0)
        eff.setYOffset(0)
        eff.setColor(QColor(110, 168, 255, 120))
        wrap.setGraphicsEffect(eff)
        eff.setEnabled(False)

        anim = QPropertyAnimation(eff, b"blurRadius", self)
        anim.setDuration(140)
        anim.setEasingCurve(QEasingCurve.OutCubic)

        class _ChipFilter(QObject):
            def __init__(self, outer, frame, effect, anim):
                super().__init__(outer)
                self.outer, self.frame, self.effect, self.anim = outer, frame, effect, anim
            def eventFilter(self, obj, ev):
                if ev.type() == QEvent.FocusIn:
                    self.effect.setEnabled(True)
                    self.anim.stop(); self.anim.setStartValue(10); self.anim.setEndValue(28); self.anim.start()
                    self.outer._set_chip_state(self.frame, active=True)
                elif ev.type() == QEvent.FocusOut:
                    self.anim.stop(); self.anim.setStartValue(self.effect.blurRadius()); self.anim.setEndValue(10); self.anim.start()
                    QTimer.singleShot(140, lambda: self.effect.setEnabled(False))
                    self.outer._set_chip_state(self.frame, active=False)
                return False

        f = _ChipFilter(self, wrap, eff, anim)
        edit.installEventFilter(f)

        def _on_text(_):
            self._set_chip_state(wrap, filled=bool(edit.text().strip()))
        edit.textChanged.connect(_on_text)
        _on_text(edit.text())

        return wrap

    def _update_create_purpose_enabled(self):
        """Enable Create Purpose only if at least one role field has text."""
        try:
            roles = [
                self.audience_edit.toPlainText().strip(),
                self.translator_edit.toPlainText().strip(),
                self.owner_edit.toPlainText().strip(),
                self.commissioner_edit.toPlainText().strip(),
            ]
        except Exception:
            roles = []
        has_any = any(bool(r) for r in roles)
        if hasattr(self, "btn_create_purpose") and self.btn_create_purpose:
            self.btn_create_purpose.setEnabled(has_any)
            self.btn_create_purpose.setToolTip(
                "Summarize the role expectations into one clear sentence."
                if has_any else "Add something to the role boxes first."
            )

    def create_purpose_from_roles(self):
        """Compose a one-sentence skopos from role expectations."""
        if not get_api_key():
            self.set_api_key_action()
            if not get_api_key():
                return

        roles = {
            "Target Audience": (self.audience_edit.toPlainText().strip() if hasattr(self,"audience_edit") else ""),
            "Translator":      (self.translator_edit.toPlainText().strip() if hasattr(self,"translator_edit") else ""),
            "Source Owner":    (self.owner_edit.toPlainText().strip() if hasattr(self,"owner_edit") else ""),
            "Commissioner":    (self.commissioner_edit.toPlainText().strip() if hasattr(self,"commissioner_edit") else "")
        }
        if not any(roles.values()):
            QMessageBox.information(self, "Add role info", "Please enter at least one role expectation.")
            return

        if self.purpose_edit.toPlainText().strip():
            if QMessageBox.question(self, "Replace purpose?",
                                    "Replace the current purpose with a new draft from roles?") != QMessageBox.Yes:
                return

        self.stage("PURP", "Create Purpose…")

        self._purp_thread = QThread(self)
        self._purp_worker = PurposeWorker(
            model=self.model_name,
            sl=self.source_language or self.src_edit.text().strip(),
            tl=self.target_language or self.tgt_edit.text().strip(),
            roles=roles
        )
        self._purp_worker.moveToThread(self._purp_thread)
        self._purp_thread.started.connect(self._purp_worker.run)
        self._purp_worker.finished.connect(self._on_purp_done)
        self._purp_worker.error.connect(self._on_purp_error)
        self._purp_worker.finished.connect(self._purp_thread.quit)
        self._purp_worker.error.connect(self._purp_thread.quit)
        self._purp_worker.finished.connect(self._purp_worker.deleteLater)
        self._purp_worker.error.connect(self._purp_worker.deleteLater)
        self._purp_thread.finished.connect(self._purp_thread.deleteLater)
        print("[PURPOSE] started…")
        self._purp_thread.start()

    def _on_purp_done(self, sentence: str):
        self.purpose_edit.setPlainText(sentence)
        self.stage("OK", "Purpose ✓")
        self._show_toast("Purpose drafted from roles.")
        print("[PURPOSE] applied.")

    def _on_purp_error(self, err: str):
        self.stage("ERR", "Purpose error")
        QMessageBox.critical(self, "Create Purpose failed", str(err))
        print("[PURPOSE] error:", err)

    def initUI(self):
        """Build the top-level layout."""
        self.setWindowTitle("PAEM-CMT")
        self.setGeometry(100, 100, 1920, 1080)
        self.stacked_widget = QWidget(self)
        self.setCentralWidget(self.stacked_widget)
        self.layout = QVBoxLayout(self.stacked_widget)
        self.layout.setContentsMargins(16, self._title_h + 8, 16, 16)
        self.layout.setSpacing(12)

        header = QWidget(self)
        h = QHBoxLayout(header); h.setContentsMargins(0,0,0,0)
        title = QLabel("✨ PAEM-CMT — Purpose-Aligned Evaluation for Customized MT")
        title.setStyleSheet("font-size:18px; font-weight:700; color:#cfe2ff;")
        self.progress_chip = QLabel("idle")
        self.progress_chip.setObjectName("chip")
        self.stage_chip = QLabel("")
        self.stage_chip.setObjectName("stageChip")
        self.stage_chip.setProperty("state", "idle")
        self.stage_chip.setProperty("pulse", "off")
        self.stage_chip.hide()
        h.addWidget(title, 1)
        h.addWidget(self.progress_chip, 0, Qt.AlignRight)
        h.addWidget(self.stage_chip,    0, Qt.AlignRight)
        self.layout.addWidget(header)

        self.landing_screen()

    def _build_title_bar(self):
        self._title_bar = QFrame(self)
        self._title_bar.setObjectName("titleBar")
        self._title_bar.setFixedHeight(self._title_h)

        bar = QHBoxLayout(self._title_bar)
        bar.setContentsMargins(0, 0, 8, 0)
        bar.setSpacing(6)
        bar.addStretch(1)

        self._btn_min = QPushButton("–", self._title_bar)
        self._btn_min.setObjectName("btnMin")
        self._btn_min.setFixedSize(26, 22)
        self._btn_min.setToolTip("Minimize")
        self._btn_min.clicked.connect(self.showMinimized)
        bar.addWidget(self._btn_min, 0, Qt.AlignRight | Qt.AlignVCenter)

        self._btn_close = QPushButton("✕", self._title_bar)
        self._btn_close.setObjectName("btnClose")
        self._btn_close.setFixedSize(26, 22)
        self._btn_close.setToolTip("Close")
        self._btn_close.clicked.connect(self.close)
        bar.addWidget(self._btn_close, 0, Qt.AlignRight | Qt.AlignVCenter)

    def _over_win_buttons(self, pos: QPoint) -> bool:
        if not hasattr(self, "_title_bar"):
            return False
        for b in (getattr(self, "_btn_min", None), getattr(self, "_btn_close", None)):
            if not b:
                continue
            gp = b.mapTo(self, QPoint(0, 0))
            if QRect(gp, b.size()).contains(pos):
                return True
        return False

    def _toggle_max_restore(self):
        if self.isMaximized():
            self.showNormal()
        else:
            self.showMaximized()

    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton and e.pos().y() <= self._title_h and not self._over_win_buttons(e.pos()):
            self._drag_pos = e.globalPos() - self.frameGeometry().topLeft()
            e.accept()
        super().mousePressEvent(e)

    def mouseMoveEvent(self, e):
        if self._drag_pos is not None and not self.isMaximized():
            self.move(e.globalPos() - self._drag_pos)
            e.accept()
        super().mouseMoveEvent(e)

    def mouseReleaseEvent(self, e):
        self._drag_pos = None
        super().mouseReleaseEvent(e)

    def mouseDoubleClickEvent(self, e):
        if e.button() == Qt.LeftButton and e.pos().y() <= self._title_h and not self._over_win_buttons(e.pos()):
            self._toggle_max_restore()
            e.accept()
        super().mouseDoubleClickEvent(e)

    def resizeEvent(self, e):
        super().resizeEvent(e)
        if hasattr(self, "_title_bar"):
            self._title_bar.setGeometry(0, 0, self.width(), self._title_h)
            self._title_bar.raise_()
        if hasattr(self, "toast") and self.toast and self.toast.isVisible():
            self.toast.move(self.width() - self.toast.width() - 24, 18)

    def landing_screen(self):
        """Setup screen."""
        self._teardown_live_screen()
        self.initUI_header_only()

        if hasattr(self, "progress_chip"):
            self.progress_chip.hide()
        if hasattr(self, "stage_chip"):
            self.stage_chip.hide()

        bar = QWidget(self)
        h = QHBoxLayout(bar); h.setContentsMargins(0,0,0,0); h.setSpacing(8)

        self.src_edit = QLineEdit(self); self.src_edit.setPlaceholderText("e.g., Turkish")
        self.src_pill = self._chip_input("🌐 Source:", self.src_edit, width=320)

        self.tgt_edit = QLineEdit(self); self.tgt_edit.setPlaceholderText("e.g., English")
        self.tgt_pill = self._chip_input("🎯 Target:", self.tgt_edit, width=320)

        self.model_edit = QLineEdit(self); self.model_edit.setText(self.model_name)
        self.model_edit.editingFinished.connect(
            lambda: setattr(self, "model_name", self.model_edit.text().strip() or self.model_name)
        )
        self.mdl_pill = self._chip_input("🤖 GPT Model:", self.model_edit, width=380)

        if getattr(self, "source_language", ""):
            self.src_edit.setText(self.source_language)
        if getattr(self, "target_language", ""):
            self.tgt_edit.setText(self.target_language)

        self.model_edit.setText(self.model_name)

        QTimer.singleShot(0, self._live_clear_setup_errors)

        self.change_src_band = QPushButton("🗁 Change Source", self)
        self.change_src_band.setObjectName("quietBand")
        self.change_src_band.clicked.connect(self.open_file_dialog)
        self.change_src_band.setVisible(False)

        self.change_tgt_band = QPushButton("🗁 Change Translation", self)
        self.change_tgt_band.setObjectName("quietBand")
        self.change_tgt_band.clicked.connect(self.open_translation_dialog)
        self.change_tgt_band.setVisible(False)

        self.set_key_btn = QPushButton("🔑 Set API Key", self)
        self.set_key_btn.setMinimumHeight(44)
        self.set_key_btn.clicked.connect(self.set_api_key_action)

        h.addWidget(self.src_pill, 0)
        h.addWidget(self.tgt_pill, 0)
        h.addWidget(self.mdl_pill, 0)
        h.addStretch(1)
        h.addWidget(self.change_src_band, 0)
        h.addWidget(self.change_tgt_band, 0)
        h.addWidget(self.set_key_btn, 0)

        self.layout.addWidget(self._make_card("🧰 Setup", bar))
        self._refresh_api_key_button_style()

        left_card,  _ = self._build_drop_preview("source")
        right_card, _ = self._build_drop_preview("target")
        self.src_card = left_card
        self.tgt_card = right_card

        self._refresh_previews()
        QTimer.singleShot(0, lambda: getattr(self, "src_preview", None) and self.src_preview._reveal_scrollbar())
        QTimer.singleShot(0, lambda: getattr(self, "tgt_preview", None) and self.tgt_preview._reveal_scrollbar())

        splitter = QSplitter(Qt.Horizontal, self)
        splitter.addWidget(left_card)
        splitter.addWidget(right_card)
        splitter.setHandleWidth(4)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)
        self.layout.addWidget(splitter, 1)
        QTimer.singleShot(0, lambda: splitter.setSizes([1, 1]))

        row = QWidget(self)
        r = QHBoxLayout(row); r.setContentsMargins(0,8,0,0); r.setSpacing(0)
        r.addStretch(1)
        next_btn = QPushButton("➡️  Next", self)
        self.next_btn = next_btn
        next_btn.setObjectName("primary"); next_btn.setMinimumHeight(46)
        next_btn.clicked.connect(self.collect_languages)
        r.addWidget(next_btn)
        self.layout.addWidget(row, 0, Qt.AlignRight)

        self._setup_warned = False
        self.src_edit.textChanged.connect(self._live_clear_setup_errors)
        self.tgt_edit.textChanged.connect(self._live_clear_setup_errors)
        self.model_edit.textChanged.connect(self._live_clear_setup_errors)

    def initUI_header_only(self):
        """Rebuild the layout while preserving the header row."""
        self.clear_layout(self.layout, keep_header=True)

    def open_translation_dialog(self):
        fn, _ = QFileDialog.getOpenFileName(self, "Open Translated File", "",
                                            "Word Files (*.docx);;HTML Files (*.html)")
        if fn:
            self.load_translation_file(fn)

    def open_file_dialog(self):
        fn, _ = QFileDialog.getOpenFileName(self, "Open .docx File", "", "Word Files (*.docx)")
        if fn:
            self.load_source_file(fn)

    def collect_languages(self):
        self.model_name = (self.model_edit.text().strip() or self.model_name)

        if not get_api_key():
            self.set_api_key_action()
            if not get_api_key():
                return

        ok = self._validate_setup_and_highlight(silent=False)
        if not ok:
            self._setup_warned = True
            return

        self.source_language = self.src_edit.text().strip()
        self.target_language = self.tgt_edit.text().strip()
        self.context_instruction_screen()

    def context_instruction_screen(self):
        self._teardown_live_screen()
        self.initUI_header_only()
        if hasattr(self, "stage_chip"):
            self.stage_chip.show()
            self.stage("RDY", "Ready")
        if hasattr(self, "progress_chip"):
            self.progress_chip.hide()

        left_split = QSplitter(Qt.Vertical, self)
        expect_card  = self._build_expectations_card()
        purpose_card = self._build_purpose_card()

        left_split.addWidget(expect_card)

        padder = QWidget(self)
        pad_l  = QVBoxLayout(padder)
        pad_l.setContentsMargins(0,16,0,0)
        pad_l.setSpacing(0)
        pad_l.addWidget(purpose_card)

        left_split.addWidget(padder)

        left_split.setHandleWidth(4)
        left_split.setStretchFactor(0, 1)
        left_split.setStretchFactor(1, 0)

        for name in ("audience_edit", "translator_edit", "owner_edit", "commissioner_edit"):
            ed = getattr(self, name, None)
            if ed:
                ed.textChanged.connect(self._update_create_purpose_enabled)

        QTimer.singleShot(0, self._update_create_purpose_enabled)

        term_card = self._build_terminology_card()

        main_split = QSplitter(Qt.Horizontal, self)
        main_split.addWidget(left_split)
        main_split.addWidget(term_card)
        main_split.setHandleWidth(4)
        main_split.setStretchFactor(0, 1)
        main_split.setStretchFactor(1, 1)
        self.layout.addWidget(main_split, 1)

        self._restore_context_ui()

        for name in ("audience_edit", "translator_edit", "owner_edit", "commissioner_edit", "purpose_edit"):
            ed = getattr(self, name, None)
            if ed:
                ed.textChanged.connect(self._lazy_sync_context_state)

        if hasattr(self, "term_table") and self.term_table:
            self.term_table.itemChanged.connect(lambda *_: self._lazy_sync_context_state())

        QTimer.singleShot(0, self._update_create_purpose_enabled)

        def _size_once():
            h = left_split.size().height() or self.height() or 800
            left_split.setSizes([max(360, h - 220), 220])
            w = main_split.size().width() or self.width() or 1200
            main_split.setSizes([int(w * 0.58), int(w * 0.42)])
        QTimer.singleShot(0, _size_once)

        bar = QWidget(self); bb = QHBoxLayout(bar); bb.setContentsMargins(0,0,0,0); bb.setSpacing(8)
        back_btn   = QPushButton("⬅️ Back", self); back_btn.clicked.connect(self._back_from_context)
        import_btn = QPushButton("📥 Import Instructions", self); import_btn.clicked.connect(self.import_instructions_json)
        export_btn = QPushButton("📤 Export Instructions", self); export_btn.clicked.connect(self.export_instructions_json)
        start_btn  = QPushButton("🚀 Start Evaluation", self); start_btn.setObjectName("primary"); start_btn.clicked.connect(self.collect_and_evaluate)
        bb.addWidget(back_btn, 0, Qt.AlignLeft); bb.addStretch(1)
        for b in (import_btn, export_btn, start_btn): bb.addWidget(b, 0)
        self.layout.addWidget(bar)   

    def _back_from_context(self):
        self._lazy_sync_context_state()
        self.landing_screen()

    def sync_weights(self):
        """Normalize the two spinboxes into self.role_weight and self.term_weight."""
        r = self.role_spin.value()
        t = self.term_spin.value()
        total = r + t or 1
        self.role_weight = r / total
        self.term_weight = t / total

    def prepare_roles(self):
        self.open_quick_fill_dialog()

    def auto_generate_brief(self):
        self.open_quick_fill_dialog()

    def auto_generate_terminology(self):
        QMessageBox.information(self, "Disabled",
                                "Terminology auto-generation was removed. Use Import to load term pairs.")

    def open_quick_fill_dialog(self):
        scrim = QWidget(self)
        scrim.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        scrim.setStyleSheet("background: rgba(10, 14, 22, 0.58);")
        scrim.setGeometry(self.rect())
        scrim.show()
        scrim.raise_()

        dlg = QDialog(self)
        dlg.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        dlg.setModal(True)
        dlg.resize(640, 420)

        dlg.setStyleSheet("""
            QDialog {
                background:#111824;
                border:1px solid #2b3d5c;
                border-radius:14px;
            }
            QLabel { color:#cfe2ff; font-weight:600; }
            QPlainTextEdit {
                background:#0f1522; border:1.5px solid #3b4f77; border-radius:12px;
                font-size:16px; padding:12px;
            }
        """)

        lay = QVBoxLayout(dlg); lay.setSpacing(12); lay.setContentsMargins(14,14,14,14)

        msg = QLabel("Just list a few keywords for each role’s expectations.\nNo full sentences needed, PAEM-CMT will expand them.")
        msg.setWordWrap(True)
        lay.addWidget(msg)

        edit = QPlainTextEdit(self)
        edit.setPlaceholderText("")
        edit.setMinimumHeight(260)
        lay.addWidget(edit, 1)

        btns = QDialogButtonBox(QDialogButtonBox.Cancel)
        send = QPushButton("Send", dlg); send.setObjectName("primary")
        btns.addButton(send, QDialogButtonBox.AcceptRole)
        lay.addWidget(btns)

        def _go():
            text = edit.toPlainText().strip()
            if not text:
                QMessageBox.information(self, "Add a few words", "Please enter a few keywords first.")
                return
            dlg.accept()
            self._start_quick_fill(text)

        send.clicked.connect(_go)
        btns.rejected.connect(dlg.reject)
        dlg.exec_()

        scrim.deleteLater()

    def _start_quick_fill(self, notes: str):
        if not get_api_key():
            self.set_api_key_action()
            if not get_api_key():
                return
        self.stage("QF", "Quick Fill…")

        self._qf_thread = QThread(self)
        self._qf_worker = QuickFillWorker(
            model=self.model_name,
            sl=self.source_language or self.src_edit.text().strip(),
            tl=self.target_language or self.tgt_edit.text().strip(),
            notes=notes
        )
        self._qf_worker.moveToThread(self._qf_thread)
        self._qf_thread.started.connect(self._qf_worker.run)
        self._qf_worker.finished.connect(self._on_qf_done)
        self._qf_worker.error.connect(self._on_qf_error)
        self._qf_worker.finished.connect(self._qf_thread.quit)
        self._qf_worker.error.connect(self._qf_thread.quit)
        self._qf_worker.finished.connect(self._qf_worker.deleteLater)
        self._qf_worker.error.connect(self._qf_worker.deleteLater)
        self._qf_thread.finished.connect(self._qf_thread.deleteLater)
        print("[QF] started…")
        self._qf_thread.start()

    def _on_qf_done(self, data: dict):
        self.audience_edit.setPlainText(data.get("target_audience",""))
        self.translator_edit.setPlainText(data.get("translator",""))
        self.owner_edit.setPlainText(data.get("source_owner",""))
        self.commissioner_edit.setPlainText(data.get("commissioner",""))
        self._show_toast("Quick Fill applied.")
        self.stage("OK", "Quick Fill ✓")
        print("[QF] applied to role fields.")

    def _on_qf_error(self, err: str):
        self.stage("ERR", "Quick Fill error")
        QMessageBox.critical(self, "Quick Fill failed", err)
        print("[QF] error:", err)

    def export_instructions_json(self):
        src, tgt, mdl = self._current_languages()
        data = {
            "languages": {"source": src, "target": tgt, "model": mdl},
            "brief": {
                "purpose":         self.purpose_edit.toPlainText().strip(),
                "target_audience": self.audience_edit.toPlainText().strip(),
                "translator":      self.translator_edit.toPlainText().strip(),
                "source_owner":    self.owner_edit.toPlainText().strip(),
                "commissioner":    self.commissioner_edit.toPlainText().strip()
            },
            "terminology": {"pairs": self._collect_term_pairs()}
        }
        fn, _ = QFileDialog.getSaveFileName(self, "Export Instructions (JSON)", "", "JSON (*.json)")
        if not fn:
            return
        try:
            with open(fn, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self._show_toast("Instructions exported.")
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def import_instructions_json(self):
        fn, _ = QFileDialog.getOpenFileName(self, "Import Instructions (JSON)", "", "JSON (*.json)")
        if not fn:
            return
        try:
            with open(fn, "r", encoding="utf-8") as f:
                data = json.load(f)

            langs = data.get("languages", {})
            self.source_language = langs.get("source", "") or ""
            self.target_language = langs.get("target", "") or ""
            if langs.get("model"):
                self.model_name = langs["model"].strip() or self.model_name

            if self._has_widget("src_edit"):   self.src_edit.setText(self.source_language)
            if self._has_widget("tgt_edit"):   self.tgt_edit.setText(self.target_language)
            if self._has_widget("model_edit"): self.model_edit.setText(self.model_name)

            brief = data.get("brief", {})
            self.purpose_edit.setPlainText(brief.get("purpose", ""))
            self.audience_edit.setPlainText(brief.get("target_audience", ""))
            self.translator_edit.setPlainText(brief.get("translator", ""))
            self.owner_edit.setPlainText(brief.get("source_owner", ""))
            self.commissioner_edit.setPlainText(brief.get("commissioner", ""))

            term = data.get("terminology", {})
            pairs = term.get("pairs", []) or []
            self.term_table.blockSignals(True)
            self.term_table.setRowCount(0)
            for src, tgt in pairs:
                r = self.term_table.rowCount()
                self.term_table.insertRow(r)
                self.term_table.setItem(r, 0, QTableWidgetItem(src))
                self.term_table.setItem(r, 1, QTableWidgetItem(tgt))
            self.term_table.blockSignals(False)

            self._lazy_sync_context_state()
            self._show_toast("Instructions imported.")
        except Exception as e:
            QMessageBox.critical(self, "Import failed", str(e))

    def collect_and_evaluate(self):
        if not getattr(self, "_last_translation_html", ""):
            QMessageBox.warning(self, "Missing translation",
                                "Please upload a translation file on the Setup screen (right preview) before evaluating.")
            return

        src, tgt, mdl = self._current_languages()
        self.source_language = src
        self.target_language = tgt
        self.model_name      = mdl

        self._lazy_sync_context_state()
        self._term_obligation_cache = {}
        self.translation_context = {
            "purpose":      self.purpose_edit.toPlainText().strip(),
            "audience":     self.audience_edit.toPlainText().strip(),
            "translator":   self.translator_edit.toPlainText().strip(),
            "owner":        self.owner_edit.toPlainText().strip(),
            "commissioner": self.commissioner_edit.toPlainText().strip(),
            "termpairs":    self._collect_term_pairs(),
        }

        self.perform_evaluation()

    def perform_evaluation(self, max_runs=40, min_runs=5, threshold=0.05):
        self.evaluation_live_screen(max_runs)
        self._start_elapsed_clock()
        if hasattr(self, "eval_btn"):
            self.eval_btn.setEnabled(False)
        self._safe_set_chip(self.progress_chip, "starting…")
        if hasattr(self, "eval_chip"):
            self._safe_set_chip(self.eval_chip, "starting…")

        if hasattr(self, "live_chart") and self.live_chart:
            self.live_chart.scores.clear()
            self.live_chart.mu.clear()
            self.live_chart.ci.clear()
            self.live_chart._seen_keys.clear()
            if hasattr(self.live_chart, "_seen_mu_keys"):
                self.live_chart._seen_mu_keys.clear()
            if hasattr(self.live_chart, "_last_appended_val"):
                self.live_chart._last_appended_val = None
                self.live_chart._last_append_ts = 0.0
            self.live_chart.update()

        self._eval_thread = QThread(self)
        self._eval_worker = EvalWorker(self, max_runs, min_runs, threshold)
        self._eval_worker.moveToThread(self._eval_thread)

        self._eval_thread.started.connect(self._eval_worker.run)
        self._eval_worker.progress.connect(self._on_eval_progress)
        self._eval_worker.error.connect(self._on_eval_error)
        self._eval_worker.error.connect(self._eval_thread.quit)
        self._eval_worker.error.connect(self._eval_worker.deleteLater)
        self._eval_worker.finished.connect(self._on_eval_finished)

        self._eval_worker.finished.connect(self._eval_thread.quit)
        self._eval_worker.finished.connect(self._eval_worker.deleteLater)
        self._eval_thread.finished.connect(self._eval_thread.deleteLater)

        self._eval_thread.start()

        self._startup_popup_timer = QTimer(self)
        self._startup_popup_timer.setSingleShot(True)
        self._startup_popup_timer.timeout.connect(
            lambda: self._open_busy_popup("Starting evaluation…")
        )
        self._startup_popup_timer.start(250)
        QApplication.processEvents()

    def _on_eval_progress(self, msg: str):
        import re, json
        if msg.startswith(("SEED_RUN:", "SEED_ANCHOR:", "RUN_HASH:", "ANCHOR_HASH:")):
            try:
                s = msg.strip()
                if hasattr(self, "console") and self.console:
                    self.console.append(s)
                if hasattr(self, "progress_chip"):
                    self._safe_set_chip(self.progress_chip, s.split(":",1)[0].lower())
            except Exception:
                pass
            return

        if not getattr(self, "_first_result_seen", False):
            if re.search(r"\brun\s+\d+/\d+\b", msg, re.I):
                self._first_result_seen = True
                try:
                    if hasattr(self, "waiting_hint"): self.waiting_hint.hide()
                except Exception:
                    pass
                try:
                    if hasattr(self, "_startup_popup_timer") and self._startup_popup_timer.isActive():
                        self._startup_popup_timer.stop()
                except Exception:
                    pass
                self._close_busy_popup()

        if msg.startswith("AUDIT:"):
            try:
                import unicodedata
                from difflib import SequenceMatcher

                payload     = json.loads(msg[len("AUDIT:"):].strip())
                attempt_no  = int(payload.get("run", 0))
                items       = payload.get("ungrounded", []) or []
                is_discard  = bool(payload.get("discarded"))
                is_anchor   = bool(payload.get("anchor"))

                def _norm_txt(x: str) -> str:
                    import html as _html
                    x = _html.unescape(x or "").replace("\u00A0", " ")
                    x = re.sub(r"(?i)<br\s*/?>", " ", x)
                    x = re.sub(r"(?s)<[^>]+>", " ", x)
                    x = unicodedata.normalize("NFKD", x)
                    x = "".join(ch for ch in x if not unicodedata.combining(ch))
                    x = re.sub(r"\s+", " ", x).strip().lower()
                    return x

                def _minor_quote_mismatch(it: dict) -> bool:
                    bullet = str(it.get("bullet", ""))
                    why    = str(it.get("why", ""))
                    exp    = it.get("exp") or None

                    m1 = re.search(r"[\"“”'’](.*?)[\"“”'’]", bullet)
                    if not m1:
                        return False
                    got = _norm_txt(m1.group(1))
                    if not got:
                        return False

                    if not exp:
                        m2 = re.search(r"Quoted text not found.*?:\s*[\"“”'’](.*?)[\"“”'’]", why)
                        if not m2:
                            return False
                        exp = m2.group(1)

                    exp = _norm_txt(exp)
                    if not exp:
                        return False

                    got = re.sub(r"(?:\u2026|\.{3})$", "", got).strip()
                    exp = re.sub(r"(?:\u2026|\.{3})$", "", exp).strip()

                    if got.startswith(exp) or exp.startswith(got):
                        return True

                    r = SequenceMatcher(None, got, exp).ratio()
                    if r >= 0.92:
                        return True

                    k = min(len(got), len(exp), 18)
                    return k >= 10 and got[:k] == exp[:k]

                MAJOR = ({"unfound_snippet"} if not self.audit_claims
                         else {"unfound_snippet", "unsupported_claim", "instruction_conflict"})

                fixed_quote_issues = []

                def _is_gating(it: dict) -> bool:
                    why = str(it.get("why", ""))
                    if why.startswith("Quoted text not found"):
                        if _minor_quote_mismatch(it):
                            fixed_quote_issues.append(it)
                            return False
                        return True
                    return it.get("code") in MAJOR

                gating = [it for it in items if _is_gating(it)]
                effective_discard = is_discard and not fixed_quote_issues

                if is_anchor and effective_discard:
                    self._handle_anchor_discard(gating or items, attempt_no)

                if attempt_no in getattr(self, "_attempt_to_run", {}):
                    disp = self._attempt_to_run[attempt_no]
                    if gating:
                        self._audit_by_run.setdefault(disp, []).extend(gating)

                if fixed_quote_issues:
                    try:
                        self._add_discard_row([{
                            "dimension": "Auto-fix",
                            "bullet": "Minor quote typo auto-fixed; continuing.",
                            "why": "non-gating: fuzzy match ≥0.96 or ≤2 edits"
                        }], attempt_no)
                    except Exception:
                        pass

                if effective_discard and attempt_no not in self._discarded_attempts_shown:
                    self._add_discard_row(gating or items, attempt_no)
                    self._discarded_attempts_shown.add(attempt_no)

            except Exception as e:
                print("[UI] AUDIT progress handling error:", e)
            return

        if ("[skip] dropped run" in msg) or ("discarded run (anchor-band)" in msg) or ("discarded run (early outlier)" in msg):
            s = msg or ""
            m = re.search(r"\[skip\]\s*dropped(?: outlier)? run\s+(\d+).*?(anchor-band|outlier).*?\[([0-9.]+)\s*,\s*([0-9.]+)\].*?score\s*=\s*([0-9.]+)", s, re.I)
            attempt_no = None; reason = None; lo = hi = sc = None
            if m:
                attempt_no = int(m.group(1))
                reason = m.group(2).lower()
                lo = float(m.group(3)); hi = float(m.group(4)); sc = float(m.group(5))
                why = f"score {sc:.2f} ∉ [{lo:.2f}, {hi:.2f}]"
            else:
                reason = "anchor-band" if "anchor-band" in s.lower() else ("outlier" if "outlier" in s.lower() else "filter")
                why = s.strip()
            r = self._add_discard_row([{"dimension":"Filter","bullet":f"Discarded by {reason}","why":why}], attempt_no)
            try: self.runs_table.item(r, 1).setText(f"discarded ({reason})")
            except Exception: pass
            return

        if msg.startswith("RUN:"):
            try:
                payload = json.loads(msg[4:].strip())
                attempt = int(payload.get("attempt") or payload.get("run") or 0)
                total   = int(payload.get("total") or 0) or None
                elapsed = payload.get("elapsed")
                score   = payload.get("score")

                if not isinstance(getattr(self, "_attempt_to_run", None), dict):
                    self._attempt_to_run = {}
                if attempt not in self._attempt_to_run:
                    self._success_count = getattr(self, "_success_count", 0) + 1
                    self._attempt_to_run[attempt] = self._success_count
                    if hasattr(self, "eval_progress") and self.eval_progress and total:
                        try: self.eval_progress.setMaximum(total)
                        except Exception: pass

                run_no = self._attempt_to_run[attempt]
                self._last_seen_run = run_no

                if hasattr(self, "waiting_hint") and self.waiting_hint:
                    try: self.waiting_hint.hide()
                    except Exception: pass

                self._ensure_run_row(run_no)
                self._set_current_run_row(run_no)
                self._ensure_status_kept(run_no)

                if run_no == 1:
                    self._set_run_cell(run_no, "Status", "anchor")

                if elapsed is not None:
                    try: self._set_run_cell(run_no, "Elapsed", f"{float(elapsed):.2f}s")
                    except Exception: pass

                t = payload.get("temp"); p = payload.get("top_p")
                if t is not None:
                    try: self._set_run_cell(run_no, "temp", f"{float(t):.3f}")
                    except Exception: pass
                if p is not None:
                    try: self._set_run_cell(run_no, "top_p", f"{float(p):.3f}")
                    except Exception: pass

                tok_val = ((payload.get("usage") or {}).get("total_tokens")
                           if "usage" in payload else payload.get("tok"))
                if tok_val is not None:
                    self._set_run_cell(run_no, "tok", str(int(tok_val)))

                if score is not None:
                    try:
                        fscore = float(score)
                        self._set_run_cell(run_no, "Score", f"{fscore:.2f}")
                        self._set_run_export(run_no, score=fscore)
                        if hasattr(self, "live_chart") and self.live_chart:
                            self.live_chart.push_score(fscore, key=f"run:{run_no}")
                    except Exception:
                        pass

                if hasattr(self, "eval_progress") and self.eval_progress:
                    try: self.eval_progress.setValue(getattr(self, "_success_count", 0))
                    except Exception: pass

            except Exception as e:
                print("[UI] RUN handler error:", e)
            return

        if msg.startswith("ROLLING:"):
            try:
                payload = json.loads(msg[len("ROLLING:"):].strip())
                mu = float(payload.get("mu", 0.0))
                ci = float(payload.get("ci", 0.0))
                guard = bool(payload.get("guard", False))
                basis_size = int(payload.get("m", 0) or 0)
                show_stability = self._should_show_live_stability(guard, basis_size, ci)

                rn = None
                if "run" in payload:
                    rn = int(payload["run"])
                elif "attempt" in payload:
                    att = int(payload.get("attempt") or 0)
                    rn = (self._attempt_to_run.get(att) if att > 0 else None)
                if not rn:
                    rn = getattr(self, "_last_seen_run", None)

                if rn:
                    self._set_run_cell(rn, "μ", f"{mu:.2f}")
                    self._set_run_cell(rn, "CI±", (f"{ci:.4f}" if show_stability else "—"))
                    self._set_run_export(
                        rn,
                        mu=mu,
                        ci=ci,
                        show_ci=bool(show_stability),
                        guard=bool(guard),
                        basis_m=int(basis_size),
                    )

                    if show_stability:
                        self._highlight_mu_ci(rn)

                if hasattr(self, "live_chart") and self.live_chart:
                    try:
                        self.live_chart.set_threshold(getattr(self, "_ci_target", getattr(self, "ci_target", 0.05)))
                        self.live_chart.push_mu_ci(mu, ci, key=f"r{rn}-mu:{mu:.4f}:{ci:.4f}")
                    except Exception:
                        pass
            except Exception as e:
                print("[UI] ROLLING handler error:", e)
            return

        if msg.startswith("DISCARD:"):
            try:
                payload = json.loads(msg.split(":",1)[1])
                rn = int(payload.get("run", 0))
                items = [{
                    "dimension":     str(payload.get("dimension", "Filter")),
                    "justification": str(payload.get("justification", "")),
                    "why":           str(payload.get("why", "")),
                }]
                self._add_discard_row(items, rn)
                return
            except Exception as e:
                print("[UI] DISCARD handler error:", e)

        try:
            s = msg or ""

            m = re.search(r"\brun\s+(\d+)\s*/\s*(\d+)\b", s, flags=re.I)
            if m:
                attempt, tot = int(m.group(1)), int(m.group(2))
                if hasattr(self, "eval_progress") and self.eval_progress:
                    try:
                        self.eval_progress.setMaximum(tot)
                    except Exception:
                        pass

                if not isinstance(getattr(self, "_attempt_to_run", None), dict):
                    self._attempt_to_run = {}
                if attempt not in self._attempt_to_run:
                    self._success_count = getattr(self, "_success_count", 0) + 1
                    self._attempt_to_run[attempt] = self._success_count

                run_no = self._attempt_to_run[attempt]

                if hasattr(self, "waiting_hint") and self.waiting_hint:
                    try:
                        self.waiting_hint.hide()
                    except Exception:
                        pass

                self._ensure_run_row(run_no)
                self._set_current_run_row(run_no)
                self._auto_scroll_runs_bottom()

                try:
                    self._set_run_cell(run_no, "Status", "running")
                    self._ensure_status_kept(run_no)
                except Exception:
                    pass

                if run_no == 1:
                    self._set_run_cell(run_no, "Status", "anchor")

                self._last_seen_run = run_no
                if hasattr(self, "eval_progress") and self.eval_progress:
                    try:
                        self.eval_progress.setValue(getattr(self, "_success_count", 0))
                    except Exception:
                        pass

            if getattr(self, "_last_seen_run", 0):
                rn = self._last_seen_run

                m = re.search(r"\|\s*([0-9]+(?:[.,][0-9]+)?)s\s*\(avg", s, flags=re.I)
                if m:
                    self._set_run_cell(rn, "Elapsed", m.group(1).replace(",", ".") + "s")
                    self._ensure_status_kept(rn)

                m = re.search(r"\bused\s+([\d,]+)\s*tok", s, flags=re.I)
                if m:
                    self._set_run_cell(rn, "tok", m.group(1).replace(",", ""))
                    self._ensure_status_kept(rn)

                m = re.search(r"\bscore[:=]?\s*([0-9]+(?:[.,][0-9]+)?)\b", s, flags=re.I)
                if m:
                    raw = m.group(1).replace(",", ".")
                    self._set_run_cell(rn, "Score", raw)
                    try:
                        val = float(raw)
                    except Exception:
                        val = None

                    self._ensure_status_kept(rn)
                    if hasattr(self, "waiting_hint") and self.waiting_hint:
                        try:
                            self.waiting_hint.hide()
                        except Exception:
                            pass
                    if val is not None and hasattr(self, "live_chart") and self.live_chart:
                        try:
                            self.live_chart.push_score(val, key=f"run:{rn}")
                        except Exception:
                            pass

        except Exception:
            pass

    def _on_eval_error(self, err: str):
        self._stop_elapsed_clock()
        self._close_busy_popup()
        QMessageBox.critical(self, "Evaluation failed", err)
        if hasattr(self, "eval_btn"):
            self.eval_btn.setEnabled(True)
        self._safe_set_chip(self.progress_chip, "error")
        if hasattr(self, "eval_chip"):
            self._safe_set_chip(self.eval_chip, "error")

    def _show_final_summary_card(self, data: dict):
        n_kept = int(data.get("checks", data.get("n_kept", data.get("n_runs", 0))))
        n_runs = n_kept
        n_attempts = int(data.get("n_attempts", n_kept))

        best_est = float(data.get("best_estimate", data.get("overall", 0.0)))
        exp_low = float(data.get("expected_range_low", best_est - float(data.get("ci", 0.0))))
        exp_high = float(data.get("expected_range_high", best_est + float(data.get("ci", 0.0))))
        exp_range_label = str(data.get("expected_range_label", f"{exp_low:.2f}–{exp_high:.2f}"))

        from PyQt5.QtCore import QEasingCurve

        class MetricTile(QFrame):
            def __init__(self, parent, title: str, value_html: str, *,
                         accent="#6ea8ff", subchip: str | None = None, tooltip: str = ""):
                super().__init__(parent)
                self.setObjectName("metricTile")
                self.setCursor(Qt.PointingHandCursor)
                self._accent = QColor(accent)
                self._accent_hi = QColor(min(self._accent.red()+35,255),
                                         min(self._accent.green()+35,255),
                                         min(self._accent.blue()+35,255))
                self._build(title, value_html, subchip, tooltip)
                self._install_hover_anim()

            def _build(self, title, value_html, subchip, tooltip):
                self.setToolTip(tooltip)
                self.setStyleSheet(f"""
                    QFrame#metricTile {{
                        background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
                                   stop:0 #1a2233, stop:1 #111827);
                        border:1.5px solid rgba(90,120,170,0.35);
                        border-radius:20px;
                    }}
                    QFrame#metricTile:hover {{
                        border:1.5px solid rgba(110,168,255,0.75);
                        background: qlineargradient(x1:0,y1:0, x2:1,y2:1,
                                   stop:0 #1b2942, stop:1 #121a27);
                    }}
                """)
                v = QVBoxLayout(self); v.setContentsMargins(16,16,16,16); v.setSpacing(6)
                top = QHBoxLayout(); top.setSpacing(8)
                lab = QLabel(title, self); lab.setStyleSheet("color:#9eb7ff; font-weight:800;")
                top.addWidget(lab); top.addStretch(1)
                v.addLayout(top)

                val = QLabel(value_html, self)
                val.setStyleSheet("font-size:30px; font-weight:900; color:#e8eefc;")
                v.addWidget(val)
                v.addStretch(1)

                if subchip:
                    chip = QLabel(subchip, self)
                    chip.setObjectName("chip")
                    chip.setStyleSheet("QLabel#chip{ padding:6px 10px; }")
                    v.addWidget(chip, 0, Qt.AlignLeft)

                eff = QGraphicsDropShadowEffect(self)
                eff.setBlurRadius(18); eff.setXOffset(0); eff.setYOffset(14)
                eff.setColor(QColor(0,0,0,140))
                self.setGraphicsEffect(eff)
                self._shadow = eff

            def _install_hover_anim(self):
                self._anim = QPropertyAnimation(self._shadow, b"blurRadius", self)
                self._anim.setDuration(180)
                self._anim.setEasingCurve(QEasingCurve.OutCubic)

            def enterEvent(self, e):
                self._anim.stop(); self._anim.setStartValue(self._shadow.blurRadius()); self._anim.setEndValue(34); self._anim.start()
                super().enterEvent(e)

            def leaveEvent(self, e):
                self._anim.stop(); self._anim.setStartValue(self._shadow.blurRadius()); self._anim.setEndValue(18); self._anim.start()
                super().leaveEvent(e)

        panel = QWidget(self); hl = QHBoxLayout(panel); hl.setContentsMargins(0,0,0,0); hl.setSpacing(12)

        kpi1 = MetricTile(
            self,
            "Best estimate",
            f"<span style='color:#cfe2ff'>{best_est:.2f}</span>",
            accent="#6ea8ff",
            tooltip="Main result from repeated checks."
        )

        kpi2 = MetricTile(
            self,
            "Expected range",
            f"{exp_range_label}",
            accent="#35e0a7",
            tooltip="Where repeated checks usually landed under the same setup."
        )

        kpi3 = MetricTile(
            self,
            "Checks",
            f"{n_kept}",
            accent="#b08cff",
            tooltip="Number of kept repeated checks."
        )

        hl.addWidget(kpi1, 1)
        hl.addWidget(kpi2, 1)
        hl.addWidget(kpi3, 1)

        box = QGroupBox("🏁 Final Result", self)
        vb  = QVBoxLayout(box); vb.setContentsMargins(16,16,16,16); vb.setSpacing(12)

        banner = QFrame(self); bl = QHBoxLayout(banner); bl.setContentsMargins(12,10,12,10); bl.setSpacing(10)
        banner.setStyleSheet("""
            QFrame { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                     stop:0 #152133, stop:1 #101826);
                     border:1px solid #2b3d5c; border-left:4px solid #35e0a7; border-radius:12px; }
        """)
        btxt = QLabel(f"✓ Finished after {n_attempts} attempts ({n_kept} kept)", self)
        btxt.setStyleSheet("font-size:20px; font-weight:900; color:#cfe2ff;")
        bl.addWidget(btxt); bl.addStretch(1)

        vb.addWidget(banner)
        vb.addWidget(panel)

        pill = QWidget(self); ph = QHBoxLayout(pill); ph.setContentsMargins(0,0,0,0); ph.setSpacing(8)
        rchip = QLabel(f"kept {n_kept} • attempts {n_attempts}", self)
        rchip.setObjectName("chip")
        rchip.setStyleSheet("QLabel#chip{ padding:6px 10px; }")
        ph.addWidget(rchip, 0)
        ph.addStretch(1)
        vb.addWidget(pill)

        row = QWidget(self); r = QHBoxLayout(row); r.setContentsMargins(0,0,0,0); r.setSpacing(8)

        btn_export_live = QPushButton("📦 Export Stats (XLSX)", self); btn_export_live.setMinimumHeight(42)
        btn_export_live.setToolTip("Spreadsheet of runs, expected range, and technical audit details.")
        btn_export_live.clicked.connect(self.export_live_snapshot)

        btn_save_html = QPushButton("💾 Save Analysis (HTML)", self); btn_save_html.setMinimumHeight(42)
        btn_save_html.setToolTip("PAEM-CMT report with the best estimate, expected range, and qualitative analysis.")
        btn_save_html.clicked.connect(self.export_report_html)

        btn_restart = QPushButton("🔄 Restart PAEM-CMT", self); btn_restart.setMinimumHeight(42)
        btn_restart.setObjectName("dangerBtn")
        btn_restart.setToolTip("Clear current session and return to the first screen.")
        btn_restart.clicked.connect(self._confirm_restart_app)

        r.addStretch(1)
        r.addWidget(btn_export_live)
        r.addWidget(btn_save_html)
        r.addWidget(btn_restart)
        vb.addWidget(row)

        card = self._make_card("📌 Summary", box)
        self._final_summary_holder.layout().addWidget(card)

    def _on_eval_finished(self, data: dict):
        self._stop_elapsed_clock()
        if hasattr(self, "eval_btn"): self.eval_btn.setEnabled(True)
        self._safe_set_chip(self.progress_chip, "done ✅")
        if hasattr(self, "eval_chip"): self._safe_set_chip(self.eval_chip, "done ✅")

        try:
            if hasattr(self, "_startup_popup_timer") and self._startup_popup_timer.isActive():
                self._startup_popup_timer.stop()
        except Exception:
            pass
        self._close_busy_popup()
        self._open_busy_popup("finalizing…")
        QApplication.processEvents()

        def _build():
            try:
                self._show_final_summary_card(data)
                self._last_eval_data = data
            except Exception as e:
                QMessageBox.warning(self, "Finished", f"Evaluation finished but UI summary failed:\n{e}")
                self._last_eval_data = data
            finally:
                self._close_busy_popup()

        QTimer.singleShot(0, _build)

    def _stabilize_run_numbers(self, res, anchor=None, prev=None):
        if not isinstance(res, dict):
            return anchor or {"dimensions": {}, "subtotals": {}, "overall": 0.0}

        out  = dict(res)
        dims = dict(out.get("dimensions", {}) or {})

        for k, v in list(dims.items()):
            v = dict(v or {})
            try:
                v["score"] = float(v.get("score", 0.0))
            except Exception:
                v["score"] = 0.0
            reasons = v.get("reasons")
            if reasons is None:
                reasons = v.get("justification")
            v["reasons"] = list(reasons or [])
            v["justification"] = v["reasons"]
            dims[k] = v

        out["dimensions"] = dims

        try:
            out["overall"] = float(out.get("overall", 0.0))
        except Exception:
            out["overall"] = 0.0

        return out

    def _term_norm_text(self, s):
        import re, unicodedata
        s = unicodedata.normalize("NFKC", (s or ""))
        s = s.replace("\u00A0", " ")
        s = s.replace("’", "'").replace("‘", "'").replace("`", "'").replace("´", "'")
        s = re.sub(r"[\u2010\u2011\u2012\u2013\u2014\u2015]", "-", s)
        s = re.sub(r"\s+", " ", s).strip().lower()
        return s

    def _term_family_pattern(self, term):
        import re

        t = self._term_norm_text(term)
        if not t:
            return None

        toks = [x for x in re.split(r"[\s\-]+", t) if x]
        if not toks:
            return None

        joiner = r"(?:[\s\-]+)"
        tail = r"(?:['’-]?\w+){0,3}"

        if len(toks) == 1:
            return rf"(?<!\w){re.escape(toks[0])}{tail}(?!\w)"

        head = joiner.join(re.escape(x) for x in toks[:-1])
        last = re.escape(toks[-1])
        return rf"(?<!\w){head}{joiner}{last}{tail}(?!\w)"

    def _count_term_family_occurrences(self, term, text):
        import re

        txt = self._term_norm_text(text)
        pat = self._term_family_pattern(term)

        if not txt or not pat:
            return 0, []

        spans = []
        try:
            for m in re.finditer(pat, txt, flags=re.I | re.UNICODE):
                spans.append((m.start(), m.end(), m.group(0)))
            return len(spans), spans
        except Exception:
            t = self._term_norm_text(term)
            if not t:
                return 0, []
            count = txt.count(t)
            return count, []

    def _term_realized_in_text(self, term, text):
        count, _ = self._count_term_family_occurrences(term, text)
        return count > 0

    def _adjudicate_term_obligations_llm(self, source_text, translation_text, items):
        import json, hashlib

        def _make_fallback(status: str):
            out = {}
            for idx, item in enumerate(items or []):
                out[idx] = {
                    "relevant_explicit": 0,
                    "hidden_fulfilled": 0,
                    "alternative_used": False,
                    "reason": "",
                    "adjudication_status": status,
                }
            return out

        defaults = _make_fallback("fallback_error")

        if not items:
            return {}

        model = (self.model_name or "gpt-5.4-2026-03-05").strip()
        cache_key_payload = {
            "logic_version": getattr(self, "term_logic_version", "obligation_v2"),
            "model": model,
            "source_text": source_text,
            "translation_text": translation_text,
            "items": items,
        }
        cache_key = hashlib.sha256(
            json.dumps(cache_key_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest()

        cache = getattr(self, "_term_obligation_cache", None)
        if isinstance(cache, dict) and cache_key in cache:
            return cache[cache_key]

        if client is None:
            out = _make_fallback("fallback_no_client")
            if isinstance(cache, dict):
                cache[cache_key] = out
            return out

        system_msg = (
            "You are auditing term obligations for a purpose-aligned MT evaluation metric.\n"
            "Your job is conservative and literal.\n"
            "You MUST NOT hallucinate, add, delete, rewrite, normalize away, or invent any source data, translation data, evidence, counts, or judgments.\n"
            "You MUST use only the exact provided source_text, translation_text, and items.\n"
            "You MUST NOT infer unstated occurrences or unstated alternative lexicalizations.\n"
            "You MUST NOT change the meaning of any term pair.\n"
            "If the evidence is insufficient, stay conservative.\n"
            "\n"
            "IMPORTANT RULES\n"
            "1) The required target term itself is mandatory. A different lexical choice is a violation.\n"
            "2) Normal inflectional/plural/possessive/case-attached forms of the SAME lexical item count as explicit realization.\n"
            "3) Count only uses of the required target term family that fulfill the related source-side term obligation in the relevant place/context. Do NOT reward random, unrelated, or displaced insertions.\n"
            "4) A missing visible occurrence may still count as fulfilled ONLY if the translation compressed, merged, or pronominalized the content without introducing a competing alternative lexical item.\n"
            "5) If more explicit occurrences of the required target term family appear for natural target-language reasons, do NOT treat that as a violation by itself. Count only source-side obligations, and reward only relevant occurrences up to that obligation count.\n"
            "6) If a competing alternative lexicalization is used for the obligation, treat that obligation as violated rather than hidden-fulfilled.\n"
            "7) If unsure, stay conservative.\n"
            "\n"
            "Return JSON only. Output a valid JSON object with this exact schema:\n"
            "{\n"
            '  "items": [\n'
            "    {\n"
            '      "index": 0,\n'
            '      "relevant_explicit": 0,\n'
            '      "hidden_fulfilled": 0,\n'
            '      "alternative_used": false,\n'
            '      "reason": ""\n'
            "    }\n"
            "  ]\n"
            "}\n"
        )

        user_payload = {
            "source_text": source_text,
            "translation_text": translation_text,
            "items": items,
        }

        try:
            rsp = client.chat.completions.create(
                model=model,
                temperature=0.0,
                max_completion_tokens=128000,
                response_format={"type": "json_object"},
                **_gpt5_effort_none_kwargs(model),
                messages=[
                    {"role": "system", "content": system_msg},
                    {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                ],
            )

            raw = (rsp.choices[0].message.content or "").strip()
            if raw.startswith("```"):
                import re
                raw = re.sub(r"^```(?:json)?|```$", "", raw, flags=re.S).strip()

            data = json.loads(raw)
            out = {}

            for row in data.get("items", []) or []:
                try:
                    idx = int(row.get("index", -1))
                except Exception:
                    continue

                if idx < 0 or idx >= len(items):
                    continue

                src_occ = int(items[idx].get("source_occurrences", 0))
                exp_occ = int(items[idx].get("explicit_target_occurrences", 0))

                try:
                    relevant_explicit = int(row.get("relevant_explicit", 0))
                except Exception:
                    relevant_explicit = 0

                try:
                    hidden_fulfilled = int(row.get("hidden_fulfilled", 0))
                except Exception:
                    hidden_fulfilled = 0

                alternative_used = bool(row.get("alternative_used", False))
                reason = str(row.get("reason", "") or "").strip()

                relevant_explicit = max(0, min(relevant_explicit, min(src_occ, exp_occ)))
                hidden_cap = max(0, src_occ - relevant_explicit)
                hidden_fulfilled = max(0, min(hidden_fulfilled, hidden_cap))

                if alternative_used:
                    hidden_fulfilled = 0

                out[idx] = {
                    "relevant_explicit": relevant_explicit,
                    "hidden_fulfilled": hidden_fulfilled,
                    "alternative_used": alternative_used,
                    "reason": reason,
                    "adjudication_status": "ok",
                }

            for idx, item in enumerate(items):
                if idx not in out:
                    out[idx] = {
                        "relevant_explicit": 0,
                        "hidden_fulfilled": 0,
                        "alternative_used": False,
                        "reason": "",
                        "adjudication_status": "fallback_partial",
                    }

            if isinstance(cache, dict):
                cache[cache_key] = out
            return out

        except Exception as e:
            try:
                print(Fore.MAGENTA + f"[WARN] term obligation adjudication fallback: {e}" + Style.RESET_ALL)
            except Exception:
                pass
            out = _make_fallback("fallback_error")
            if isinstance(cache, dict):
                cache[cache_key] = out
            return out

    def _compute_termlist_stats(self, text_html):
        ctx = getattr(self, "translation_context", {}) or {}
        pairs = ctx.get("termpairs", []) or []

        source_text = strip_html_tags(getattr(self, "docx_content", "") or "")
        raw_text = strip_html_tags(text_html or "")

        required_pairs = [
            ((src or "").strip(), (tgt or "").strip())
            for src, tgt in pairs
            if (src or "").strip() and (tgt or "").strip()
        ]

        pair_decisions = []
        active_items = []

        for pair_index, (src, tgt) in enumerate(required_pairs):
            src_occ, _ = self._count_term_family_occurrences(src, source_text)
            tgt_occ, _ = self._count_term_family_occurrences(tgt, raw_text)

            row = {
                "source_term": src,
                "target_term": tgt,
                "source_occurrences": int(src_occ),
                "explicit_target_occurrences": int(tgt_occ),
                "relevant_explicit": 0,
                "hidden_fulfilled": 0,
                "fulfilled": 0,
                "missed": 0,
                "alternative_used": False,
                "status": "not_applicable" if int(src_occ) == 0 else "pending",
                "adjudication_status": "not_applicable" if int(src_occ) == 0 else "pending",
                "reason": "",
            }
            pair_decisions.append(row)

            if int(src_occ) > 0:
                active_items.append({
                    "pair_index": pair_index,
                    "source_term": src,
                    "target_term": tgt,
                    "source_occurrences": int(src_occ),
                    "explicit_target_occurrences": int(tgt_occ),
                })

        adjudicated = self._adjudicate_term_obligations_llm(source_text, raw_text, active_items)

        for active_idx, item in enumerate(active_items):
            pair_index = int(item["pair_index"])
            row = pair_decisions[pair_index]
            dec = adjudicated.get(active_idx, {}) or {}

            src_occ = int(row["source_occurrences"])
            tgt_occ = int(row["explicit_target_occurrences"])

            adjudication_status = str(dec.get("adjudication_status", "ok") or "ok")

            relevant_explicit = int(dec.get("relevant_explicit", 0) or 0)
            relevant_explicit = max(0, min(relevant_explicit, min(src_occ, tgt_occ)))

            hidden_cap = max(0, src_occ - relevant_explicit)
            hidden_fulfilled = int(dec.get("hidden_fulfilled", 0) or 0)
            hidden_fulfilled = max(0, min(hidden_fulfilled, hidden_cap))

            alternative_used = bool(dec.get("alternative_used", False))

            if adjudication_status != "ok":
                relevant_explicit = 0
                hidden_fulfilled = 0
                alternative_used = False

            if alternative_used:
                hidden_fulfilled = 0

            fulfilled = max(0, min(src_occ, relevant_explicit + hidden_fulfilled))
            missed = max(0, src_occ - fulfilled)

            row["relevant_explicit"] = relevant_explicit
            row["hidden_fulfilled"] = hidden_fulfilled
            row["fulfilled"] = fulfilled
            row["missed"] = missed
            row["alternative_used"] = alternative_used
            row["adjudication_status"] = adjudication_status
            row["status"] = "provisional" if adjudication_status != "ok" else ("fulfilled" if missed == 0 else "violated")
            row["reason"] = str(dec.get("reason", "") or "").strip()

        matched_pairs = []
        missing_pairs = []

        for row in pair_decisions:
            if int(row.get("source_occurrences", 0)) <= 0:
                continue
            pair = (row["source_term"], row["target_term"])
            if int(row.get("missed", 0)) == 0:
                matched_pairs.append(pair)
            else:
                missing_pairs.append(pair)

        total_terms = sum(int(r.get("source_occurrences", 0)) for r in pair_decisions if int(r.get("source_occurrences", 0)) > 0)
        matched = sum(int(r.get("fulfilled", 0)) for r in pair_decisions if int(r.get("source_occurrences", 0)) > 0)
        misses = max(0, total_terms - matched)
        miss_rate = (misses / total_terms) if total_terms else 0.0

        if total_terms == 0:
            term_score = None
        elif matched >= total_terms:
            term_score = 5.0
        elif matched == 0:
            term_score = 0.0
        else:
            term_score = 5.0 * (matched / total_terms)

        active_statuses = [
            str(r.get("adjudication_status", "ok"))
            for r in pair_decisions
            if int(r.get("source_occurrences", 0)) > 0
        ]

        if not active_statuses:
            adjudication_status = "not_applicable"
        elif all(s == "ok" for s in active_statuses):
            adjudication_status = "ok"
        else:
            adjudication_status = "provisional"

        return {
            "required_pairs": required_pairs,
            "matched_pairs": matched_pairs,
            "missing_pairs": missing_pairs,
            "total_terms": total_terms,
            "matched": matched,
            "misses": misses,
            "miss_rate": miss_rate,
            "term_score": term_score,
            "catastrophic_term_violation": (total_terms > 0 and misses == total_terms),
            "pair_decisions": pair_decisions,
            "logic_version": getattr(self, "term_logic_version", "obligation_v3_relaxed_matcher"),
            "adjudication_status": adjudication_status,
        }

    def _recompute_single_run_metric(self, res):
        import re, math, unicodedata

        out = dict(res or {})
        flat = self._flatten_dimensions(out.get("dimensions", {}) or {})

        dims = {}
        for d in self.STANDARD_HEADINGS:
            info = dict(flat.get(d, {}) or {})
            try:
                score = float(info.get("score", 0.0))
            except Exception:
                score = 0.0

            reasons = info.get("justification")
            if reasons is None:
                reasons = info.get("reasons", [])
            if isinstance(reasons, str):
                reasons = [reasons]
            else:
                reasons = list(reasons or [])

            dims[d] = {
                "score": score,
                "reasons": reasons,
                "justification": reasons[:],
            }

        def _norm(s):
            return unicodedata.normalize("NFKC", (s or "")).lower().strip()

        def _negatives(dim_info):
            rs = list((dim_info or {}).get("reasons") or [])
            c = 0
            for r in rs:
                s = str(r or "").lstrip()
                if s.startswith("✓"):
                    continue
                if "—" in s or " - " in s:
                    c += 1
            return c

        def _smooth_drop(neg_count: int, unit=0.45, k=0.85, cap=2.0):
            if neg_count <= 0:
                return 0.0
            base = (1 - math.exp(-k * neg_count)) / (1 - math.exp(-k))
            return min(cap, unit * neg_count * 0.35 + cap * 0.65 * base)

        term_stats = self._compute_termlist_stats(getattr(self, "_last_translation_html", "") or "")
        required_pairs = term_stats["required_pairs"]
        total_terms = term_stats["total_terms"]
        misses = term_stats["misses"]
        miss_rate = term_stats["miss_rate"]
        term_score_det = term_stats["term_score"]
        catastrophic_term_violation = term_stats["catastrophic_term_violation"]
        pair_decisions = term_stats.get("pair_decisions", [])

        role_dims = ["Target Audience", "Intended Purpose", "Translator", "Source Owner", "Commissioner"]

        for dim_name, dim in list(dims.items()):
            s = float(dim.get("score", 0.0))

            if dim_name == "Terminology Adherence":
                if term_score_det is not None:
                    s = float(term_score_det)
            else:
                if dim_name in role_dims and total_terms > 0 and misses > 0:
                    s = max(0.0, s - miss_rate * 0.80)

                if (
                    dim_name in role_dims
                    and catastrophic_term_violation
                    and (term_score_det is not None)
                    and float(term_score_det) <= 1e-9
                ):
                    s = 2.50 * ((max(0.0, min(5.0, s)) / 5.0) ** 1.35)

            dims[dim_name]["score"] = s

        if "Terminology Adherence" in dims:
            req_targets = [tgt for _, tgt in required_pairs]

            def _valid_term_strength(bullet: str) -> bool:
                b = str(bullet or "")
                if not b.lstrip().startswith("✓"):
                    return True
                return any(self._term_realized_in_text(tgt, b) for tgt in req_targets)

            dims["Terminology Adherence"]["reasons"] = [
                b for b in (dims["Terminology Adherence"].get("reasons") or [])
                if _valid_term_strength(b)
            ]
            dims["Terminology Adherence"]["justification"] = dims["Terminology Adherence"]["reasons"][:]

        if "Terminology Adherence" in dims:
            term_dim_score = float((dims.get("Terminology Adherence") or {}).get("score", 0.0))
            if term_dim_score <= 1e-9:
                dims["Terminology Adherence"]["reasons"] = [
                    b for b in (dims["Terminology Adherence"].get("reasons") or [])
                    if not str(b or "").lstrip().startswith("✓")
                ]
                dims["Terminology Adherence"]["justification"] = dims["Terminology Adherence"]["reasons"][:]

        role_vals = [dims[d]["score"] for d in role_dims if d in dims]
        role_score = (sum(role_vals) / len(role_vals)) if role_vals else 0.0
        term_score = float((dims.get("Terminology Adherence", {}) or {}).get("score", 0.0))

        shown_scores = [
            float((dims.get(d, {}) or {}).get("score", 0.0))
            for d in self.STANDARD_HEADINGS
            if d in dims
        ]
        overall_base = (sum(shown_scores) / len(shown_scores)) if shown_scores else 0.0

        overall = overall_base

        out["overall"] = float(overall)
        out["dimensions"] = dims
        out["subtotals"] = {
            "Role Satisfaction": role_score,
            "Terminology Adherence": term_score,
        }
        out["_term_total"] = total_terms
        out["_term_misses"] = misses
        out["_term_pair_decisions"] = pair_decisions
        out["_term_logic_version"] = getattr(self, "term_logic_version", "obligation_v3_relaxed_matcher")
        out["_term_adjudication_status"] = term_stats.get("adjudication_status", "not_applicable")
        out["dimensions"] = self._freeze_scored_reasons(out.get("dimensions", {}))
        return out    

    def _perform_evaluation_core(self, max_runs=40, min_runs=5, threshold=0.05, on_pulse=lambda *_: None):
        """
        Run repeated PAEM-CMT scoring until the internal stability check is sufficient
        or the run limit is reached.
        """
        import re, json, statistics
        from difflib import SequenceMatcher
        import unicodedata

        def ui_pulse(msg: str):
            on_pulse(msg)

        def fuzzy_unique(reasons, threshold=0.65):
            unique = []
            for r in reasons:
                if not any(SequenceMatcher(None, r, u).ratio() > threshold for u in unique):
                    unique.append(r)
            return unique

        if not isinstance(max_runs, int) or max_runs < 1:
            max_runs = 40
        if not isinstance(min_runs, int) or min_runs < 1:
            min_runs = 5
        if not isinstance(threshold, (float, int)) or threshold <= 0:
            threshold = 0.05

        ctx = getattr(self, "translation_context", {})
        roles = {
            "Intended Purpose":   ctx.get("purpose", ""),
            "Target Audience":    ctx.get("audience", ""),
            "Translator":         ctx.get("translator", ""),
            "Source Owner":       ctx.get("owner", ""),
            "Commissioner":       ctx.get("commissioner", "")
        }
        roles = {k: v for k, v in roles.items() if v}

        pairs = ctx.get("termpairs", []) or []
        termlist = "\n".join(f"{a} → {b}" if b else f"{a} → (target TBD)" for a,b in pairs)

        translation_html = (getattr(self, "_last_translation_html", "") or "").strip()
        if not translation_html:
            pv = getattr(self, "tgt_preview", None)
            try:
                translation_html = (pv.toHtml().strip() if pv else "")
            except Exception:
                translation_html = ""
        if not translation_html:
            raise RuntimeError("No translation content available. Upload a translation file on the Setup screen.")
        translation = translation_html

        source_html      = self.extract_text_with_formatting(self.file_path)
        translation_text = strip_html_tags(translation)
        source_text      = strip_html_tags(source_html)
        translation_for_judge = translation_text
        source_for_judge      = source_text

        roles_list = "\n".join(f"- {r}: {roles[r]}" for r in roles)

        paemcmt_prompt = f"""
PAEM-CMT is a purpose-aligned evaluation metric that assesses how well a machine translation output has been customized to the expectations of a specific translation task. Use the provided context — translation instructions, roles, and termlist — to evaluate the translation output on two main pillars:

⚠️ *Use the full 0.00–5.00 scale.
Do not hesitate to assign scores below 3.50 when medium or major issues appear.

SCORING LEGEND (0.00–5.00; use two decimals)
0.00–0.99  total failure of customization; harmful/inverted intent
1.00–1.99  wrong or minimal customization; major mismatches
2.00–2.49  attempted but ineffective; many criteria unmet
2.50–2.99  attempted; mixed; several medium issues; major revision needed
3.00–3.50  partly customized; some core criteria met; clear issues remain
3.51–3.99  generally customized; minor/occasional medium issues
4.00–4.50  well customized; small polish items
4.51–5.00  fully aligned; minor or no changes needed

SCORING GUARDRAILS
• Compute the numeric as 5.00 − Σ(deductions). Output two decimals. Never add “bonus” points for positives; positives are descriptive, not compensatory.
• Use the full 0.00–5.00 range. If medium or major issues exist, scores must not exceed 3.90.
• Severities use two-decimal values **inside ranges** (no snapping, no centroids):
  minor 0.14–0.22, medium 0.32–0.42, major 0.58–0.70, critical 1.80–2.05.
  Do not use the exact values (0.18, 0.36, 0.62, 1.90). Do not default to midpoints; distribute values across issues according to evidence strength.
• Deduct once per underlying issue for the whole text (after merging near-duplicates). Repetitions escalate one severity level; do not add another line.
• Conservative tie-break: if unsure between two severities, pick the lower one.
• Be consistent across runs; however, when two severities are tied or borderline, apply a deterministic tiebreak: lower severity if the evidence is weak, higher if it’s pervasive. Small ±0.01 (borderline) nudges are acceptable on those ties.
• Terminology violations (from the provided termlist) propagate into role dimensions; do not keep roles high when terminology contradicts the brief.
• Never round up; always output two decimals.


ISSUE COUNTING RULES
• Work negative-first: identify all defects before listing positives.
• One line per underlying defect type: term mismatch (per unique term), brief/role contradiction, meaning error, register/tone violation, formatting/structure miss, etc.
• If the same defect recurs, escalate its severity once (e.g., minor→medium). Do not add another line.
• Merge bullets that describe the same defect before assigning a single severity/deduction.
• Positives never reduce deductions; they only describe alignment achieved.
• Assign each defect to exactly one most-relevant dimension (no duplicates across dimensions).
• Compute each dimension’s numeric independently using only that dimension’s defects; do not copy the overall numeric into every dimension.
If a dimension has no specific defects, omit that dimension instead of repeating another dimension’s numeric.


SCORING PIPELINE (strict order)
1) Detect defects (negative-first) and normalize to these categories:
   terminology_violation (per unique term), brief_contradiction, meaning_error,
   register_tone, format_structure, other.
2) Merge near-duplicates; count each underlying defect once per text. If it repeats, escalate one severity level.
3) Assign severities using those ranges; pick the precise value deterministically to avoid clustering:
   Let s = SHA1( (meta.run_seed + "|" + evidence_text[:16]).lower() ).
   Let r = int(s[:2], 16) / 255.0.
   Pick value = lower_bound + r * (band_width - 0.01) (clamp to band).
   If a defect repeats (escalation), bias to the upper half of the band.
   Mark (borderline) only if you must move > +0.02.
4) Compute D = sum(deductions). Provisional = 5.00 − D.
5) Apply the EVIDENCE-BASED CEILINGS above to get Final. Round to two decimals. Output Final.

EVIDENCE-BASED CEILINGS (apply after computing Provisional)
• If any critical defect exists → Final = min(Provisional, 3.20).
• Else if any major defect exists → Final = min(Provisional, 3.80).
• Else if count(medium) ≥ 3 → Final = min(Provisional, 3.60).
• Else if count(minor) ≥ 5 → Final = min(Provisional, 4.20).
• Output Final (two decimals).




Terminological Customization is NORMATIVE (uses the provided termlist):
- The required target term itself is mandatory. A different lexical choice is a terminology violation.
- Count normal inflectional / plural / possessive / case-attached forms of the SAME lexical item as realization of the required term.
- Do NOT reward a required target term if it appears in an unrelated place. It must fulfill the relevant source-side term obligation.
- If fewer explicit target-term occurrences appear because the translation merges clauses/sentences, pronominalizes, or avoids unnecessary repetition, that may still be compliant ONLY when no competing alternative lexicalization is introduced.
- Treat a violated required term as a cross-impact on every related role as well, not only a terminology issue.
- If every required term is violated, do NOT keep affected role dimensions in the high-good range; treat this as a major cross-role failure.
- Use explicit negative bullets for these cross-impacts.
- Count at most one deduction bullet per unique source→target term pair. If the same pair is violated in multiple places, escalate severity rather than duplicating bullets.
— REPORTING POLARITY OVERRIDE —
- Any competing alternative lexicalization of a required term is always a weakness.
- Only correct realization of the required target term family may be reported as a strength.
- Do not claim “term not used” if the quoted snippet already contains the required target term family.

MINIMUM DEDUCTIONS (floors per defect)
- Missing/incorrect required term from the termlist (per unique term pair with one or more violated obligations): ≥ medium (≥0.34).
- If two or more unique required target terms have violated obligations, treat Terminology Adherence as a major failure; its score must not exceed 3.80.
- If every source-side obligation created by the termlist is violated, Terminology Adherence must not exceed 3.50.
- In Terminology Adherence, a strength bullet is allowed only if its quoted translation span contains the required target term family. Do not praise unrelated terminology there.
- Direct contradiction to the brief/role (audience, tone, persona, constraints): ≥ major (≥0.56).
- Meaning-changing mistranslation of a task-critical span: ≥ major (≥0.56).
- Systematic register/tone mismatch affecting most of the text: ≥ medium (≥0.34).
- Required structure/format instruction not followed (when specified): ≥ medium (≥0.34).


1. Role Expectation Satisfaction:
Evaluate how well the translation meets the expectations of each role listed below. Each role has its own perspective:
{roles_list if roles_list else "- (No roles provided, skip this section)"}

2. Terminological Customization:
Evaluate the translation’s use of the provided termlist. Consider:
- Termlist Match Rate: Were source-side term obligations fulfilled with the required target term family?
- Relevance of Use: Were required target-term occurrences used in the relevant places, rather than inserted randomly or unrelatedly?
- Obligation Fulfillment Under Restructuring: If fewer explicit target-term occurrences appear because of sentence merging, pronominalization, or compression, was the obligation still fulfilled without introducing a competing alternative lexicalization?
- Extra Explicit Uses: If more explicit target-term occurrences appear for natural target-language reasons, count only source-side obligations. Do not treat extra relevant occurrences as a violation by themselves, and do not award bonus credit beyond full fulfillment.
- Integration Quality: Were the required target terms integrated naturally where they were used?
- Catastrophic all-term failure: If every required term is violated, affected role dimensions must not remain in the high-good range.

Only include dimensions for roles or termlists that are present.

**JUSTIFICATION & SUGGESTION RULES – read carefully**

• For every dimension return at most 5 bullet-level items. Fewer is fine.
• STRICT bullet format:
  <issue statement> — "<verbatim snippet (≤ 15 words)>"
  - The snippet MUST be a verbatim substring from the Translation; Use Source only as background evidence when needed to explain a translation choice; Source differences do NOT count as concerns unless they clearly reduce brief fulfillment.
  - If you cannot find ANY snippet (≤15 words) that supports the issue: SKIP this bullet entirely.
  - Use straight double quotes "…". Exactly ONE quoted snippet per bullet.
  - To skip middle words, use two anchors separated by a comma or semicolon inside the quotes
    ; both anchors must appear in-order in the same sentence.
• One-snippet-one-time per dimension: do not reuse the same quoted text in multiple bullets in that dimension.
• No repetition or paraphrasing of the same idea across bullets.
• Keep the issue statement short (≤ 20 words), content-focused.

BRIEF-ONLY PRIORITY
• Primary and only criterion: how well the Translation fulfills the brief and role expectations.
• A source-vs-translation difference is NOT a weakness by itself.
• Do NOT penalize non-retention of source wording, framing, list structure, explicitness, or examples unless the brief clearly requires that element.
• Use the Source only to understand what the Translation is doing, not as a scoring target.
• A concern must be framed in terms of brief fulfillment, audience fit, tone, task effectiveness, or required terminology — not in terms of source loss alone.
• If a Translation differs from the Source but still fulfills the brief well, do not penalize it.
• If a source-side omission or shift is mentioned, explicitly connect it to brief failure; otherwise drop it.

ONE-CLAIM RULE
• Each bullet must make only one evaluative claim.
• Do not combine two different concerns into one bullet.
• Do not combine praise and caution in the same bullet.
• If a point contains both a strength and a limitation, split them or keep only the dominant one.
• Avoid conjunction-led bundling such as “and”, “but”, “while”, “although” when they join separate evaluative claims.

CONCERN FRAMING RULE
• A concern bullet must be written from the problem side, not from the benefit side.
• Do not open a concern bullet with praise, mitigation, or partial credit.
• In a concern bullet, do not begin with wording such as:
  - is good / works well / is effective
  - is natural / fluent / readable / clear / smooth
  - is warm / polite / professional / engaging
  - helps / supports / improves / strengthens
  - preserves / captures / conveys well
  unless the bullet is actually a strength.
• If a point contains both a benefit and a drawback, either:
  - split it into two bullets, or
  - keep only the drawback if the bullet belongs under concerns.
• A concern bullet should name the brief-relevant problem first, then support it with evidence.
• Prefer formulations like:
  - is too ...
  - is less ...
  - omits ...
  - adds ...
  - shifts ...
  - weakens ...
  - overstates ...
  - understates ...
  - makes ... less/more ...
• Do not use “X is good, but Y...” structure in a concern bullet.
• If the overall judgment is negative, phrase the bullet as negative throughout.
  
ANTI-OVERREACH / CALIBRATION RULES
• Judge the Translation against the brief, not against source retention.
• Do not convert a source-vs-translation difference into a concern unless it clearly reduces brief fulfillment.
• When evidence supports only a source difference, do not report it as a weakness.
• Treat moderate issues as moderate: describe the brief-relevant problem itself, not source loss alone.
• Do not over-penalize arguable stylistic alternatives unless they clearly reduce task fulfillment.
• Changes in framing, list composition, explicitness, or tone count as concerns only when they materially affect purpose, audience fit, role expectations, or required terminology.
• If an issue is real but limited, state it narrowly and score it proportionally.
• Prefer narrow, proportional judgments over broad interpretive claims.

JUDGMENT STYLE POLICY
• Be strict in scoring, but neutral in phrasing.
• Prefer textual observations over rhetorical or predictive language.
• Use verbs like: adds, omits, shifts, narrows, softens, strengthens, makes explicit, makes less explicit, is more formal, is less direct.
• Avoid speculative consequence wording unless directly supported by the brief or source.
• Avoid phrases like: could backfire, may reduce trust, risks backlash, could alienate readers, may undermine impact, unless that consequence is explicitly grounded in the brief/source.
• For arguable stylistic issues, use restrained wording such as: slightly, somewhat, a bit more, a bit less, more/less natural, more/less direct.
• Escalate tone only for clear major errors, contradictions, invented content, or direct brief violations.
• When such a major/critical brief-level concern is output as a bullet, prefix that concern bullet with "! ".
• Do not overstate list reshaping, tone shifts, or framing changes unless they clearly alter task fulfillment.

STRENGTHS CAPTURE
• After listing negative defects for a dimension, add 0–2 concise strengths only if they are cleanly positive and do NOT contradict any concern anywhere.
• Strengths MUST be genuinely positive properties of the current Translation (e.g., correct required term usage, accurate meaning, structure followed).
• Never praise a deviation from the termlist or the brief. If a required term was not used (or was replaced), that is NEVER a strength.
• Keep strengths in the SAME "justification" list; prefix each with "✓ ".
• If a concern is major or critical relative to the brief, prefix it with "! ". Do NOT use "! " for minor or moderate concerns.
• A strength bullet must contain only praise. Do NOT include any caveat, contrast, reservation, or downside in a strength bullet.
• Do NOT use “but”, “however”, “while”, “although”, or similar contrastive framing inside a strength bullet.
• Strict format still applies: <statement> — "<verbatim snippet (≤15 words)>"
• If no true strengths exist for that dimension, output none.
• For strong but not perfect dimensions, include at most one concise minor note only if it clearly explains why the score is not higher; do not force a note when none exists, never hallucinate on given data or judgement.



SELF-CHECK (hard gate before you output):
1) Does EVERY bullet contain exactly one pair of "quotes"? If not, drop that bullet.
2) Is the inside ≤ 15 words? If not, shorten or drop.
3) Does the quote appear verbatim (case-insensitive) in Translation or Source? If not, pick a different quote or drop.
4) If the bullet is based on a Source-vs-Translation difference, have you made its relevance to brief fulfillment explicit? If not, drop it.
5) If the bullet only says that something from the Source is missing, without showing why the brief needed it, drop it.

CONSISTENCY & POLARITY GUARD (hard gate)
• Global no-contradiction: Do not output a strength that negates a weakness about the same underlying point anywhere (any dimension).
• Terminology polarity lock: If any provided source→target term is violated (missing, replaced, or mistranslated), that event MUST appear as a weakness in "Terminological Customization" and may appear as cross-impact weaknesses in affected roles. It may NOT be framed as a strength in any dimension.
• Allowed terminology strength examples (when true): "✓ Required term used consistently — "<quote>""; "✓ Terms integrated naturally — "<quote>"".
• Forbidden strength examples (never output): "✓ Uses alternative to required term …", "✓ Avoids required term …", "✓ Chooses different term for nuance …".
• If a previous bullet would be contradicted by a planned strength, DROP the strength instead of rephrasing it as positive.
Example (terminology, language-agnostic): If the termlist mandates target term "<T_REQUIRED>" for source concept "<S_CONCEPT>", but the Translation uses a different term "<T_ALT>", report a weakness in "Terminological Customization" (and in any impacted Roles). Do NOT frame the deviation as a strength; phrase strengths only as correct use of "<T_REQUIRED>", never as "avoids '<T_ALT>'".




•  After the bullet list, add **"suggestion": "…"**:  
  - If the dimension’s score ≤ 4.8 → one concrete, actionable fix (≤ 20 words).  
  - If the score > 4.8 → `"why not 5"`.

EVIDENCE & QUOTING RULES (deterministic mode)
• ZERO-QUOTE RULE: *Every bullet MUST include exactly one quoted snippet* (≤ 15 words) taken from the Source or the Translation. If you cannot find a valid snippet, DO NOT output that bullet; replace it with a different, evidence-backed bullet or reduce the count.
• SIGNATURE: Each item in "justification" MUST contain one em dash ` — ` and exactly one pair of straight double quotes `"…"`.
• SOURCE OF QUOTES: Only quote Source/Translation text. Never quote the brief/roles/termlist or your own wording.
• EVIDENCE-FIRST ALGORITHM (follow in this order for every bullet):
  1) Pick a concrete, checkable point.
  2) Select a 5–12 word span from Source/Translation that anchors that point.
  3) Compose `<issue> — "<span>"`. If no span fits, SKIP this bullet.
• ELLIPSIS: You MAY use “…” or “...” inside the quote to collapse the middle only; left/right parts must match the original exactly. Do not add trailing ellipsis to complete phrases.
• ABSENCE CLAIMS (e.g., “no explicit mention of X”, “does not mention privacy/togetherness”):
  – Do NOT quote the absent term or the brief. Instead, quote a short, relevant sentence/fragment from the Translation (or Source) as context, and name the absent concept in the unquoted issue text.
  – Example pattern (language-agnostic): `Does not explicitly mention privacy — "<relevant translation fragment>"`
• LIST EVIDENCE: When citing a list, you MAY quote the list of content items as they appear (e.g., "item1, item2, item3"). Minor function words/punctuation may appear between items in the actual Translation; item order must be preserved.
• TERMINOLOGY BULLETS: When praising or criticizing a term, quote the actual occurrence (or nearest 5–12 word span) from the Translation. Do NOT quote the termlist itself.
• ANTI-CONTRADICTION: Do not claim “term not used” if your quoted snippet contains that term (or a clear inflection).

Return a JSON object with this exact structure (do **NOT** rename or remove keys):

{{
  "dimensions": {{
    "Intended Purpose":      {{ 
      "score": X.XX, 
      "justification": [ "…", … ],
      "suggestion":   "one concrete fix"
    }},
    "Target Audience":       {{ 
      "score": X.XX, 
      "justification": [ "…", … ],
      "suggestion":   "one concrete fix"
    }},
    "Translator":            {{ 
      "score": X.XX, 
      "justification": [ "…", … ],
      "suggestion":   "one concrete fix"
    }},
    "Source Owner":          {{ 
      "score": X.XX, 
      "justification": [ "…", … ],
      "suggestion":   "one concrete fix"
    }},
    "Commissioner":          {{ 
      "score": X.XX, 
      "justification": [ "…", … ],
      "suggestion":   "one concrete fix"
    }},
    "Terminology Adherence": {{ 
      "score": X.XX, 
      "justification": [ "…", … ],
      "suggestion":   "one concrete fix"
    }}
  }},
  "subtotals": {{
    "Role Satisfaction": X.XX,
    "Terminology Adherence": X.XX
  }},
  "overall": X.XX
}}


Ensure all scores are between 0.00 and 5.00 and use the same scale. Keep justifications short and issue-focused. IMPORTANT: Use **exactly** these six JSON keys under \""dimensions"\"—no others. Return all dimension justifications in English only.
"""

        messages = [
            {"role": "system", "content": "You are a professional evaluator using the PAEM-CMT metric. PAEM-CMT is a purpose-aligned evaluation metric that assesses how well a machine translation output has been customized to the expectations of a specific translation task. Use the full 0.00–5.00 scale. Do **not** hesitate to assign scores below 3.50 when you spot medium or major issues. *Bullets without a quoted snippet are INVALID*. If you cannot find a ≤15-word verbatim quote from the Translation (or Source if needed), DO NOT output that bullet. The only scoring criterion is brief fulfillment. Source differences do not matter unless they clearly reduce how well the Translation fulfills the brief. Be strict in scoring and evidence use, but neutral in wording. Do not use speculative consequence language unless explicitly supported by the brief."},
            {"role": "user", "content": json.dumps({
                "name":        "PAEM-CMT",
                "description": paemcmt_prompt.strip(),
                "content": {
                    "instructions": self.translation_instructions,
                    "source":       source_for_judge,
                    "translation":  translation_for_judge,
                    "roles":        roles,
                    "termlist":     termlist
                }
            }, ensure_ascii=False)}
        ]

        import os, sys
        enc = None
        try:
            from tiktoken.load import load_tiktoken_bpe
            from tiktoken import Encoding
            base_dir = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
            bpe_path = os.path.join(base_dir, "o200k_base.tiktoken")
            if os.path.isfile(bpe_path):
                bpe_ranks = load_tiktoken_bpe(bpe_path)
                enc = Encoding(
                    name="o200k_base.local",
                    pat_str=r"""'s|'t|'re|'ve|'m|'ll|'d| ?\p{L}+| ?\p{N}+| ?[^ \s\p{L}\p{N}]+|\s+(?!\S)|\s+""",
                    mergeable_ranks=bpe_ranks,
                    special_tokens={
                        "<|endoftext|>": 100257,
                        "<|fim_prefix|>": 100258,
                        "<|fim_middle|>": 100259,
                        "<|fim_suffix|>": 100260,
                        "<|endofprompt|>": 100276,
                    },
                )
        except Exception as e:
            print(f"[Tokenizer] Local .tiktoken not used: {e}")

        if enc is None:
            try:
                import tiktoken
                enc = tiktoken.get_encoding("o200k_base")
            except Exception:
                enc = None

        def count_tokens(msgs):
            if enc:
                return sum(len(enc.encode(m.get("content", ""))) + 4 for m in msgs) + 2
            text = "".join(m.get("content", "") for m in msgs)
            return max(1, len(text) // 4)

        EVAL_MODEL = (self.model_name or "gpt-5.4-2026-03-05").strip()
        MAX_COMPLETION = COMPLETION_CAP
        print(Fore.MAGENTA + f"[Token Usage] requesting up to {MAX_COMPLETION} completion tokens (OUTPUT cap)")

        EPS_CI   = float(threshold)
        self._ci_target = float(EPS_CI)
        MAX_RUNS = max_runs
        MIN_RUNS = max(5, min_runs)
        all_scores = []
        stats = LiveStats(on_update=ui_pulse)
        prompt_tok_total     = 0
        completion_tok_total = 0
        EVAL_MODEL = self.model_name

        VARIANCE_MODE = "organic"

        EVAL_TEMP_BASE = 0.26
        TEMP_JITTER    = 0.025

        TOP_P_BASE     = 0.27
        TOPP_JITTER    = 0.015

        MIXTURE_LOCK_DELTA = 0.45

        ANCHOR_BAND_ENABLE       = False
        BAND_EPS            = 0.01
        BAND_CENTER_BLEND   = 0.35
        
        ANCHOR_BAND_TARGET_RUNS  = 30
        EARLY_SKIP_ENABLE        = False

        BAND_WARMUP_KEEP        = 4
        BAND_EXTRA_WARMUP       = 0.06
        BAND_RECENTER_AT        = 2
        BAND_RECENTER_WINDOW    = 5
        BAND_SPIKE_STEP         = 0.04
        BAND_SPIKE_MAX          = 0.12

        BASE_SEED = 1337
        ALPHA     = 0.05

        CI_SMOOTH_MODE = "roll3"
        AUDIT_MODE = str(getattr(self, "audit_mode", "deterministic_v1"))
        AUDIT_MAJOR_DROP_THRESHOLD = int(getattr(self, "audit_major_drop_threshold", 1))

        def _ci_stats(values):
            arr = [float(x) for x in values]
            n = len(arr)
            if n == 0:
                return {"mean": 0.0, "sd": 0.0, "ci": 0.0, "n": 0}

            mu = statistics.mean(arr)
            sd = statistics.stdev(arr) if n > 1 else 0.0

            try:
                from scipy.stats import t as _t
                crit = _t.ppf(1 - ALPHA / 2, df=n - 1) if n > 1 else float("inf")
            except Exception:
                _TCRIT_95 = {
                    1: 12.706, 2: 4.303, 3: 3.182, 4: 2.776, 5: 2.571,
                    6: 2.447, 7: 2.365, 8: 2.306, 9: 2.262, 10: 2.228,
                    11: 2.201, 12: 2.179, 13: 2.160, 14: 2.145, 15: 2.131,
                    16: 2.120, 17: 2.110, 18: 2.101, 19: 2.093, 20: 2.086,
                    24: 2.064, 30: 2.042
                }
                crit = _TCRIT_95.get(n - 1, 1.96) if n > 1 else float("inf")

            ci = crit * sd / math.sqrt(n) if n > 0 else 0.0
            return {"mean": float(mu), "sd": float(sd), "ci": float(ci), "n": n}

        def _linear_quantile(sorted_vals, q):
            vals = [float(x) for x in sorted_vals]
            n = len(vals)
            if n == 0:
                return 0.0
            if n == 1:
                return vals[0]

            q = max(0.0, min(1.0, float(q)))
            pos = (n - 1) * q
            lo = int(math.floor(pos))
            hi = int(math.ceil(pos))
            if lo == hi:
                return vals[lo]
            frac = pos - lo
            return vals[lo] + (vals[hi] - vals[lo]) * frac

        def _user_facing_result_summary(raw_overall_scores):
            vals = sorted(float(x) for x in raw_overall_scores)
            n = len(vals)
            if n == 0:
                return {
                    "best_estimate": 0.0,
                    "expected_range_low": 0.0,
                    "expected_range_high": 0.0,
                    "expected_range_label": "0.00–0.00",
                    "checks": 0,
                    "user_range_method": "none",
                }

            best_estimate = float(statistics.median(vals))

            if n < 5:
                low = float(min(vals))
                high = float(max(vals))
                method = "minmax_small_n_raw_kept_runs"
            else:
                low = float(_linear_quantile(vals, 0.10))
                high = float(_linear_quantile(vals, 0.90))
                method = "central_80_empirical_raw_kept_runs"

            low = min(low, best_estimate)
            high = max(high, best_estimate)

            return {
                "best_estimate": best_estimate,
                "expected_range_low": low,
                "expected_range_high": high,
                "expected_range_label": f"{low:.2f}–{high:.2f}",
                "checks": n,
                "user_range_method": method,
            }

        DEBUG_DUMP_JSON = False
        DEBUG_DUMP_PATH = "paem_debug.jsonl"

        def _norm(s: str, strip_accents: bool = False) -> str:
            import unicodedata, re, html as _html
            s = unicodedata.normalize("NFKC", s or "")
            s = _html.unescape(s).replace("\u00A0", " ")
            s = re.sub(r"(?i)<br\s*/?>", " ", s)
            s = re.sub(r"(?s)<[^>]+>", " ", s)
            s = re.sub(r"[\u200B\u200C\u200D\uFEFF\u2060\u00AD]", "", s)
            s = (s.replace("“", '"').replace("”", '"')
                   .replace("‘", "'").replace("’", "'")
                   .replace("«", '"').replace("»", '"')
                   .replace("‹", "'").replace("›", "'")
                   .replace("\u2026", ".")
                   .replace("\u2013", "-").replace("\u2014", "-"))
            s = re.sub(r"\s+", " ", s).strip().lower()
            if strip_accents:
                s = unicodedata.normalize("NFD", s)
                s = "".join(ch for ch in s if not unicodedata.combining(ch))
                s = unicodedata.normalize("NFC", s)
            return s

        SRC_NORM = _norm(source_text)
        TRG_NORM = _norm(translation_text)

        def _nopunct(s: str) -> str:
            s = re.sub(r"[^\w\s]", "", s)
            return re.sub(r"\s+", " ", s).strip()

        SRC_NOPUNCT = _nopunct(SRC_NORM)
        TRG_NOPUNCT = _nopunct(TRG_NORM)

        def _is_grounded(q_norm: str,
                         words: int,
                         src_override: str = None,
                         trg_override: str = None) -> bool:
            """
            Verify that a quoted snippet exists in source or translation text.
            Supports both legacy and override-based calls.
            """

            if not q_norm:
                return False

            src_text = src_override if src_override is not None else source_text
            trg_text = trg_override if trg_override is not None else translation_text

            SRC_NORM      = _norm(src_text, strip_accents=False)
            TRG_NORM      = _norm(trg_text, strip_accents=False)
            SRC_NORM_ACC  = _norm(src_text, strip_accents=True)
            TRG_NORM_ACC  = _norm(trg_text, strip_accents=True)

            if q_norm in SRC_NORM or q_norm in TRG_NORM:
                return True

            if words >= 4:
                qp = _nopunct(q_norm)
                if qp and (qp in _nopunct(SRC_NORM) or qp in _nopunct(TRG_NORM)):
                    return True

                qa = _norm(q_norm, strip_accents=True)
                if qa in SRC_NORM_ACC or qa in TRG_NORM_ACC:
                    return True

                if words >= 5:
                    qpa = _nopunct(qa)
                    if qpa and (qpa in _nopunct(SRC_NORM_ACC) or qpa in _nopunct(TRG_NORM_ACC)):
                        return True

            return False

        def _best_fuzzy_hit(haystack: str, needle: str, ratio_ok: float = 0.96) -> tuple[float, str | None]:
            """Find the best approximate match of `needle` inside `haystack`."""
            if not needle:
                return 0.0, None
            if needle in haystack:
                return 1.0, needle

            n = needle.strip()
            def _anchors(x: str) -> list[str]:
                x = re.sub(r"\s+", " ", x)
                c = []
                a = x[:6].strip()
                b = x[-6:].strip()
                if len(a) >= 3: c.append(a)
                if len(b) >= 3 and b != a: c.append(b)
                return c or [x[:3]]

            best_r, best_sub = 0.0, None
            for frag in _anchors(n):
                for m in re.finditer(re.escape(frag), haystack):
                    start = max(0, m.start() - 10)
                    for L in (len(n)-2, len(n)-1, len(n), len(n)+1, len(n)+2):
                        if L <= 0:
                            continue
                        end = min(len(haystack), start + L + 20)
                        window = haystack[start:end]
                        for off in range(0, min(20, max(0, len(window)-L+1))):
                            sub = window[off:off+L]
                            r = SequenceMatcher(None, n, sub).ratio()
                            if r > best_r:
                                best_r, best_sub = r, sub
                                if best_r >= ratio_ok:
                                    return best_r, best_sub
            return best_r, best_sub

        def _replace_first_quoted(bullet: str, old_inner: str, new_inner: str) -> str:
            """Replace the first quoted segment equal to old_inner with new_inner."""
            pat = r'([\"“”\'’«‹])\s*' + re.escape(old_inner) + r'\s*([\"“”\'’»›])'
            return re.sub(pat, lambda m: f'{m.group(1)}{new_inner}{m.group(2)}', bullet, count=1)

        QUOTE_RE = re.compile(
            r'“([^”]+)”|\"([^"]+)\"|«([^»]+)»|‹([^›]+)›|(?<!\w)\'([^\'\n]+)\'(?!\w)|‘([^’]+)’'
        )

        def _extract_quotes(text: str) -> list[str]:
            return [g for m in QUOTE_RE.finditer(text or "") for g in m.groups() if g]

        def _deterministic_audit(result_dict: dict) -> list[dict]:
            """Validate bullets and auto-fix minor quote typos when possible."""
            def _clean_edge(s: str) -> str:
                return s.strip().strip('.,;:!?…“”"\'‘’«»()[]{}')

            def _anchors_in_order(quoted: str) -> bool:
                """
                Accept comma-/semicolon-separated lists if each item appears in order
                in the source or translation.
                """
                if not quoted or ("," not in quoted and ";" not in quoted):
                    return False
                parts = re.split(r"\s*[,;]\s*", quoted)
                parts = [p for p in (parts or []) if p and p.strip()]
                if len(parts) < 2:
                    return False
                anchors = [_norm(p) for p in parts]

                def seq_find(text: str) -> bool:
                    pos = 0
                    for a in anchors:
                        i = text.find(a, pos)
                        if i < 0:
                            return False
                        pos = i + len(a)
                    return True

                return seq_find(TRG_NORM) or seq_find(SRC_NORM)

            def _ellipsis_ok(quoted: str) -> bool:
                """
                Treat "…" or "..." inside the quote as a collapsed middle segment.
                """
                if not quoted or ("..." not in quoted and "…" not in quoted):
                    return False
                raw_parts = re.split(r"(?:\u2026|\.{3})", quoted)
                parts = [p.strip() for p in raw_parts if p and p.strip()]
                if not parts:
                    return False
                parts_norm = [_norm(p) for p in parts]

                def seq_find(text: str) -> bool:
                    pos = 0
                    for pn in parts_norm:
                        i = text.find(pn, pos)
                        if i < 0:
                            return False
                        pos = i + len(pn)
                    return True

                return seq_find(TRG_NORM) or seq_find(SRC_NORM)

            issues = []
            dims = result_dict.get("dimensions", {}) or {}

            for dim_name, info in dims.items():
                bullets = info.get("justification", []) or []
                if isinstance(bullets, str):
                    bullets = [bullets]

                seen_snips = set()
                patched_any = False
                new_bullets = list(bullets)

                for idx, b in enumerate(bullets):
                    m_dash = re.search(r"\s[—–-]\s", b)
                    if m_dash:
                        head, tail = b[:m_dash.start()], b[m_dash.end():]
                    elif "—" in b:
                        head, tail = b.split("—", 1)
                    else:
                        head, tail = b, ""

                    pre_raw  = _extract_quotes(head)
                    post_raw = _extract_quotes(tail)
                    pre   = [_clean_edge(q) for q in pre_raw  if q and q.strip()]
                    quotes = [_clean_edge(q) for q in post_raw if q and q.strip()]

                    if not quotes:
                        if pre:
                            issues.append({
                                "dimension": dim_name,
                                "justification": b,
                                "why": "Examples-only quotes before —; expected a ≤15-word snippet after —."
                            })
                            continue
                        issues.append({
                            "dimension": dim_name,
                            "justification": b,
                            "why": "No quoted snippet found (must include one ≤ 15 words)."
                        })
                        continue

                    found_any = False
                    saw_too_long_only = True
                    chosen_norm = None
                    too_long_seen = False

                    for q in quotes:
                        wcnt = len([w for w in q.split() if w.strip()])
                        if wcnt > 15:
                            too_long_seen = True
                            continue
                        saw_too_long_only = False

                        qn = _norm(q)

                        if _is_grounded(qn, wcnt):
                            chosen_norm = qn
                            found_any = True
                            break

                        if ("..." in q or "…" in q) and _ellipsis_ok(q):
                            chosen_norm = _norm(q)
                            found_any = True
                            break

                        if ("," in q or ";" in q) and _anchors_in_order(q):
                            chosen_norm = _norm(q)
                            found_any = True
                            break

                        r1, hit1 = _best_fuzzy_hit(SRC_NORM, qn)
                        r2, hit2 = _best_fuzzy_hit(TRG_NORM, qn)
                        best_r, best_hit = (r1, hit1) if r1 >= r2 else (r2, hit2)

                        if best_r >= 0.94 and best_hit:
                            patched = _replace_first_quoted(b, q, best_hit)
                            new_bullets[idx] = patched
                            patched_any = True
                            chosen_norm = _norm(best_hit)
                            found_any = True
                            issues.append({
                                "dimension": dim_name,
                                "justification": patched,
                                "why": "auto-fixed minor quote typo"
                            })
                            break

                    if found_any and chosen_norm is not None:
                        if chosen_norm in seen_snips:
                            issues.append({
                                "dimension": dim_name,
                                "justification": new_bullets[idx],
                                "why": "Same snippet repeated within this dimension."
                            })
                        else:
                            seen_snips.add(chosen_norm)

                        if too_long_seen:
                            issues.append({
                                "dimension": dim_name,
                                "justification": new_bullets[idx],
                                "why": "One or more quoted snippets exceed 15 words."
                            })
                        continue

                    if saw_too_long_only and too_long_seen:
                        issues.append({
                            "dimension": dim_name,
                            "justification": b,
                            "why": "Quoted snippet longer than 15 words (must be ≤ 15)."
                        })
                        continue

                    example = quotes[0] if quotes else ""
                    short = (example[:40] + "…") if len(example) > 40 else example
                    issues.append({
                        "dimension": dim_name,
                        "justification": b,
                        "why": f'Quoted text not found verbatim in source or translation: "{short}"',
                        "exp": example
                    })

                if patched_any:
                    info["justification"] = new_bullets

            return issues

        def _llm_dedupe(result_dict: dict) -> dict:
            prompt = {"role": "user", "content":
                "Here is a PAEM-CMT evaluation JSON.\n\n"
                "TASK: For *each* dimension under \"dimensions\"\n"
                " • Keep MAX 5 bullets.\n"
                " • If two bullets express the SAME point, merge them into ONE.\n"
                " • Preserve the JSON schema exactly.\n"
                " • Preserve the original severity.\n"
                " • Do NOT intensify wording when merging.\n"
                " • Prefer the least speculative and most text-grounded phrasing.\n"
                " • If two phrasings are equally valid, choose the more neutral one.\n"
                " • Prefer bullets tied directly to brief fulfillment.\n"
                " • Drop bullets that complain only about source non-retention without explicit brief relevance.\n"
                " • Do NOT strengthen a source difference into a concern unless brief failure is explicit.\n"
                " • If a bullet is negative overall, rewrite it fully as a concern.\n"
                " • Remove leading praise, mitigation, or concessive framing from concern bullets.\n"
                " • Do NOT preserve 'X is good, but ...' or 'Although X works, ...' structures in merged concern bullets.\n\n"
                "Return ONLY valid JSON."
            }
            try:
                rsp = client.chat.completions.create(
                    model=EVAL_MODEL,
                    messages=[
                        {"role": "system", "content": "You are a meticulous editor of JSON evaluation data. Return valid JSON only."},
                        prompt,
                        {"role": "user", "content": json.dumps(result_dict, ensure_ascii=False)}
                    ],
                    max_completion_tokens=COMPLETION_CAP,
                    temperature=0.0,
                    response_format={"type": "json_object"},
                    **_gpt5_effort_none_kwargs(EVAL_MODEL),
                )
                print("[DEDUPE] response received.")
                deduped = (rsp.choices[0].message.content or "").strip()

                if deduped.startswith("```"):
                    deduped = re.sub(r"^```(?:json)?|```$", "", deduped, flags=re.S).strip()

                m = re.search(r"\{.*\}", deduped, re.S)
                if not m:
                    raise ValueError("No JSON object returned by dedupe helper.")

                return json.loads(m.group())
            except Exception as e:
                print(Fore.MAGENTA + f"[WARN] LLM dedupe skipped: {e}" + Style.RESET_ALL)
                return result_dict

        def _validate_main_eval_result(obj: dict) -> dict:
            if not isinstance(obj, dict):
                raise ValueError("Main eval output is not a JSON object.")

            dims = obj.get("dimensions")
            if not isinstance(dims, dict):
                raise ValueError("Main eval output is missing 'dimensions'.")

            required_dims = list(self.STANDARD_HEADINGS)
            missing_dims = [k for k in required_dims if k not in dims]
            extra_dims = [k for k in dims.keys() if k not in required_dims]

            if missing_dims:
                raise ValueError(f"Main eval output is missing dimensions: {missing_dims}")
            if extra_dims:
                raise ValueError(f"Main eval output has unexpected dimensions: {extra_dims}")

            for dim_name in required_dims:
                info = dims.get(dim_name)
                if not isinstance(info, dict):
                    raise ValueError(f"Dimension '{dim_name}' is not an object.")
                if "score" not in info or "justification" not in info:
                    raise ValueError(f"Dimension '{dim_name}' is missing 'score' or 'justification'.")
                score = float(info.get("score", 0.0))
                if score < 0.0 or score > 5.0:
                    raise ValueError(f"Dimension '{dim_name}' has out-of-range score: {score}")
                if not isinstance(info.get("justification"), list):
                    raise ValueError(f"Dimension '{dim_name}' justification is not a list.")

            subs = obj.get("subtotals")
            if not isinstance(subs, dict):
                raise ValueError("Main eval output is missing 'subtotals'.")
            for key in ("Role Satisfaction", "Terminology Adherence"):
                if key not in subs:
                    raise ValueError(f"Main eval output is missing subtotal '{key}'.")
                float(subs[key])

            if "overall" not in obj:
                raise ValueError("Main eval output is missing 'overall'.")
            float(obj["overall"])

            return obj

        def _parse_main_eval_result(raw_text: str) -> tuple[dict, str]:
            text = (raw_text or "").strip()
            if text.startswith("```"):
                text = re.sub(r"^```(?:json)?|```$", "", text, flags=re.S).strip()

            strict_err = None
            try:
                parsed = json.loads(text)
                parsed = _validate_main_eval_result(parsed)
                return parsed, "json_object_strict"
            except Exception as e:
                strict_err = e

            if not bool(getattr(self, "eval_parse_fallback_enabled", True)):
                raise ValueError(f"Strict main eval JSON parse failed: {strict_err}")

            m = re.search(r"\{.*\}", text, re.S)
            if not m:
                raise ValueError(f"Main eval JSON parse failed and no fallback object was found: {strict_err}")

            parsed = json.loads(m.group())
            parsed = _validate_main_eval_result(parsed)
            parsed["_eval_parse_warning"] = f"brace_fallback:{type(strict_err).__name__}"
            return parsed, "brace_fallback"

        def _score_once(model: str = EVAL_MODEL, temp: float = 0.5, top_p=None, seed: int = 42):
            """Single evaluation call."""
            nonlocal prompt_tok_total, completion_tok_total
            import random
            rng = random.Random(int(seed) ^ 0x9E3779B9)
            dbg_tag = f"seed={int(seed)}"
            msgs = [m.copy() for m in messages]
            try:
                u = json.loads(msgs[-1]["content"])
                meta = u.setdefault("meta", {})
                meta["run_seed"] = int(seed)
                meta["nonce"] = f"{seed}-{rng.randint(1000,9999)}"
                msgs[-1]["content"] = json.dumps(u, ensure_ascii=False)
            except Exception:
                msgs[-1]["content"] = (msgs[-1]["content"] + f"\n<!-- meta_run_id:{seed} -->")

            run_no = max(0, int(seed) - BASE_SEED)
            OFF_TEMP = (-0.12, -0.06, 0.00, 0.06, 0.12)
            OFF_TOPP = (-0.10, -0.05, 0.00, 0.05, 0.10)

            level = (0.75 ** (run_no // 10))

            if temp is None:
                j = OFF_TEMP[run_no % len(OFF_TEMP)] * TEMP_JITTER * level
                temp_run = max(0.0, min(1.0, EVAL_TEMP_BASE + j))
            else:
                temp_run = float(temp)

            if top_p is None:
                j = OFF_TOPP[run_no % len(OFF_TOPP)] * TOPP_JITTER * level
                top_p_run = max(0.05, min(1.0, TOP_P_BASE + j))
            else:
                top_p_run = float(top_p)

            raw = client.chat.completions.with_raw_response.create(
                model=model,
                messages=msgs,
                max_completion_tokens=MAX_COMPLETION,
                temperature=temp_run,
                top_p=top_p_run,
                response_format={"type": "json_object"},
                seed=(seed if getattr(self, "_seed_anchor", True) else None),
                **_gpt5_effort_none_kwargs(model),
            )
            print(f"[cand] {dbg_tag} temp={temp_run:.3f} top_p={top_p_run:.3f} (single-shot)", flush=True)

            headers = raw.http_response.headers
            chat = raw.parse()
            print("[EVAL] response received.")
            usage = {
                "prompt_tokens":      getattr(chat.usage, "prompt_tokens", 0),
                "completion_tokens":  getattr(chat.usage, "completion_tokens", 0),
                "total_tokens":       getattr(chat.usage, "total_tokens", 0),
            }
            prompt_tok_total     += usage.get("prompt_tokens", 0)
            completion_tok_total += usage.get("completion_tokens", 0)

            content = (chat.choices[0].message.content or "").strip()
            result, parse_mode = _parse_main_eval_result(content)
            result["_eval_parse_mode"] = parse_mode

            result = _llm_dedupe(result)

            det_issues = _deterministic_audit(result) if USE_DETERMINISTIC_AUDIT else []

            merged = list(det_issues or [])

            for it in merged:
                why = str(it.get("why", ""))
                if   why.startswith("Quoted text not found"):
                    it["code"] = "unfound_snippet"
                elif why.startswith("No quoted snippet"):
                    it["code"] = "no_snippet"
                elif ("longer than 15 words" in why) or ("exceed 15 words" in why):
                    it["code"] = "too_long"
                elif why.startswith("Same snippet repeated"):
                    it["code"] = "format_violation"
                elif why.startswith("auto-fixed"):
                    it["code"] = "format_violation"
                else:
                    it.setdefault("code", "format_violation")

            MAJOR_CODES = {"unfound_snippet", "no_snippet"}

            major_hits = sum(1 for it in merged if it.get("code") in MAJOR_CODES)

            hallucinated = (major_hits >= AUDIT_MAJOR_DROP_THRESHOLD)

            return result, usage, headers, hallucinated, merged, float(temp_run), float(top_p_run)

        print(Fore.CYAN + f"Running PAEM-CMT with repeated runs + internal stability (min {MIN_RUNS}, max {MAX_RUNS}, CI≤±{EPS_CI}).")

        try:
            start = time.monotonic()
            for attempt in range(3):
                ui_pulse(f"SEED_ANCHOR:{BASE_SEED}")
                anchor, usage, headers, halluc, issues, _, _ = _score_once(temp=0.0, top_p=TOP_P_BASE, seed=BASE_SEED)
                h_anchor = hashlib.sha1(json.dumps(anchor, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()[:8]
                ui_pulse("ANCHOR_HASH:" + h_anchor)
                print(Fore.CYAN + f"seed {BASE_SEED} | hash {h_anchor}" + Style.RESET_ALL)

                if halluc or issues:
                    ui_pulse("AUDIT:" + json.dumps({
                        "run": 1,
                        "anchor": True,
                        "discarded": bool(halluc),
                        "ungrounded": issues
                    }, ensure_ascii=False))

                if not halluc:
                    break
                print(Fore.RED + f"⚠️  Anchor attempt {attempt+1} discarded – audit failure(s)")
                ui_pulse("anchor retry due to audit…")
            if halluc:
                print(Fore.YELLOW + "⚠️  Audit kept flagging bullets – keeping first result anyway")

            _raw = float((anchor or {}).get("overall", 0.0))
            anchor = self._stabilize_run_numbers(anchor, anchor=None, prev=[])
            anchor["_raw_overall"] = _raw
            anchor = self._recompute_single_run_metric(anchor)

            elapsed = time.monotonic() - start
            stats.push(elapsed, usage, headers, run_idx=0, total_runs=MAX_RUNS, score=anchor["overall"])
            ui_pulse("RUN:" + json.dumps({
                "attempt": 1,
                "total":   MAX_RUNS,
                "elapsed": elapsed,
                "usage":   {"total_tokens": usage.get("total_tokens", 0)},
                "score":   float(anchor.get("overall", 0.0)),
                "temp":    0.0,
                "top_p":   float(TOP_P_BASE),
            }, ensure_ascii=False))

            all_scores.append(anchor)

            if ANCHOR_BAND_ENABLE:
                k = 1 if CI_SMOOTH_MODE == "raw" else (3 if CI_SMOOTH_MODE == "roll3" else 5)
                m_eff = max(3, ANCHOR_BAND_TARGET_RUNS // k)

                try:
                    from scipy.stats import t as _t
                    crit_band = _t.ppf(1 - ALPHA/2, df=m_eff - 1)
                except Exception:
                    crit_band = 1.96

                sd_max = EPS_CI * math.sqrt(m_eff) / max(crit_band, 1e-9)
                BAND_TOLERANCE = 1.35
                band_half = min(0.55, BAND_TOLERANCE * (3.0 ** 0.5) * sd_max)

                anchor_c = float(anchor.get("overall", 0.0))
                self._anchor_band = (anchor_c - band_half, anchor_c + band_half)
                print(Fore.MAGENTA + f"[band] anchor ±{band_half:.2f} → [{self._anchor_band[0]:.2f}, {self._anchor_band[1]:.2f}] for CI≤±{EPS_CI}")
                self._band_center = anchor_c
                self._band_consec_drops = 0

        except Exception as e:
            raise RuntimeError(f"Anchor run failed: {e}")

        for i in range(1, MAX_RUNS):
            try:
                start = time.monotonic()
                ui_pulse(f"SEED_RUN:{BASE_SEED + i}")
                if VARIANCE_MODE == "deterministic":
                    temp_arg, top_p_arg = EVAL_TEMP_BASE, TOP_P_BASE
                else:
                    anchor_c = float(all_scores[0].get("overall", 0.0)) if all_scores else 0.0
                    recent_all = [float(s.get("overall", 0.0)) for s in all_scores[-4:]]
                    recent = recent_all
                    mixture = False
                    if len(recent) >= 4:
                        cl, ce = _k2_split(recent)
                        if len(ce) == 2 and abs(ce[0] - ce[1]) >= MIXTURE_LOCK_DELTA and min(len(cl[0]), len(cl[1])) >= 2:
                            mixture = True
                    jump = (len(all_scores) >= 2 and abs(all_scores[-1]["overall"] - all_scores[-2]["overall"]) > 0.12)
                    drift = (len(all_scores) >= 1 and abs(all_scores[-1]["overall"] - anchor_c) >= (MIXTURE_LOCK_DELTA / 2.0))
                    med = statistics.median(recent) if recent else 0.0
                    mad = (statistics.median([abs(x - med) for x in recent]) if len(recent) >= 2 else 0.0)

                    if mixture or jump or drift or mad >= 0.05:
                        temp_arg, top_p_arg = EVAL_TEMP_BASE, TOP_P_BASE
                    else:
                        temp_arg, top_p_arg = None, None

                res, usage, headers, halluc, issues, temp_used, top_p_used = _score_once(
                    temp=temp_arg, top_p=top_p_arg, seed=BASE_SEED + i
                )

                if halluc or issues:
                    ui_pulse("AUDIT:" + json.dumps({
                        "run": i+1,
                        "discarded": bool(halluc),
                        "ungrounded": issues
                    }, ensure_ascii=False))
                if halluc:
                    print(Fore.RED + f"⚠️  run {i+1} discarded – {len(issues)} audit failure(s)")
                    ui_pulse("discarded run (audit)")
                    continue

                _raw = float((res or {}).get("overall", 0.0))
                res = self._stabilize_run_numbers(res, anchor=all_scores[0] if all_scores else None, prev=all_scores)
                res["_raw_overall"] = _raw
                res = self._recompute_single_run_metric(res)
                anchor_val = float(all_scores[0].get("overall", 0.0)) if all_scores else float(res.get("overall", 0.0))

                if ANCHOR_BAND_ENABLE and hasattr(self, "_anchor_band"):
                    base_lo, base_hi = self._anchor_band
                    base_half = 0.5 * (base_hi - base_lo)
                    center = getattr(self, "_band_center", (base_hi + base_lo) * 0.5)
                    kept_n = len(all_scores)

                    if not hasattr(self, "_recent_all"):
                        from collections import deque
                        self._recent_all = deque(maxlen=8)
                    self._recent_all.append(float(res.get("overall", 0.0)))

                    warm_bonus  = BAND_EXTRA_WARMUP if kept_n < BAND_WARMUP_KEEP else 0.0
                    spike_bonus = min(BAND_SPIKE_MAX, getattr(self, "_band_consec_drops", 0) * BAND_SPIKE_STEP)

                    half_dyn = base_half + warm_bonus + spike_bonus
                    half_dyn = max(half_dyn, getattr(self, "_band_half_prev", 0.0) * 0.90)

                    vals_all = list(self._recent_all)
                    if kept_n == 0 and len(vals_all) >= 3:
                        vals_all.sort()
                        mid = vals_all[len(vals_all)//2] if len(vals_all)%2 else 0.5*(vals_all[len(vals_all)//2-1]+vals_all[len(vals_all)//2])
                        q1  = vals_all[max(0, (len(vals_all)*1)//4)]
                        q3  = vals_all[min(len(vals_all)-1, (len(vals_all)*3)//4)]
                        iqr = max(1e-9, (q3 - q1))
                        center   = float(mid)
                        half_dyn = max(half_dyn, 0.5 * iqr)

                    if kept_n >= 2:
                        vals_k = [float(s.get("overall", 0.0)) for s in all_scores[-BAND_RECENTER_WINDOW:]]
                        if vals_k:
                            vals_k.sort()
                            med_k = vals_k[len(vals_k)//2] if (len(vals_k) % 2) else 0.5 * (vals_k[len(vals_k)//2 - 1] + vals_k[len(vals_k)//2])
                            target = max(base_lo, min(float(med_k), base_hi))
                            prev_center = center
                            blended = 0.70 * center + 0.30 * target
                            step_cap = 0.06
                            if abs(blended - prev_center) > step_cap:
                                blended = prev_center + (step_cap if blended > prev_center else -step_cap)
                            center = blended
                            self._band_center = center

                    center = max(base_lo, min(center, base_hi))

                    cand_lo, cand_hi = center - half_dyn, center + half_dyn
                    lo = max(cand_lo, base_lo - BAND_EPS)
                    hi = min(cand_hi, base_hi + BAND_EPS)

                    if lo > hi:
                        lo, hi = hi, lo
                    min_half = max(0.5 * base_half, 0.06)
                    mid = max(base_lo, min(center, base_hi))
                    if (hi - lo) < (2 * min_half):
                        lo = max(base_lo, mid - min_half)
                        hi = min(base_hi, mid + min_half)

                    val = float(res.get("overall", 0.0))
                    if (val < (lo - BAND_EPS)) or (val > (hi + BAND_EPS)):
                        why_txt = (
                            f"score {val:.2f} ∉ dyn[{lo:.2f}, {hi:.2f}] "
                            f"base[{base_lo:.2f}, {base_hi:.2f}]"
                        )
                        print(Fore.YELLOW + f"[skip] dropped run {i+1} outside anchor-band ({why_txt})")
                        ui_pulse("DISCARD:" + json.dumps({
                            "run": i+1,
                            "kind": "anchor-band",
                            "dimension": "Filter",
                            "justification": "Outside anchor band",
                            "score": val,
                            "band": [float(lo), float(hi)],
                            "band_base": [float(base_lo), float(base_hi)],
                            "why": why_txt
                        }, ensure_ascii=False))

                        self._band_consec_drops = getattr(self, "_band_consec_drops", 0) + 1
                        continue

                early_phase = len(all_scores) >= 2 and (i + 1) <= 8
                if EARLY_SKIP_ENABLE and early_phase and abs(res["overall"] - anchor_val) >= (MIXTURE_LOCK_DELTA * 0.9):
                    print(Fore.YELLOW + f"[skip] dropped outlier run {i+1} near start (Δ={abs(res['overall']-anchor_val):.2f} ≥ {MIXTURE_LOCK_DELTA*0.9:.2f})")
                    ui_pulse("discarded run (early outlier)")
                    continue

                elapsed = time.monotonic() - start
                stats.push(elapsed, usage, headers, run_idx=i, total_runs=MAX_RUNS, score=res["overall"])
                ui_pulse("RUN:" + json.dumps({
                    "attempt": i+1,
                    "total":   MAX_RUNS,
                    "elapsed": elapsed,
                    "usage":   {"total_tokens": usage.get("total_tokens", 0)},
                    "score":   float(res.get("overall")),
                    "temp":    float(temp_used),
                    "top_p":   float(top_p_used),
                }, ensure_ascii=False))

                all_scores.append(res)
                self._band_consec_drops = 0
                try:
                    cur_half = None
                    if 'hi' in locals() and 'lo' in locals():
                        cur_half = 0.5 * (hi - lo)
                    elif hasattr(self, "_anchor_band"):
                        _blo, _bhi = self._anchor_band
                        cur_half = 0.5 * (_bhi - _blo)
                    if cur_half is not None:
                        self._band_half_prev = max(getattr(self, "_band_half_prev", 0.0), cur_half)
                    self._band_center = (
                        (1.0 - BAND_CENTER_BLEND) * getattr(self, "_band_center", float(res["overall"]))
                        + BAND_CENTER_BLEND * float(res["overall"])
                    )
                except Exception:
                    pass

                h = hashlib.sha1(json.dumps(res, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()[:8]
                ui_pulse("RUN_HASH:" + h)
                print(Fore.CYAN + f"seed {BASE_SEED + i} | hash {h}" + Style.RESET_ALL)

            except Exception as e:
                print(Fore.RED + f"Run {i+1} failed: {e}")
                ui_pulse(f"run {i+1} error")
                continue

            if len(all_scores) < MIN_RUNS:
                continue

            def _arr_for_ci_from_all_scores(scores, mode):
                vals = [float(r.get("overall", 0.0)) for r in scores]
                if mode == "raw":
                    return vals
                k = 3 if mode == "roll3" else 5
                out = []
                for i in range(k, len(vals) + 1, k):
                    chunk = vals[i - k:i]
                    out.append(sorted(chunk)[k // 2])
                return out

            arr = _arr_for_ci_from_all_scores(all_scores, CI_SMOOTH_MODE)
            n   = len(arr)

            arr_ci   = list(arr)
            idx_trim = None
            if n >= 6:
                mu0  = statistics.median(arr)
                mad0 = statistics.median([abs(x - mu0) for x in arr]) or 0.0
                s0   = 1.4826 * mad0
                resid = [abs(x - mu0) for x in arr]
                j = max(range(n), key=lambda k: resid[k])
                rj = resid[j]
                thr_abs = 0.25
                thr_rob = 2.5 * s0 if s0 > 0 else float("inf")
                if rj > max(thr_abs, thr_rob) and resid.count(rj) == 1:
                    idx_trim = j
                    del arr_ci[j]

            n_ci = len(arr_ci)

            locked_by_mix = False
            if all_scores:
                anchor_c = float(all_scores[0].get("overall", 0.0))
                _seq_for_mix = list(arr_ci)
                if n_ci < 3:
                    vals_all = [float(r.get("overall", 0.0)) for r in all_scores]
                    _seq_for_mix = vals_all[-5:] if len(vals_all) >= 5 else vals_all[:]

            if len(_seq_for_mix) >= 3:
                clusters, centers = _k2_split(_seq_for_mix)
                if len(centers) == 2 and abs(centers[0] - centers[1]) >= MIXTURE_LOCK_DELTA:
                    j = 0 if abs(centers[0] - anchor_c) <= abs(centers[1] - anchor_c) else 1
                    keep = clusters[j]
                    if len(keep) >= 2:
                        arr_ci = keep
                        n_ci = len(arr_ci)
                        locked_by_mix = True
                        print(Fore.MAGENTA + f"[mix] locked CI to anchor-mode @ {centers[j]:.2f} (Δ={abs(centers[0]-centers[1]):.2f})")

            vals_all = [float(r.get("overall", 0.0)) for r in all_scores]
            fallback_k = 5 if CI_SMOOTH_MODE == "roll5" else (3 if CI_SMOOTH_MODE == "roll3" else 5)

            use_raw_fallback = False

            seq = arr_ci
            m = len(seq)

            seq = [max(0.0, min(5.0, float(x))) for x in seq]

            if m <= 1:
                mu = statistics.mean(seq)
                sd = 0.0
                half_ci = float("inf")
            else:
                if use_raw_fallback:
                    mu = statistics.mean(seq)
                    sd = statistics.stdev(seq)
                else:
                    if m >= 3:
                        mu, sd = _huber_mean(seq)
                    else:
                        mu = statistics.mean(seq)
                        sd = statistics.stdev(seq)
            try:
                from scipy.stats import t as _t
                crit = _t.ppf(1 - ALPHA/2, df=m - 1)
            except Exception:
                _TCRIT_95 = {
                    1: 12.706, 2: 4.303, 3: 3.182, 4: 2.776, 5: 2.571,
                    6: 2.447, 7: 2.365, 8: 2.306, 9: 2.262, 10: 2.228,
                    11: 2.201, 12: 2.179, 13: 2.160, 14: 2.145, 15: 2.131,
                    16: 2.120, 17: 2.110, 18: 2.101, 19: 2.093, 20: 2.086,
                    24: 2.064, 30: 2.042
                }
                crit = _TCRIT_95.get(m - 1, 1.96)

            half_ci = crit * sd / math.sqrt(m)

            guarded = False
            if sd < 1e-6 and (n_ci < 3 or len(all_scores) < max(10, 2 * fallback_k)):
                half_ci = EPS_CI + 1e-6
                guarded = True

            conf_pct = f"{(1-ALPHA)*100:.0f}"
            tag = []
            if idx_trim is not None: tag.append("trim1")
            if 'guarded' in locals() and guarded:
                GUARD_MIN_KEPT = 5
                tag.append(f"guard m={m}/{GUARD_MIN_KEPT}")
            tag = (" [" + ",".join(tag) + "]") if tag else ""

            _half = (f"{half_ci:.4f}" if math.isfinite(half_ci) else "∞")
            print(f"  → average {Fore.YELLOW}{mu:.2f}{Style.RESET_ALL}, stability ±{Fore.GREEN}{_half}{Style.RESET_ALL}{tag}")
            ui_pulse(f"average {mu:.2f}  stability ±{_half}{tag}")
            ui_pulse("ROLLING:" + json.dumps({
                "mu": float(mu),
                "ci": (float(half_ci) if math.isfinite(half_ci) else 1e9),
                "guard": bool(('guarded' in locals()) and guarded),
                "m": int(m),
                "run": len(all_scores),
            }, ensure_ascii=False))

            k = 1 if CI_SMOOTH_MODE == "raw" else (3 if CI_SMOOTH_MODE == "roll3" else 5)

            enough_cohorts = (
                (CI_SMOOTH_MODE == "raw")
                or (n_ci >= 3)
                or locked_by_mix
            )

            allow_converge = (m >= 3) and enough_cohorts and (not use_raw_fallback)
            if allow_converge and (half_ci < EPS_CI - 1e-9) and not (('guarded' in locals()) and guarded):

                attempts_fin = i + 1
                kept_fin     = len(all_scores)
                mu_fin = float(mu)
                sd_fin = float(sd)
                ci_fin = float(half_ci)

                scores_for_final = list(all_scores)

                print(Fore.BLUE + f"Stable enough at attempt {attempts_fin} (stability ±{half_ci:.4f} ≤ {EPS_CI}).")
                conf_pct = f"{(1-ALPHA)*100:.0f}"
                user_summary = _user_facing_result_summary([float(r.get("overall", 0.0)) for r in all_scores])

                print(
                    Style.BRIGHT + Fore.GREEN + f"\n✓ Finished after {attempts_fin} attempts ({kept_fin} kept)" +
                    Fore.YELLOW + f"\n   Best estimate: {user_summary['best_estimate']:.2f}" +
                    Fore.CYAN   + f"\n   Expected range: {user_summary['expected_range_label']}" +
                    Fore.MAGENTA+ f"\n   Checks: {user_summary['checks']}\n" + Style.RESET_ALL
                )

                final = self.aggregate_paemcmt_scores(scores_for_final) or {}
                if not isinstance(final, dict):
                    final = {}

                final.setdefault("dimensions", {})
                final.setdefault("subtotals", {})

                ow = float(final.get("overall", mu_fin))
                final["overall_weighted"] = ow
                final["overall_ci_mean"] = float(mu_fin)
                final["overall"] = ow
                final["n_runs"] = kept_fin
                final["n_kept"] = kept_fin
                final["n_attempts"] = attempts_fin
                final["std_internal_process"] = sd_fin
                final["ci"] = ci_fin
                final["alpha"] = ALPHA
                raw_stats = _ci_stats([float(r.get("overall", 0.0)) for r in all_scores])
                user_summary = _user_facing_result_summary([float(r.get("overall", 0.0)) for r in all_scores])
                final["best_estimate"] = float(user_summary["best_estimate"])
                final["expected_range_low"] = float(user_summary["expected_range_low"])
                final["expected_range_high"] = float(user_summary["expected_range_high"])
                final["expected_range_label"] = str(user_summary["expected_range_label"])
                final["checks"] = int(user_summary["checks"])
                final["user_range_method"] = str(user_summary["user_range_method"])
                final["ci_smooth_mode"] = CI_SMOOTH_MODE
                final["overall_ci_mean_smoothed"] = float(final.get("overall_ci_mean", final.get("overall", 0.0)))
                final["ci_smoothed"] = float(final.get("ci", 0.0))

                final["overall_ci_mean_raw"] = raw_stats["mean"]
                final["ci_raw"] = raw_stats["ci"]
                final["std_overall"] = raw_stats["sd"]
                final["std_overall_raw"] = raw_stats["sd"]
                final["audit_mode"] = AUDIT_MODE
                final["audit_major_drop_threshold"] = AUDIT_MAJOR_DROP_THRESHOLD
                parse_modes = sorted({str(r.get("_eval_parse_mode", "unknown")) for r in all_scores})
                final["eval_parse_mode"] = parse_modes[0] if len(parse_modes) == 1 else "mixed"
                final["eval_parse_modes"] = parse_modes
                ui_pulse(
                    f"best estimate {final['best_estimate']:.2f}  expected range {final['expected_range_label']}"
                )
                ui_pulse("finished ✅")
                return final

        kept_fin = len(all_scores)
        attempts_fin = MAX_RUNS
        if kept_fin == 0:
            raise RuntimeError("No successful runs.")

        arr_fin = [float(r.get("overall", 0.0)) for r in all_scores]
        nfin = kept_fin

        def _mad(seq):
            m = statistics.median(seq)
            return statistics.median([abs(x - m) for x in seq]) or 1e-9

        def _huber(seq, c=1.25, it=10):
            mu = statistics.median(seq)
            s  = 1.4826 * _mad(seq)
            if s < 1e-6:
                mu = statistics.mean(seq)
                sd = statistics.stdev(seq) if len(seq) > 1 else 0.0
                return mu, sd
            for _ in range(it):
                w = []
                for x in seq:
                    z = (x - mu) / s
                    w.append(1.0 if abs(z) <= c else (c / (abs(z) + 1e-9)))
                mu = sum(w_i * x_i for w_i, x_i in zip(w, seq)) / (sum(w) or 1e-9)
            var = sum(min((x - mu)**2, (c * s)**2) for x in seq) / max(len(seq) - 1, 1)
            return mu, (var ** 0.5)

        if nfin >= 3:
            mu_fin, sd_fin = _huber(arr_fin)
        else:
            mu_fin = statistics.mean(arr_fin)
            sd_fin = statistics.stdev(arr_fin) if nfin > 1 else 0.0

        try:
            from scipy.stats import t as _t
            crit_fin = _t.ppf(1 - ALPHA/2, df=nfin - 1) if nfin > 1 else float("inf")
        except Exception:
            _TCRIT_95 = {
                1: 12.706, 2: 4.303, 3: 3.182, 4: 2.776, 5: 2.571,
                6: 2.447, 7: 2.365, 8: 2.306, 9: 2.262, 10: 2.228,
                11: 2.201, 12: 2.179, 13: 2.160, 14: 2.145, 15: 2.131,
                16: 2.120, 17: 2.110, 18: 2.101, 19: 2.093, 20: 2.086,
                24: 2.064, 30: 2.042
            }
            crit_fin = _TCRIT_95.get(nfin - 1, 1.96) if nfin > 1 else float("inf")

        ci_fin = crit_fin * sd_fin / math.sqrt(nfin)

        print(Fore.YELLOW + f"Reached max runs ({MAX_RUNS}) without meeting the stability goal. Returning the final estimate.")

        final = self.aggregate_paemcmt_scores(all_scores) or {}
        if not isinstance(final, dict):
            final = {}
        final.setdefault("dimensions", {})
        final.setdefault("subtotals", {})
        final.setdefault("overall", float(mu_fin))

        ow = float(final.get("overall", mu_fin))
        final["overall_weighted"] = ow
        final["overall_ci_mean"] = float(mu_fin)
        final["overall"] = ow
        final["n_runs"] = kept_fin
        final["n_kept"] = kept_fin
        final["n_attempts"] = attempts_fin
        final["std_internal_process"] = sd_fin
        final["ci"] = ci_fin
        final["alpha"] = ALPHA
        raw_stats = _ci_stats([float(r.get("overall", 0.0)) for r in all_scores])
        user_summary = _user_facing_result_summary([float(r.get("overall", 0.0)) for r in all_scores])
        final["best_estimate"] = float(user_summary["best_estimate"])
        final["expected_range_low"] = float(user_summary["expected_range_low"])
        final["expected_range_high"] = float(user_summary["expected_range_high"])
        final["expected_range_label"] = str(user_summary["expected_range_label"])
        final["checks"] = int(user_summary["checks"])
        final["user_range_method"] = str(user_summary["user_range_method"])
        final["ci_smooth_mode"] = CI_SMOOTH_MODE
        final["overall_ci_mean_smoothed"] = float(final.get("overall_ci_mean", final.get("overall", 0.0)))
        final["ci_smoothed"] = float(final.get("ci", 0.0))

        final["overall_ci_mean_raw"] = raw_stats["mean"]
        final["ci_raw"] = raw_stats["ci"]
        final["std_overall"] = raw_stats["sd"]
        final["std_overall_raw"] = raw_stats["sd"]
        final["audit_mode"] = AUDIT_MODE
        final["audit_major_drop_threshold"] = AUDIT_MAJOR_DROP_THRESHOLD
        parse_modes = sorted({str(r.get("_eval_parse_mode", "unknown")) for r in all_scores})
        final["eval_parse_mode"] = parse_modes[0] if len(parse_modes) == 1 else "mixed"
        final["eval_parse_modes"] = parse_modes
        ui_pulse(
            f"best estimate {final['best_estimate']:.2f}  expected range {final['expected_range_label']}"
        )
        return final 

    def _flatten_dimensions(self, dimensions: dict, parent: str = "") -> dict:
        """Flatten nested dimensions into {path: leaf} where leaf has score and justification."""
        flat = {}
        for key, info in dimensions.items():
            name = f"{parent}.{key}" if parent else key
            if isinstance(info, dict):
                if "score" in info and "justification" in info:
                    flat[name] = info
                else:
                    flat.update(self._flatten_dimensions(info, parent=name))
        return flat

    def _merge_similar_bullets_llm(
        self,
        bullets,
        max_items: int = 5,
        model: str | None = None,
    ):
        bullets = [str(b).strip() for b in bullets if b and str(b).strip()]
        seen = set()
        bullets = [b for b in bullets if (b.lower() not in seen and not seen.add(b.lower()))]
        if len(bullets) <= 1:
            return bullets[:]

        try:
            rsp = client.chat.completions.create(
                model=model or self.model_name,
                temperature=0.0,
                max_completion_tokens=128000,
                **_gpt5_effort_none_kwargs(model or self.model_name),
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "You are a concise QA assistant.\n"
                            "Given a list of critique bullets:\n"
                            "• Combine bullets that refer to the same underlying issue.\n"
                            "• NEVER alter text inside double quotes.\n"
                            f"• Output at most {max_items} merged bullets, one per line, no numbering."
                        )
                    },
                    {
                        "role": "user",
                        "content": "\n".join(f"- {b}" for b in bullets)
                    }
                ],
            )
            raw = (rsp.choices[0].message.content or "").strip()

            import re
            merged = []
            for ln in raw.splitlines():
                t = ln.strip()
                if not t:
                    continue
                t = re.sub(r'^\s*(?:[-•*]|\d+[.)])\s*', '', t)
                if t:
                    merged.append(t)
            return merged[:max_items] or bullets[:max_items]

        except Exception as e:
            print(f"[WARN] LLM merge failed: {e}")
            return bullets[:max_items]

    def aggregate_paemcmt_scores(self, results):
        from collections import defaultdict
        from difflib import SequenceMatcher
        import re, math, unicodedata

        def _huber_mean(xs, c=1.5, iters=10):
            if not xs:
                return 0.0
            m = sum(xs) / len(xs)
            for _ in range(iters):
                diffs = [x - m for x in xs]
                w = []
                for d in diffs:
                    ad = abs(d)
                    w.append(1.0 if ad <= c else (c / ad))
                denom = sum(w) or 1.0
                m = sum(x * w_i for x, w_i in zip(xs, w)) / denom
            return m

        def _dedupe_fuzzy(seq, thresh=0.65):
            out = []
            for s in seq:
                if not any(SequenceMatcher(None, s, t).ratio() > thresh for t in out):
                    out.append(s)
            return out

        def _dedupe_by_quote(seq):
            out, seen = [], set()
            qre = re.compile(r'“([^”]+)”|\"([^"]+)\"|‘([^’]+)’|\'([^\'\n]+)\'|«([^»]+)»|‹([^›]+)›')
            for s in seq:
                m = qre.search(s or "")
                q = next((g for g in (
                    m.group(1) if m else None,
                    m.group(2) if m else None,
                    m.group(3) if m else None,
                    m.group(4) if m else None,
                    m.group(5) if m else None,
                    m.group(6) if m else None
                ) if g), "")
                key = (" ".join(q.split()).lower()) if q else None
                if key and key in seen:
                    continue
                if key:
                    seen.add(key)
                out.append(s)
            return out

        _dedupe = lambda seq: _dedupe_fuzzy(seq, thresh=0.65)[:5]

        dim_scores = defaultdict(list)
        dim_reasons_raw = defaultdict(list)
        subtotal_scores = defaultdict(list)

        for res in results:
            flat_dims = self._flatten_dimensions(res.get("dimensions", {}))

            for dim_name, info in flat_dims.items():
                dim_scores[dim_name].append(info.get("score", 0.0))

                just = info.get("justification", [])
                if isinstance(just, str):
                    just_list = [just]
                elif isinstance(just, list):
                    just_list = []
                    for j in just:
                        if isinstance(j, list):
                            just_list.extend(j)
                        else:
                            just_list.append(str(j))
                else:
                    just_list = [str(just)]

                dim_reasons_raw[dim_name].extend(just_list)

            for cat, val in res.get("subtotals", {}).items():
                subtotal_scores[cat].append(val)

        averaged_dims = {}
        for dim, vals in dim_scores.items():
            raw = list(dim_reasons_raw[dim])
            reasons = _dedupe_by_quote(raw)
            reasons = _dedupe(reasons)
            reasons = reasons[:5]
            averaged_dims[dim] = {
                "score": _huber_mean(vals, c=1.25),
                "reasons": reasons
            }

        averaged_subtotals = {
            cat: _huber_mean(vals, c=1.25)
            for cat, vals in subtotal_scores.items()
        }

        if not averaged_dims:
            last = next((r for r in reversed(results)
                         if isinstance(r, dict) and (r.get("dimensions") or {})), None)
            if last:
                averaged_dims = {
                    k: {
                        "score": float(((v or {}).get("score", 0.0))),
                        "reasons": (v or {}).get("justification", [])
                    }
                    for k, v in (last.get("dimensions") or {}).items()
                }
            else:
                averaged_dims = {k: {"score": 0.0, "reasons": []} for k in self.STANDARD_HEADINGS}

        dims_canon = {}
        for k in getattr(self, "STANDARD_HEADINGS", [
            "Intended Purpose", "Target Audience", "Translator",
            "Source Owner", "Commissioner", "Terminology Adherence"
        ]):
            if k in averaged_dims:
                dims_canon[k] = averaged_dims[k]
        for k, v in averaged_dims.items():
            if k not in dims_canon:
                dims_canon[k] = v
        averaged_dims = dims_canon

        max_score = 5.0
        role_score = averaged_subtotals.get("Role Satisfaction", 0.0)
        term_score = averaged_subtotals.get("Terminology Adherence", 0.0)

        def _norm(s):
            return unicodedata.normalize("NFKC", (s or "")).lower().strip()

        def _negatives(dim):
            rs = (averaged_dims.get(dim, {}) or {}).get("reasons", []) or []
            c = 0
            for r in rs:
                if isinstance(r, dict):
                    if str(r.get("polarity", "")).lower().startswith("neg"):
                        c += 1
                elif isinstance(r, str):
                    if "—" in r or " - " in r:
                        c += 1
            return c

        term_stats = self._compute_termlist_stats(getattr(self, "_last_translation_html", "") or "")
        required_pairs = term_stats["required_pairs"]
        total_terms = term_stats["total_terms"]
        misses = term_stats["misses"]
        miss_rate = term_stats["miss_rate"]
        term_score_det = term_stats["term_score"]
        catastrophic_term_violation = term_stats["catastrophic_term_violation"]
        pair_decisions = term_stats.get("pair_decisions", [])

        def _smooth_drop(neg_count: int, unit=0.45, k=0.85, cap=2.0):
            if neg_count <= 0:
                return 0.0
            base = (1 - math.exp(-k * neg_count)) / (1 - math.exp(-k))
            return min(cap, unit * neg_count * 0.35 + cap * 0.65 * base)

        role_dims = ["Target Audience", "Intended Purpose", "Translator", "Source Owner", "Commissioner"]

        for dim_name, dim in list(averaged_dims.items()):
            s = float(dim.get("score", 0.0))

            if dim_name == "Terminology Adherence":
                if term_score_det is not None:
                    s = float(term_score_det)

            averaged_dims[dim_name]["score"] = s

        if "Terminology Adherence" in averaged_dims:
            req_targets = [tgt for _, tgt in required_pairs]

            def _valid_term_strength(bullet: str) -> bool:
                b = str(bullet or "")
                if not b.lstrip().startswith("✓"):
                    return True
                return any(self._term_realized_in_text(tgt, b) for tgt in req_targets)

            averaged_dims["Terminology Adherence"]["reasons"] = [
                b for b in (averaged_dims["Terminology Adherence"].get("reasons") or [])
                if _valid_term_strength(b)
            ]
            averaged_dims["Terminology Adherence"]["justification"] = averaged_dims["Terminology Adherence"]["reasons"][:]

            term_dim_score = float((averaged_dims.get("Terminology Adherence") or {}).get("score", 0.0))
            if term_dim_score <= 1e-9:
                averaged_dims["Terminology Adherence"]["reasons"] = [
                    b for b in (averaged_dims["Terminology Adherence"].get("reasons") or [])
                    if not str(b or "").lstrip().startswith("✓")
                ]
                averaged_dims["Terminology Adherence"]["justification"] = averaged_dims["Terminology Adherence"]["reasons"][:]

        for dim_name in averaged_dims:
            averaged_dims[dim_name]["justification"] = list((averaged_dims[dim_name] or {}).get("reasons") or [])

        role_vals = [averaged_dims[d]["score"] for d in role_dims if d in averaged_dims]
        role_score = sum(role_vals) / len(role_vals) if role_vals else 0.0

        term_score = float(averaged_dims.get("Terminology Adherence", {}).get(
            "score",
            averaged_subtotals.get("Terminology Adherence", 0.0)
        ))
        averaged_subtotals["Terminology Adherence"] = term_score

        averaged_subtotals["Role Satisfaction"] = role_score

        shown_scores = [
            float((averaged_dims.get(d, {}) or {}).get("score", 0.0))
            for d in self.STANDARD_HEADINGS
            if d in averaged_dims
        ]
        overall_base = (sum(shown_scores) / len(shown_scores)) if shown_scores else 0.0

        overall = overall_base

        averaged_dims = self._freeze_scored_reasons(averaged_dims)
        return {
            "dimensions": averaged_dims,
            "subtotals": averaged_subtotals,
            "overall": overall,
            "term_pair_decisions": pair_decisions,
            "term_logic_version": getattr(self, "term_logic_version", "obligation_v3_relaxed_matcher"),
            "term_adjudication_status": term_stats.get("adjudication_status", "not_applicable"),
        }

    def evaluation_live_screen(self, total_runs: int):
        """Live evaluation view with runs table and chart."""
        self.initUI_header_only()
        if hasattr(self, "stage_chip"): self.stage_chip.hide()
        if hasattr(self, "progress_chip"): self.progress_chip.show()

        self._audit_by_run = {}
        self._first_result_seen = False
        self._last_live_stability_signature = None
        root = QWidget(self)
        rl = QVBoxLayout(root); rl.setContentsMargins(0,0,0,0); rl.setSpacing(10)

        title = QLabel("🚀 Live Evaluation", self)
        title.setStyleSheet("font-size:22px; font-weight:800; color:#cfe2ff;")
        rl.addWidget(title, 0)

        self.eval_progress = QProgressBar(self)
        self.eval_progress.setRange(0, max(1, int(total_runs)))
        self.eval_progress.setValue(0)
        self.eval_progress.setTextVisible(False)
        self.eval_progress.setFixedHeight(10)
        self.eval_progress.setStyleSheet("""
            QProgressBar { background:#141a26; border:1px solid #2b3d5c; border-radius:6px; }
            QProgressBar::chunk { background:#6ea8ff; border-radius:6px; }
        """)
        rl.addWidget(self.eval_progress, 0)

        runs_box = QGroupBox("📈 Runs", self)
        vb_runs = QVBoxLayout(runs_box); vb_runs.setContentsMargins(10,10,10,10)
        self._init_runs_table()

        self.waiting_hint = QLabel("⏳ Waiting for first results…", self)
        self.waiting_hint.setAlignment(Qt.AlignCenter)
        self.waiting_hint.setStyleSheet("color:#a9bfdc; font-weight:700; padding:10px;")
        vb_runs.addWidget(self.waiting_hint)
        vb_runs.addWidget(self.runs_table)
        self._add_shadow(runs_box)

        chart_box = QGroupBox("📊 Live Score Pattern", self)
        vb_chart = QVBoxLayout(chart_box); vb_chart.setContentsMargins(10,10,10,10)
        self.live_chart = LiveChart(self, threshold=getattr(self, "_ci_target", 0.05), active_after=10)
        vb_chart.addWidget(self.live_chart, 1)
        self._add_shadow(chart_box)

        split = QSplitter(Qt.Horizontal, self)
        split.setHandleWidth(4)
        split.addWidget(runs_box)
        split.addWidget(chart_box)
        split.setStretchFactor(0, 3)
        split.setStretchFactor(1, 2)
        rl.addWidget(split, 1)

        self.layout.addWidget(self._make_card("🧪 Evaluation", root), 1)

        self._final_summary_holder = QWidget(self)
        self._final_summary_holder.setLayout(QVBoxLayout())
        self._final_summary_holder.layout().setContentsMargins(0,0,0,0)
        self._final_summary_holder.layout().setSpacing(0)
        self.layout.addWidget(self._final_summary_holder, 0)

    def _handle_anchor_discard(self, items: list, attempt_no: int):
        """Reset first-kept-run UI state after an anchor discard."""
        try:
            if hasattr(self, "waiting_hint") and self.waiting_hint:
                self.waiting_hint.setText("⏳ anchor was discarded — retrying…")
                self.waiting_hint.show()
        except Exception:
            pass

        try:
            self._success_count = 0
            self._last_seen_run = 0
            self._attempt_to_run = {}
            self._run_export = {}
            self._last_live_stability_signature = None
        except Exception:
            pass

        try:
            r = self._run_rows.pop(1, None)
            if r is not None and self._has_widget("runs_table"):
                self.runs_table.removeRow(r)
                self._run_rows = {}
                for rr in range(self.runs_table.rowCount()):
                    it = self.runs_table.item(rr, 0)
                    if it and (it.text() or "").strip().isdigit():
                        self._run_rows[int(it.text().strip())] = rr
        except Exception:
            pass

        try:
            if hasattr(self, "live_chart") and self.live_chart:
                self.live_chart.scores[:] = []
                self.live_chart.mu[:] = []
                self.live_chart.ci[:] = []
                self.live_chart.update()
        except Exception:
            pass

        try:
            if hasattr(self, "eval_progress") and self.eval_progress:
                self.eval_progress.setValue(0)
        except Exception:
            pass

        self._safe_set_chip(self.progress_chip, "anchor discarded — retrying…")

    def _classify_bullets_api(self, data: dict) -> dict:
        """
        Label each bullet as 'pos' or 'neg'.
        Returns {dimension: ["pos"/"neg", ...]} in the same order as the input bullets.
        """
        import json, re
        dims = data.get("dimensions", {}) or {}
        payload = {k: (v.get("reasons") or []) for k, v in dims.items()}

        if not any(payload.values()):
            return {}

        try:
            prompt_rules = (
                "You will receive a JSON object mapping dimension name → list of bullets.\n"
                "TASK: For each bullet, output 'pos' if it is a strength/positive observation,\n"
                "or 'neg' if it is a weakness/problem.\n\n"
                "HARD RULES:\n"
                "• Do not rewrite or correct bullets; DO NOT CHANGE TEXT.\n"
                "• Keep the same order. Output JSON ONLY with the same keys.\n"
                "• For each dimension, return a list of strings ('pos'/'neg') with\n"
                "  exactly the same length as the input list for that dimension.\n"
                "• No explanations, no extra keys."
            )

            rsp = client.chat.completions.create(
                model=self.model_name,
                temperature=0.0,
                max_completion_tokens=128000,
                **_gpt5_effort_none_kwargs(self.model_name),
                messages=[
                    {"role": "system", "content": "You label only; never rewrite content. Return JSON only."},
                    {"role": "user",   "content": prompt_rules},
                    {"role": "user",   "content": json.dumps(payload, ensure_ascii=False)}
                ],
                timeout=60,
            )
            raw = (rsp.choices[0].message.content or "").strip()
            if raw.startswith("```"):
                raw = re.sub(r"^```(?:json)?|```$", "", raw, flags=re.S).strip()
            labels = json.loads(raw)

            out = {}
            for dim, bullets in payload.items():
                lab = labels.get(dim, [])
                if isinstance(lab, list) and len(lab) == len(bullets) and all(x in ("pos","neg") for x in lab):
                    out[dim] = lab
            return out
        except Exception as e:
            print(f"[Classify] labeling failed; using fallback: {e}")
            return {}

    def _wrap_as_paged(self, html: str) -> str:
        css = """
        <style>
          html,body{margin:0;padding:0;background:#0f1115}
          .sheet{width:794px; margin:18px auto; background:#fff; color:#111; 
                 box-shadow:0 6px 24px rgba(0,0,0,.35); padding:40px 52px; 
                 line-height:1.45; font:16px/1.45 Segoe UI,Inter,Helvetica,Arial,sans-serif;}
          .sheet h1,.sheet h2,.sheet h3{page-break-after:avoid}
          .page-break{break-before:page}
          @media print{
            body{background:#fff}
            .sheet{box-shadow:none; margin:0; width:auto; padding:0}
          }
        </style>
        """
        if "</head>" in html.lower():
            return re.sub(r"</head>", css + "</head>", html, flags=re.I)
        if "<body" in html.lower():
            return css + html
        return f"<!doctype html><html><head>{css}</head><body><div class='sheet'>{html}</div></body></html>"

    def _confirm_restart_app(self):
        """Ask for confirmation and optionally clear the API key before restart."""
        msg = QMessageBox(self)
        msg.setWindowTitle("Restart PAEM-CMT")
        msg.setText(
            "This will clear the current session (files, context, results) "
            "and return to the first screen."
        )
        msg.setInformativeText("Do you want to continue?")
        msg.setIcon(QMessageBox.Warning)
        msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)

        cb = QCheckBox("Also clear API key")
        msg.setCheckBox(cb)

        if msg.exec_() == QMessageBox.Yes:
            self.restart_app(clear_api_key=cb.isChecked())

    def restart_app(self, clear_api_key: bool = False):
        """Reset session state and return to the landing screen."""
        try:
            self._teardown_live_screen()
        except Exception:
            pass
        try:
            if hasattr(self, "_elapsed_timer") and self._elapsed_timer:
                self._elapsed_timer.stop()
        except Exception:
            pass
        self._elapsed_only_mode = False

        for tname in ("_eval_thread", "_qf_thread", "_purp_thread", "_label_thread"):
            th = getattr(self, tname, None)
            try:
                if th and hasattr(th, "isRunning") and th.isRunning():
                    th.quit()
                    th.wait(1500)
            except Exception:
                pass
            try:
                setattr(self, tname, None)
            except Exception:
                pass

        self.source_language = ""
        self.target_language = ""
        self.model_name = "gpt-5.4-2026-03-05"
        self.role_weight = 0.5
        self.term_weight = 0.5
        self.term_pairs = []
        self._context_state = {
            "purpose": "",
            "audience": "",
            "translator": "",
            "owner": "",
            "commissioner": "",
            "termpairs": [],
        }
        self._term_obligation_cache = {}
        self._last_live_stability_signature = None
        self.term_logic_version = "obligation_v3_relaxed_matcher"
        self.eval_parse_fallback_enabled = False
        self.audit_mode = "deterministic_v1"
        self.audit_major_drop_threshold = 1
        for attr, val in (
            ("file_path", ""),
            ("docx_content", ""),
            ("_last_translation_html", ""),
            ("translation_instructions", ""),
            ("translation_context", {}),
            ("_last_eval_data", {}),
            ("_bullet_labels", {}),
            ("_audit_by_run", {}),
            ("_audit_by_row", {}),
            ("_attempt_to_run", {}),
        ):
            try:
                setattr(self, attr, val)
            except Exception:
                pass
        self._success_count = 0
        self._skip_translation = False

        try:
            if hasattr(self, "console") and self.console:
                self.console.clear()
        except Exception:
            pass
        try:
            self._safe_set_chip(self.progress_chip, "idle")
            self._update_stage_chip("idle", "Ready")
        except Exception:
            pass

        if clear_api_key:
            try:
                set_api_key("")
                globals()["client"] = None
                self._refresh_api_key_button_style()
                print("[Restart] API key cleared.")
            except Exception:
                pass

        self.clear_layout(keep_header=True)
        self.landing_screen()
        self._show_toast("Session reset.")

    def _get_term_audit_payload(self, data=None):
        payload = data if isinstance(data, dict) else {}

        rows = list(
            payload.get("term_pair_decisions")
            or payload.get("_term_pair_decisions")
            or []
        )

        logic_version = str(
            payload.get("term_logic_version")
            or payload.get("_term_logic_version")
            or getattr(self, "term_logic_version", "obligation_v3_relaxed_matcher")
        )

        adjudication_status = str(
            payload.get("term_adjudication_status")
            or payload.get("_term_adjudication_status")
            or "not_applicable"
        )

        if not rows:
            try:
                term_stats = self._compute_termlist_stats(getattr(self, "_last_translation_html", "") or "")
                rows = list(term_stats.get("pair_decisions") or [])
                logic_version = str(term_stats.get("logic_version") or logic_version)
                adjudication_status = str(term_stats.get("adjudication_status") or adjudication_status)
            except Exception:
                rows = []

        norm_rows = []
        for row in rows:
            if not isinstance(row, dict):
                continue

            src_occ = int(row.get("source_occurrences", 0) or 0)
            exp_occ = int(row.get("explicit_target_occurrences", 0) or 0)
            rel_exp = int(row.get("relevant_explicit", 0) or 0)
            hidden = int(row.get("hidden_fulfilled", 0) or 0)
            fulfilled = int(row.get("fulfilled", max(0, min(src_occ, rel_exp + hidden))) or 0)
            missed = int(row.get("missed", max(0, src_occ - fulfilled)) or 0)

            norm_rows.append({
                "source_term": str(row.get("source_term", "") or ""),
                "target_term": str(row.get("target_term", "") or ""),
                "source_occurrences": src_occ,
                "explicit_target_occurrences": exp_occ,
                "relevant_explicit": rel_exp,
                "hidden_fulfilled": hidden,
                "fulfilled": fulfilled,
                "missed": missed,
                "alternative_used": bool(row.get("alternative_used", False)),
                "status": str(row.get("status", "not_applicable") or "not_applicable"),
                "adjudication_status": str(row.get("adjudication_status", adjudication_status) or adjudication_status),
                "reason": str(row.get("reason", "") or "").strip(),
                "logic_version": logic_version,
            })

        active_rows = [r for r in norm_rows if r["source_occurrences"] > 0]
        total = sum(r["source_occurrences"] for r in active_rows)
        fulfilled = sum(r["fulfilled"] for r in active_rows)
        missed = max(0, total - fulfilled)
        catastrophic = bool(total > 0 and missed == total)

        violated_rows = sorted(
            [r for r in active_rows if r["missed"] > 0],
            key=lambda r: (r["alternative_used"], r["missed"], r["source_occurrences"]),
            reverse=True,
        )

        return {
            "rows": active_rows,
            "logic_version": logic_version,
            "adjudication_status": adjudication_status,
            "total_obligations": total,
            "fulfilled": fulfilled,
            "missed": missed,
            "catastrophic": catastrophic,
            "violated_rows": violated_rows,
        }

    def export_live_snapshot(self):
        try:
            if not hasattr(self, "runs_table") or not hasattr(self, "live_chart"):
                QMessageBox.information(self, "Not available", "Live view not active.")
                return

            fn, _ = QFileDialog.getSaveFileName(
                self, "Export Live", "",
                "Excel Workbook (*.xlsx);;PDF (*.pdf);;PNG (*.png)"
            )
            if not fn:
                return

            if fn.lower().endswith(".xlsx"):
                if Workbook is None:
                    QMessageBox.warning(
                        self, "Missing dependency",
                        "Install openpyxl to export XLSX:\n\n    pip install openpyxl"
                    )
                    return

                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.chart import LineChart, Reference
                from openpyxl.chart.axis import ChartLines

                C_TEXT    = "1C274C"
                C_SUB     = "4B5563"
                C_HEAD    = "F2F4F8"
                C_GRID    = "D9DEE8"
                C_ZEBRA   = "FAFBFD"
                C_ANCHOR  = "E8F0FE"
                C_KEPT    = "EAF7EE"
                C_DISCARD = "FDECEA"
                C_WHITE   = "FFFFFF"

                def _fill(hex6): return PatternFill("solid", fgColor=hex6)
                def _thin_border(color=C_GRID):
                    s = Side(style="thin", color=color)
                    return Border(left=s, right=s, top=s, bottom=s)

                wb = Workbook()

                ws = wb.active
                ws.title = "Runs"
                headers = ["Run","Status","Score","Tokens","Elapsed","Temp","Top-p","Average","Stability"]
                ws.append(headers)

                for r in range(self.runs_table.rowCount()):
                    row = []
                    for c in range(self.runs_table.columnCount()):
                        it = self.runs_table.item(r, c)
                        row.append("" if it is None else it.text())
                    ws.append(row)

                for r in range(2, ws.max_row + 1):
                    run_v = str(ws.cell(r, 1).value or "").strip()
                    if not run_v.isdigit():
                        continue

                    raw = (getattr(self, "_run_export", {}) or {}).get(int(run_v), {})
                    if "score" in raw:
                        ws.cell(r, 3, float(raw["score"]))
                    if "mu" in raw:
                        ws.cell(r, 8, float(raw["mu"]))
                    if "ci" in raw and bool(raw.get("show_ci", True)):
                        ws.cell(r, 9, float(raw["ci"]))

                for c, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=c, value=h)
                    cell.font = Font(b=True, color=C_TEXT)
                    cell.fill = _fill(C_HEAD)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = _thin_border()

                widths = [8,12,8,8,12,8,8,8,10]
                for i, w in enumerate(widths, start=1):
                    ws.column_dimensions[chr(64+i)].width = w
                ws.freeze_panes = "A2"

                for r in range(2, ws.max_row+1):
                    status = (ws.cell(r, 2).value or "").lower()
                    base = _fill(C_ZEBRA if r % 2 else C_WHITE)
                    row_fill = base
                    if "anchor" in status:   row_fill = _fill(C_ANCHOR)
                    elif "kept" in status:   row_fill = _fill(C_KEPT)
                    elif "discard" in status:row_fill = _fill(C_DISCARD)

                    for c in range(1, ws.max_column+1):
                        cell = ws.cell(r, c)
                        cell.fill = row_fill
                        cell.border = _thin_border()
                        if c in (1,3,4,5,6,7,8,9):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.font = Font(color=C_TEXT)

                    ws.cell(r, 8).font = Font(b=True, color=C_TEXT)
                    ws.cell(r, 9).font = Font(b=True, color=C_TEXT)

                for r in range(2, ws.max_row + 1):
                    for c, fmt in ((3, "0.0000"), (6, "0.000"), (7, "0.000"), (8, "0.0000"), (9, "0.0000")):
                        try:
                            v = float(ws.cell(r, c).value)
                            ws.cell(r, c, v).number_format = fmt
                        except Exception:
                            pass

                tr = wb.create_sheet("Trend")
                tr.append(["Index","Score","Average","Stability","Average+Stability","Average−Stability",""])

                def _f(v):
                    try:
                        if v is None:
                            return None
                        s = str(v).strip()
                        if not s or s == "—":
                            return None
                        return float(s.replace(",", "."))
                    except Exception:
                        return None

                for r in range(2, ws.max_row + 1):
                    run_v = (ws.cell(r, 1).value or "").strip()
                    status = (ws.cell(r, 2).value or "").strip().lower()

                    if not run_v.isdigit():
                        continue
                    if "discard" in status:
                        continue

                    idx = int(run_v)
                    sc  = _f(ws.cell(r, 3).value)
                    mu  = _f(ws.cell(r, 8).value)
                    ci  = _f(ws.cell(r, 9).value)

                    up = (mu + ci) if (mu is not None and ci is not None) else None
                    lo = (mu - ci) if (mu is not None and ci is not None) else None

                    tr.append([idx, sc, mu, ci, up, lo, None])

                for c in range(1, 7):
                    cell = tr.cell(1, c)
                    cell.font = Font(b=True, color=C_TEXT)
                    cell.fill = _fill(C_HEAD)
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = _thin_border()

                for r in range(2, tr.max_row+1):
                    for c in range(1, 7):
                        tr.cell(r,c).border = _thin_border()
                        tr.cell(r,c).font = Font(color=C_TEXT)
                    if tr.cell(r,2).value is not None: tr.cell(r,2).number_format = "0.0000"
                    if tr.cell(r,3).value is not None: tr.cell(r,3).number_format = "0.0000"
                    if tr.cell(r,4).value is not None: tr.cell(r,4).number_format = "0.0000"
                    if tr.cell(r,5).value is not None: tr.cell(r,5).number_format = "0.0000"
                    if tr.cell(r,6).value is not None: tr.cell(r,6).number_format = "0.0000"

                tr.freeze_panes = "A2"
                tr.column_dimensions["A"].width = 8
                tr.column_dimensions["B"].width = 8
                tr.column_dimensions["C"].width = 8
                tr.column_dimensions["D"].width = 10
                tr.column_dimensions["E"].width = 10
                tr.column_dimensions["F"].width = 10

                target = float(getattr(self, "_ci_target", getattr(self, "ci_target", 0.05)))
                tr["H1"] = "Stability goal"
                tr["H2"] = target
                tr["H1"].font = Font(b=True, color=C_TEXT)
                tr["H1"].fill = _fill(C_HEAD)
                tr["H1"].alignment = Alignment(horizontal="center")
                tr["H1"].border = _thin_border()
                tr["H2"].number_format = "0.0000"
                tr["H2"].border = _thin_border()

                if tr.max_row >= 3:
                    rows = tr.max_row
                    Nrows = rows - 1

                    chart = LineChart()
                    chart.title = "PAEM-CMT Score Pattern"

                    vals = []
                    for r in range(2, tr.max_row+1):
                        for c in (2,3,5,6):
                            v = tr.cell(r, c).value
                            try:
                                vals.append(float(v))
                            except (TypeError, ValueError):
                                pass
                    lo, hi = 0.0, 5.0
                    if vals:
                        lo = max(0.0, min(vals) - 0.10)
                        hi = min(5.0, max(vals) + 0.10)
                        if hi - lo < 0.50:
                            pad = (0.50 - (hi - lo)) / 2
                            lo = max(0.0, lo - pad)
                            hi = min(5.0, hi + pad)

                    chart.y_axis.title = "Score (0–5)"
                    chart.y_axis.scaling.min = lo
                    chart.y_axis.scaling.max = hi
                    chart.y_axis.majorUnit = 0.10 if (hi - lo) <= 1.0 else 0.25
                    chart.y_axis.minorUnit = 0.05 if (hi - lo) <= 1.0 else 0.10
                    chart.y_axis.number_format = "0.00"
                    chart.y_axis.axPos = "l"
                    chart.y_axis.tickLblPos = "nextTo"
                    chart.y_axis.majorGridlines = ChartLines()
                    chart.y_axis.minorGridlines = ChartLines()

                    Nrows = tr.max_row - 1
                    chart.x_axis.title = "Run"
                    chart.x_axis.tickLblPos = "low"
                    chart.x_axis.majorGridlines = ChartLines()
                    if Nrows >= 50:
                        chart.x_axis.tickLblSkip = 4
                    elif Nrows >= 30:
                        chart.x_axis.tickLblSkip = 2

                    chart.legend.position = "r"
                    chart.legend.overlay = False
                    chart.style = 2

                    cats   = Reference(tr, min_col=1, min_row=2, max_row=tr.max_row)
                    s_score= Reference(tr, min_col=2, min_row=1, max_row=tr.max_row)
                    s_mu   = Reference(tr, min_col=3, min_row=1, max_row=tr.max_row)
                    s_up   = Reference(tr, min_col=5, min_row=1, max_row=tr.max_row)
                    s_lo   = Reference(tr, min_col=6, min_row=1, max_row=tr.max_row)

                    chart.add_data(s_score, titles_from_data=True)
                    chart.add_data(s_mu,    titles_from_data=True)
                    chart.add_data(s_up,    titles_from_data=True)
                    chart.add_data(s_lo,    titles_from_data=True)
                    chart.set_categories(cats)

                    chart.height = 18
                    chart.width  = 34

                    tr.add_chart(chart, "J2")

                ds = wb.create_sheet("Discarded")
                ds.append(["Run", "Dimension", "Bullet", "Why"])

                discarded_rows = 0

                for run_no, items in (getattr(self, "_audit_by_run", {}) or {}).items():
                    for it in items or []:
                        ds.append([run_no, it.get("dimension",""), it.get("justification",""), it.get("why","")])
                        discarded_rows += 1

                for _, items in (getattr(self, "_audit_by_row", {}) or {}).items():
                    for it in items or []:
                        ds.append(["—", it.get("dimension",""), it.get("justification",""), it.get("why","")])
                        discarded_rows += 1

                for c in range(1, 5):
                    cell = ds.cell(1, c)
                    cell.font = Font(b=True, color=C_TEXT)
                    cell.fill = _fill(C_HEAD)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = _thin_border()

                ds.freeze_panes = "A2"
                ds.column_dimensions["A"].width = 8
                ds.column_dimensions["B"].width = 20
                ds.column_dimensions["C"].width = 60
                ds.column_dimensions["D"].width = 100

                for r in range(2, ds.max_row+1):
                    base_fill = _fill(C_ZEBRA if r % 2 else C_WHITE)
                    for c in range(1, 5):
                        cell = ds.cell(r, c)
                        cell.fill = base_fill
                        cell.font = Font(color=C_TEXT)
                        cell.border = _thin_border()
                        wrap = c in (3,4)
                        cell.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal="left")
                    ds.cell(r, 1).fill = _fill(C_DISCARD)

                term_payload = self._get_term_audit_payload(getattr(self, "_last_eval_data", {}) or {})

                ta = None
                for s in wb.sheetnames:
                    if s.lower() == "terminology_audit":
                        ta = wb[s]
                if ta is None:
                    ta = wb.create_sheet("Terminology_Audit")

                ta.delete_rows(1, ta.max_row)

                ta_headers = [
                    "source_term",
                    "target_term",
                    "source_occurrences",
                    "explicit_target_occurrences",
                    "relevant_explicit",
                    "hidden_fulfilled",
                    "fulfilled",
                    "missed",
                    "alternative_used",
                    "status",
                    "adjudication_status",
                    "reason",
                    "logic_version",
                ]
                ta.append(ta_headers)

                for row in term_payload.get("rows", []):
                    ta.append([
                        row.get("source_term", ""),
                        row.get("target_term", ""),
                        int(row.get("source_occurrences", 0) or 0),
                        int(row.get("explicit_target_occurrences", 0) or 0),
                        int(row.get("relevant_explicit", 0) or 0),
                        int(row.get("hidden_fulfilled", 0) or 0),
                        int(row.get("fulfilled", 0) or 0),
                        int(row.get("missed", 0) or 0),
                        "Yes" if row.get("alternative_used", False) else "No",
                        row.get("status", ""),
                        row.get("adjudication_status", term_payload.get("adjudication_status", "")),
                        row.get("reason", ""),
                        row.get("logic_version", term_payload.get("logic_version", "")),
                    ])

                for c in range(1, len(ta_headers) + 1):
                    cell = ta.cell(1, c)
                    cell.font = Font(b=True, color=C_TEXT)
                    cell.fill = _fill(C_HEAD)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = _thin_border()

                ta.freeze_panes = "A2"
                ta.auto_filter.ref = f"A1:L{max(2, ta.max_row)}"

                ta.column_dimensions["A"].width = 24
                ta.column_dimensions["B"].width = 24
                ta.column_dimensions["C"].width = 12
                ta.column_dimensions["D"].width = 14
                ta.column_dimensions["E"].width = 12
                ta.column_dimensions["F"].width = 12
                ta.column_dimensions["G"].width = 10
                ta.column_dimensions["H"].width = 10
                ta.column_dimensions["I"].width = 12
                ta.column_dimensions["J"].width = 14
                ta.column_dimensions["K"].width = 70
                ta.column_dimensions["L"].width = 18

                for r in range(2, ta.max_row + 1):
                    hidden_fulfilled = int(ta.cell(r, 6).value or 0)
                    status = str(ta.cell(r, 10).value or "").strip().lower()

                    base_fill = _fill(C_ZEBRA if r % 2 else C_WHITE)
                    row_fill = base_fill
                    if "violated" in status:
                        row_fill = _fill(C_DISCARD)
                    elif hidden_fulfilled > 0:
                        row_fill = _fill(C_ANCHOR)
                    else:
                        row_fill = _fill(C_KEPT)

                    for c in range(1, len(ta_headers) + 1):
                        cell = ta.cell(r, c)
                        cell.fill = row_fill
                        cell.font = Font(color=C_TEXT)
                        cell.border = _thin_border()

                        if c in (3, 4, 5, 6, 7, 8):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.number_format = "0"
                        elif c in (9, 10):
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                stab = None
                for s in wb.sheetnames:
                    if s.lower() == "stability_audit":
                        stab = wb[s]
                if stab is None:
                    stab = wb.create_sheet("Stability_Audit")

                stab.delete_rows(1, stab.max_row)

                snap = getattr(self, "_last_eval_data", {}) or {}

                best_est = float(snap.get("best_estimate", snap.get("overall_weighted", snap.get("overall", 0.0))) or 0.0)
                exp_low = float(snap.get("expected_range_low", best_est - float(snap.get("ci", 0.0))) or 0.0)
                exp_high = float(snap.get("expected_range_high", best_est + float(snap.get("ci", 0.0))) or 0.0)
                exp_range_label = str(snap.get("expected_range_label", f"{exp_low:.2f}–{exp_high:.2f}"))
                checks = int(snap.get("checks", snap.get("n_kept", snap.get("n_runs", 0))) or 0)
                user_range_method = str(snap.get("user_range_method", "") or "")

                overall_live = float(snap.get("overall_weighted", snap.get("overall", 0.0)) or 0.0)
                raw_mean = float(snap.get("overall_ci_mean_raw", overall_live) or 0.0)
                raw_ci = float(snap.get("ci_raw", snap.get("ci", 0.0)) or 0.0)
                smoothed_mean = float(snap.get("overall_ci_mean_smoothed", snap.get("overall_ci_mean", overall_live)) or 0.0)
                smoothed_ci = float(snap.get("ci_smoothed", snap.get("ci", 0.0)) or 0.0)
                smooth_mode = str(snap.get("ci_smooth_mode", "raw") or "raw")
                raw_sd = float(snap.get("std_overall_raw", 0.0) or 0.0)
                internal_sd = float(snap.get("std_internal_process", 0.0) or 0.0)

                stab.append(["Item", "Value"])
                stab.append(["Best estimate", best_est])
                stab.append(["Expected range", exp_range_label])
                stab.append(["Checks", checks])
                stab.append(["User-facing range source", "Kept raw reruns"])
                stab.append(["User-facing range method", user_range_method])

                stab.append(["Technical stability source", "Smoothed signal"])
                stab.append(["Smoothing mode", smooth_mode])
                stab.append(["Internal aggregate overall", overall_live])
                stab.append(["Observed run pattern mean", raw_mean])
                stab.append(["Observed run pattern spread (raw CI)", raw_ci])
                stab.append(["Observed run pattern variation", raw_sd])
                stab.append(["Internal process variation", internal_sd])
                stab.append(["Internal stability mean", smoothed_mean])
                stab.append(["Internal stability spread", smoothed_ci])
                stab.append(["Stability goal", target])

                for c in range(1, 3):
                    cell = stab.cell(1, c)
                    cell.font = Font(b=True, color=C_TEXT)
                    cell.fill = _fill(C_HEAD)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = _thin_border()

                stab.freeze_panes = "A2"
                stab.column_dimensions["A"].width = 34
                stab.column_dimensions["B"].width = 24

                for r in range(2, stab.max_row + 1):
                    for c in range(1, 3):
                        cell = stab.cell(r, c)
                        cell.border = _thin_border()
                        cell.font = Font(color=C_TEXT)
                        cell.fill = _fill(C_ZEBRA if r % 2 else C_WHITE)
                        if c == 1:
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                for r in range(2, stab.max_row + 1):
                    val = stab.cell(r, 2).value
                    if isinstance(val, (int, float)):
                        stab.cell(r, 2).number_format = "0.0000"

                sp = None
                for s in wb.sheetnames:
                    if s.lower() == "stability_path":
                        sp = wb[s]
                if sp is None:
                    sp = wb.create_sheet("Stability_Path")

                sp.delete_rows(1, sp.max_row)

                sp_headers = [
                    "Run",
                    "Score",
                    "Average_raw",
                    "Stability_raw",
                    "Stability_shown",
                    "Guarded",
                    "Basis_size",
                ]
                sp.append(sp_headers)

                run_export = getattr(self, "_run_export", {}) or {}
                for run_no in sorted(run_export.keys()):
                    row = run_export.get(run_no, {}) or {}
                    sp.append([
                        int(run_no),
                        row.get("score", ""),
                        row.get("mu", ""),
                        row.get("ci", ""),
                        "Yes" if bool(row.get("show_ci", False)) else "No",
                        "Yes" if bool(row.get("guard", False)) else "No",
                        int(row.get("basis_m", 0) or 0),
                    ])

                for c in range(1, len(sp_headers) + 1):
                    cell = sp.cell(1, c)
                    cell.font = Font(b=True, color=C_TEXT)
                    cell.fill = _fill(C_HEAD)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = _thin_border()

                sp.freeze_panes = "A2"
                sp.auto_filter.ref = f"A1:G{max(2, sp.max_row)}"

                sp.column_dimensions["A"].width = 8
                sp.column_dimensions["B"].width = 10
                sp.column_dimensions["C"].width = 12
                sp.column_dimensions["D"].width = 12
                sp.column_dimensions["E"].width = 14
                sp.column_dimensions["F"].width = 10
                sp.column_dimensions["G"].width = 10

                for r in range(2, sp.max_row + 1):
                    guarded = str(sp.cell(r, 6).value or "").strip().lower() == "yes"
                    shown = str(sp.cell(r, 5).value or "").strip().lower() == "yes"

                    if guarded:
                        row_fill = _fill(C_DISCARD)
                    elif shown:
                        row_fill = _fill(C_KEPT)
                    else:
                        row_fill = _fill(C_ZEBRA if r % 2 else C_WHITE)

                    for c in range(1, len(sp_headers) + 1):
                        cell = sp.cell(r, c)
                        cell.fill = row_fill
                        cell.font = Font(color=C_TEXT)
                        cell.border = _thin_border()

                        if c in (1, 7):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.number_format = "0"
                        elif c in (2, 3, 4):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.number_format = "0.0000"
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                meta = None
                for s in wb.sheetnames:
                    if s.lower() == "meta":
                        meta = wb[s]
                if meta is None:
                    meta = wb.create_sheet("Meta")

                meta.delete_rows(1, meta.max_row)
                meta.append(["Stability goal", target])
                meta.append(["Source lang", self.source_language])
                meta.append(["Target lang", self.target_language])
                meta.append(["Model", self.model_name])
                meta.append(["Runs kept", max(0, tr.max_row - 1)])
                meta.append(["Discarded bullets (rows)", discarded_rows])
                meta.append(["User-facing estimate label", "Best estimate"])
                meta.append(["User-facing range label", "Expected range"])
                meta.append(["Technical support sheet", "Stability_Audit"])
                meta.append(["Technical path sheet", "Stability_Path"])
                meta.append(["Term logic", term_payload.get("logic_version", "obligation_v2")])
                meta.append(["Term obligations", int(term_payload.get("total_obligations", 0) or 0)])
                meta.append(["Term fulfilled", int(term_payload.get("fulfilled", 0) or 0)])
                meta.append(["Term missed", int(term_payload.get("missed", 0) or 0)])
                meta.append(["Term catastrophic", "Yes" if term_payload.get("catastrophic", False) else "No"])

                for r in range(1, meta.max_row+1):
                    meta.cell(r,1).font = Font(b=True, color=C_TEXT)
                    meta.cell(r,1).fill = _fill(C_HEAD)
                    meta.cell(r,1).alignment = Alignment(horizontal="left", vertical="center")
                    meta.cell(r,1).border = _thin_border()
                    meta.cell(r,2).fill = _fill(C_WHITE)
                    meta.cell(r,2).font = Font(color=C_SUB)
                    meta.cell(r,2).border = _thin_border()
                    if r == 1:
                        meta.cell(r,2).number_format = "0.0000"
                meta.column_dimensions["A"].width = 22
                meta.column_dimensions["B"].width = 34

                wb.save(fn)
                QMessageBox.information(self, "Saved", f"Exported to {fn}")
                return

            table_pm = self.runs_table.viewport().grab()
            chart_pm = self.live_chart.grab()
            gap = 16
            W = table_pm.width() + chart_pm.width() + gap
            H = max(table_pm.height(), chart_pm.height())

            if fn.lower().endswith(".png"):
                img = QImage(W, H, QImage.Format_ARGB32)
                img.fill(QColor(0,0,0,0))
                pa = QPainter(img)
                pa.fillRect(QRect(0,0,W,H), QColor("#0b1016"))
                pa.drawPixmap(0, 0, table_pm)
                pa.drawPixmap(table_pm.width()+gap, 0, chart_pm)
                pa.end()
                img.save(fn)
                QMessageBox.information(self, "Saved", f"Exported to {fn}")
                return

            from PyQt5.QtGui import QPdfWriter
            pdf = QPdfWriter(fn)
            pdf.setPageSizeMM(QSizeF(W/3.78, H/3.78))
            pdf.setResolution(96)
            pa = QPainter(pdf)
            pa.fillRect(QRect(0,0,W,H), QColor("#0b1016"))
            pa.drawPixmap(0, 0, table_pm)
            pa.drawPixmap(table_pm.width()+gap, 0, chart_pm)
            pa.end()
            QMessageBox.information(self, "Saved", f"Exported to {fn}")

        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))

    def export_report_html(self):
        data = getattr(self, "_last_eval_data", None)
        if not data:
            QMessageBox.information(self, "No results", "Run an evaluation first.")
            return

        fn, _ = QFileDialog.getSaveFileName(
            self, "Save Analysis Report (HTML)", "",
            "HTML Files (*.html *.htm)"
        )
        if not fn:
            return

        self._open_busy_popup("Preparing HTML report")

        th = QThread(self)
        worker = ReportBuildWorker(self, data)
        worker.moveToThread(th)

        def _done(labels: dict, html: str):
            try:
                self._bullet_labels = labels or {}
                wrapped = self._wrap_as_paged(html)
                with open(fn, "w", encoding="utf-8") as f:
                    f.write(wrapped)
                QMessageBox.information(self, "Saved", f"Report saved to {fn}")
            except Exception as e:
                QMessageBox.critical(self, "Save error", str(e))
            finally:
                self._close_busy_popup()

        def _fail(msg: str):
            try:
                QMessageBox.critical(self, "Report error", msg)
            finally:
                self._close_busy_popup()

        # signals
        worker.finished.connect(_done)
        worker.error.connect(_fail)

        # keep refs alive + quit/cleanup on both paths
        self._report_thread = th
        self._report_worker = worker
        worker.finished.connect(th.quit)
        worker.error.connect(th.quit)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        th.finished.connect(th.deleteLater)

        def _cleanup():
            self._report_thread = None
            self._report_worker = None
        th.finished.connect(_cleanup)

        th.started.connect(worker.run)
        th.start()

    def _teardown_live_screen(self):
        """Stop live-view timers safely."""
        try:
            if hasattr(self, "_row_pulse_timer") and self._row_pulse_timer:
                self._row_pulse_timer.stop()
                self._row_pulse_timer.deleteLater()
        except Exception:
            pass
        self._row_pulse_timer = None
        self._current_run_row = None

    def _freeze_scored_reasons(self, dimensions: dict):
        if not isinstance(dimensions, dict):
            return dimensions

        for key, info in dimensions.items():
            if not isinstance(info, dict):
                continue

            if "score" in info and ("justification" in info or "reasons" in info):
                raw = list(info.get("justification") or info.get("reasons") or [])
                info["scored_reasons"] = raw[:]
            else:
                self._freeze_scored_reasons(info)

        return dimensions

    def _sanitize_report_reasons(self, dim_name, score, reasons):
        """
        Final report-only cleanup.

        Does not change scores.
        Only rewrites or drops bullets for final-report polarity hygiene.

        Rules:
        - never add new bullets
        - never change scores
        - never hallucinate new facts or judgments
        - never change quoted snippets
        - strengths stay strengths (keep leading '✓ ')
        - major/critical concerns keep leading '! '
        - concerns must be problem-first and not open with praise
        """
        import re, json

        CONTRAST = re.compile(r'\b(?:but|however|although|though|while|yet)\b', re.I)

        NEG_CUE = re.compile(
            r'\b(?:'
            r'not|missing|omits?|omitted|absent|wrong|mismatch|shift(?:s|ed|ing)?|'
            r'less|too|overly|formal|institutional|unsupported|violation|replace(?:d|s|ment)?|'
            r'instead|narrow(?:s|ed|ing)?|weak(?:en|ens|ened|ening)?|'
            r'understate(?:s|d)?|overstate(?:s|d)?|not used|incorrect|fails?|'
            r'problem|issue|concern|reduces brief fulfillment'
            r')\b',
            re.I
        )

        POS_LEAD = re.compile(
            r'^\s*(?:'
            r'adds|keeps|helps|supports|improves|strengthens|preserves|captures|conveys|'
            r'reassures|normalizes|reduces|includes|provides|offers|'
            r'makes explicit|made explicit|is easy(?: to)?|is clear|is natural|is warm|is polite|'
            r'is professional|is direct|is readable|is fluent|is smooth|'
            r'used correctly|required term is used correctly|integrated naturally|'
            r'free and accessible|framed positively|stated clearly|made clear|'
            r'framed as acceptable|quick to scan|easy to scan|easy to read|easy to grasp|'
            r'explicitly respects|respects privacy|includes a clear|'
            r'help-seeking is framed|the symptom line is|the message is|the wording is|'
            r'it does include|it includes|it keeps|it adds|it helps|it supports'
            r')\b',
            re.I
        )

        QUOTE_RE = re.compile(r'“([^”]+)”|\"([^\"]+)\"|‘([^’]+)’|\'([^\'\n]+)\'|«([^»]+)»|‹([^›]+)›')

        def _extract_quote(s: str) -> str:
            m = QUOTE_RE.search(s or "")
            if not m:
                return ""
            return next((g for g in m.groups() if g), "")

        cleaned = []

        for reason in (reasons or []):
            s = str(reason or "").strip()
            is_strength = s.lstrip().startswith("✓")
            is_major = s.lstrip().startswith("!")

            if is_strength:
                cleaned.append(s)
                continue

            # if a concern has a concessive structure, keep only the negative side
            m = CONTRAST.search(s)
            if m:
                tail = s[m.end():].strip(" ,;:-")
                if tail:
                    s = tail

            # if a concern still starts like praise and contains no real negative cue, drop it
            if POS_LEAD.search(s) and not NEG_CUE.search(s):
                continue

            if is_major and not s.lstrip().startswith("!"):
                s = "! " + s.lstrip()
            cleaned.append(s)

        if dim_name == "Terminology Adherence" and float(score) <= 1e-9:
            cleaned = [
                s for s in cleaned
                if not str(s or "").lstrip().startswith("✓")
            ]

        cleaned = cleaned[:5]

        if not cleaned:
            return cleaned

        try:
            model = (self.model_name or "gpt-5.4-2026-03-05").strip()

            payload = []
            for idx, s in enumerate(cleaned):
                payload.append({
                    "index": idx,
                    "text": s,
                    "quote": _extract_quote(s),
                    "is_strength": str(s).lstrip().startswith("✓"),
                    "is_major": str(s).lstrip().startswith("!"),
                })

            system_msg = {
                "role": "system",
                "content": (
                    "You are polishing evaluation bullets for a final report. "
                    "You MUST NOT add facts, judgments, risks, implications, or evidence. "
                    "You MUST NOT strengthen severity. "
                    "You MUST NOT add new bullets. "
                    "You MUST preserve every quoted snippet exactly as given. "
                    "You may only rewrite wording for polarity hygiene or return an empty string to drop a bullet. "
                    "Strength bullets must remain strengths and keep the leading '✓ '. "
                    "Major/critical concern bullets must keep the leading '! '. "
                    "Concern bullets must be problem-first and must not open with praise, mitigation, or partial credit. "
                    "Do not output 'X is good, but Y...' or similar concessive structures for concerns. "
                    "Return valid JSON only."
                )
            }

            user_msg = {
                "role": "user",
                "content": (
                    "Rewrite each bullet for final-report clarity.\n\n"
                    "Hard rules:\n"
                    "1) Keep the same order and same indexes.\n"
                    "2) Do NOT add bullets.\n"
                    "3) Do NOT add facts or judgments.\n"
                    "4) Do NOT change any quoted snippet in any way.\n"
                    "5) Keep strengths as strengths (with leading '✓ ').\n"
                    "6) Keep major/critical concerns marked with leading '! '.\n"
                    "7) Rewrite concerns so they are fully negative/problem-first.\n"
                    "8) If a concern begins with praise or partial credit, rewrite it from the problem side.\n"
                    "9) If a bullet cannot be repaired without violating the rules, return an empty string for that bullet.\n"
                    "10) If Terminology Adherence score is 0.00, there must be no strength bullets.\n\n"
                    "Return exactly this schema:\n"
                    "{\"items\": [{\"index\": 0, \"text\": \"...\"}]}\n\n"
                    f"Dimension: {dim_name}\n"
                    f"Score: {float(score):.2f}\n"
                    f"Items: {json.dumps(payload, ensure_ascii=False)}"
                )
            }

            rsp = client.chat.completions.create(
                model=model,
                messages=[system_msg, user_msg],
                max_completion_tokens=128000,
                temperature=0.0,
                response_format={"type": "json_object"},
                **_gpt5_effort_none_kwargs(model),
            )

            raw = (rsp.choices[0].message.content or "").strip()
            parsed = json.loads(raw)
            items = parsed.get("items") or []

            repaired = []
            by_index = {}
            for it in items:
                try:
                    by_index[int(it.get("index"))] = str(it.get("text") or "")
                except Exception:
                    pass

            for idx, original in enumerate(cleaned):
                candidate = by_index.get(idx, original).strip()

                if candidate == "":
                    continue

                orig_quote = _extract_quote(original)
                cand_quote = _extract_quote(candidate)

                if orig_quote != cand_quote:
                    repaired.append(original)
                    continue

                if dim_name == "Terminology Adherence" and float(score) <= 1e-9:
                    if candidate.lstrip().startswith("✓"):
                        continue

                repaired.append(candidate)

            return repaired[:5]

        except Exception:
            return cleaned[:5]

    def _build_report_html(self, data):
        """Build the final HTML report in the app's dark theme."""
        from PyQt5.QtCore import QDate
        import html, re

        SENT_END = re.compile(r'(?<=[.!?])\s+')

        def _sentencise(html_blob: str, prefix: str) -> list[str]:
            """HTML → text → sentence spans."""
            txt = strip_html_tags(html_blob or "")
            txt = re.sub(r"\s+", " ", txt).strip()
            sents = [s.strip() for s in SENT_END.split(txt) if s.strip()]
            return [f'<span id="{prefix}{i}">{html.escape(s)}</span>' for i, s in enumerate(sents)]

        orig_html = self.extract_text_with_formatting(self.file_path)
        tran_html = getattr(self, "_last_translation_html", "")
        term_payload = self._get_term_audit_payload(data)
        orig_spans = _sentencise(orig_html, "s")
        tran_spans = _sentencise(tran_html, "t")

        span_map: dict[str, str] = {}
        for i, span in enumerate(orig_spans):
            key = re.sub(r"<[^>]+>", "", span)[:80]
            span_map[key] = f"s{i}"
        for i, span in enumerate(tran_spans):
            key = re.sub(r"<[^>]+>", "", span)[:80]
            span_map[key] = f"t{i}"

        def _find_id(snippet: str) -> str:
            """Return the span id whose sentence contains the quoted snippet."""
            q = (snippet or "").strip()[:80].lower()
            for sent, _id in span_map.items():
                if q and q in sent.lower():
                    return _id
            return ""

        def _term_findings_html(payload: dict) -> str:
            rows = list(payload.get("rows") or [])
            violated = list(payload.get("violated_rows") or [])
            items = []

            for r in violated[:3]:
                msg = f'{r["source_term"]} → {r["target_term"]}: {r["fulfilled"]}/{r["source_occurrences"]} obligations fulfilled'
                if r["alternative_used"]:
                    msg += "; competing alternative lexicalization detected"
                elif r["missed"] > 0:
                    msg += f'; {r["missed"]} obligation(s) missed'
                if r["reason"]:
                    msg += f" — {r['reason']}"
                items.append(f"<li>{html.escape(msg)}</li>")

            if not items:
                hidden_rows = [r for r in rows if r["hidden_fulfilled"] > 0]
                for r in hidden_rows[:2]:
                    msg = (
                        f'{r["source_term"]} → {r["target_term"]}: '
                        f'{r["hidden_fulfilled"]} obligation(s) counted as hidden-but-fulfilled after restructuring'
                    )
                    if r["reason"]:
                        msg += f" — {r['reason']}"
                    items.append(f"<li>{html.escape(msg)}</li>")

            if not items:
                ok_rows = [r for r in rows if r["source_occurrences"] > 0 and r["missed"] == 0]
                for r in ok_rows[:2]:
                    msg = (
                        f'{r["source_term"]} → {r["target_term"]}: '
                        f'all {r["source_occurrences"]} obligation(s) fulfilled'
                    )
                    items.append(f"<li>{html.escape(msg)}</li>")

            return "".join(items) or "<li>—</li>"

        labels_map = getattr(self, "_bullet_labels", {}) or {}

        def _fallback_label(s: str) -> str:
            s = (s or "").strip()
            low = s.lower()
            head = s[:1]

            neg = bool(re.search(
                r"\b("
                r"does\s+not\s+use|doesn['’]t\s+use|avoid(?:s|ed)?|"
                r"replac(?:e|es|ed|ing)|instead\s+of|alternate\s+to|alternative\s+to|"
                r"wrong|mismatch|fail|error|incorrect|inconsistent|"
                r"not\s+.*(aligned|consistent)|required\s+(but\s+)?absent|missing\s+required\s+term|"
                r"lowered|violation|contradiction|mistranslat"
                r")\b", re.I
            ))

            pos = (
                head in ("+", "✓", "✅") or
                bool(re.search(
                    r"\b("
                    r"correct(ly)?\s+use[s]?|uses\s+required\s+term|"
                    r"aligned|appropriate|consistent|natural|clear|accurate|faithful|"
                    r"improved|well[- ]structured|meets\s+brief|follows\s+(instructions|format)"
                    r")\b", re.I
                ))
            )

            if neg and pos:
                return "neg"
            if neg:
                return "neg"
            return "pos" if pos else "neg"

        css = """
        body{background:#0f1115;color:#E6E6E6;font-family:Segoe UI,Inter,Arial,sans-serif;margin:0;padding:18px}
        h1{color:#88FF88;margin:6px 0 4px 0}
        h2{color:#88CCFF;margin:8px 0}
        h3{color:#CFE2FF;margin:10px 0 6px}
        .card{background:#171a21;border:1px solid #232a36;border-radius:16px;padding:14px;margin:12px 0;}
        .meta{display:flex;gap:12px;flex-wrap:wrap;color:#c7d0e0}
        .pill{background:#14233e;border:1px solid #2c58a0;border-radius:999px;padding:6px 10px;font-size:12px}
        table{width:100%;table-layout:fixed;border-collapse:separate;border-spacing:0 6px;margin:8px 0}
        th,td{background:#10131a;border:1px solid #252b36;padding:8px 10px;vertical-align:top;word-break:break-word;}
        th{background:#171a21;color:#c7d0e0;border-radius:12px 12px 0 0;border-bottom:1px solid #252b36;}
        tr td:first-child{border-radius:10px 0 0 10px}
        tr td:last-child{border-radius:0 10px 10px 0}
        a.bullet{color:#FFC36A;text-decoration:none;}
        a.bullet:hover{filter:brightness(1.1);}
        .hl{background:#55550033;transition:background .2s;}

        .headrow{display:flex;align-items:flex-start;justify-content:space-between;gap:12px}
        .header-left{display:flex;flex-direction:column;gap:8px}
        .header-left .meta{margin-top:2px}
        .kpi{min-width:200px;text-align:right;background:#10151f;border:1px solid #252b36;border-radius:12px;padding:10px 12px}
        .kpi .big{font-size:36px;font-weight:800;color:#cfe2ff;line-height:1}
        .kpi .sub{font-size:13px;color:#9fb2d0;margin-top:4px}
        .kpi .pm{color:#7be3a4}

        .col1{display:grid;grid-template-columns:1fr;gap:14px}
        .box{background:#10151f;border:1px solid #252b36;border-radius:12px;padding:10px}
        .box h4{margin:0 0 6px 0;color:#cfe2ff}
        ul{margin:0;padding-left:18px}
        ul.good li a{color:#7be3a4}
        ul.bad  li a{color:#ffb3b3}
        .term-grid{display:grid;grid-template-columns:repeat(3,minmax(120px,1fr));gap:10px;margin:10px 0 12px}
        .term-stat{background:#10151f;border:1px solid #252b36;border-radius:12px;padding:10px}
        .term-stat .k{font-size:12px;color:#9fb2d0;margin-bottom:4px}
        .term-stat .v{font-size:20px;font-weight:800;color:#cfe2ff;line-height:1.15}
        .term-box{background:#10151f;border:1px solid #252b36;border-radius:12px;padding:10px;margin-top:10px}
        .term-box h4{margin:0 0 6px 0;color:#cfe2ff}
        .term-note{font-size:12px;color:#9fb2d0;margin-top:8px}
        """

        js = """
        <script>
          document.querySelectorAll('a.bullet').forEach(a=>{
            const id=a.dataset.ref;
            const tgt=id && document.getElementById(id);
            const enter=()=>{a.classList.add('hl');tgt&&tgt.classList.add('hl');};
            const leave=()=>{a.classList.remove('hl');tgt&&tgt.classList.remove('hl');};
            a.addEventListener('mouseenter',enter);
            a.addEventListener('mouseleave',leave);
          });
        </script>
        """

        html_chunks = ['<html><head><meta charset="utf-8"><style>', css, '</style></head><body>']

        best_hdr = float(data.get("best_estimate", data.get("overall_weighted", data.get("overall", 0.0))))
        exp_low_hdr = float(data.get("expected_range_low", best_hdr - float(data.get("ci", 0.0))))
        exp_high_hdr = float(data.get("expected_range_high", best_hdr + float(data.get("ci", 0.0))))
        exp_range_hdr = str(data.get("expected_range_label", f"{exp_low_hdr:.2f}–{exp_high_hdr:.2f}"))
        checks_hdr = int(data.get("checks", data.get("n_kept", data.get("n_runs", 0))))

        html_chunks += [
            '<div class="card"><div class="headrow">',
            '<div class="header-left"><h1>PAEM-CMT Report</h1>',
            '<div class="meta">',
            f'<span class="pill">Lang: {html.escape(self.source_language)} → {html.escape(self.target_language)}</span>',
            f'<span class="pill">Model: {html.escape(self.model_name)}</span>',
            f'<span class="pill">Date: {QDate.currentDate().toString()}</span>',
            '</div></div>',
            f'<div class="kpi"><div class="big">{best_hdr:.2f}</div>',
            f'<div class="sub">Best estimate across kept runs • Expected range: {html.escape(exp_range_hdr)}</div></div>',
            '</div></div>'
        ]

        html_chunks += [
            '<div class="card"><h2>Original vs. Translation</h2>',
            '<table style="table-layout:fixed;width:100%">',
            '<colgroup>',
            '<col style="width:50%"><col style="width:50%">',
            '</colgroup>',
            '<tr><th>Original</th><th>Translation</th></tr>',
            '<tr><td>',
            '<br>'.join(orig_spans) or "_",
            '</td><td>',
            '<br>'.join(tran_spans) or "_",
            '</td></tr></table></div>',
        ]

        role_dims = {
            "Intended Purpose",
            "Target Audience",
            "Translator",
            "Source Owner",
            "Commissioner",
        }

        catastrophic_terms = bool(term_payload.get("catastrophic", False))

        def _show_strengths_for_dimension(label: str, score: float) -> bool:
            if float(score) < 2.50:
                return False

            if catastrophic_terms and label in role_dims:
                return False

            return True

        def _is_major_concern_text(reason_text: str) -> bool:
            return str(reason_text or "").lstrip().startswith("!")

        def _single_column_title(label: str, score: float, has_pos: bool, has_major: bool, has_minor: bool) -> str:
            if (not has_pos) and (not has_major) and (not has_minor):
                return ""

            if float(score) >= 4.00 and (not has_pos) and (not has_major) and has_minor:
                return "Notes"

            if has_pos and (not has_major) and (not has_minor):
                return "Strengths"

            return "Concerns"

        BLURB = {
            "Intended Purpose":      "Is the translation laser-focused on the communicative goal?",
            "Target Audience":       "Is the language tailored to the readers’ culture and knowledge?",
            "Translator":            "Does the translator’s strategy serve the task effectively?",
            "Source Owner":          "Does it respect the owner’s intent, tone and constraints?",
            "Commissioner":          "Does it meet the commissioner’s brief and risk expectations?",
            "Terminology Adherence": "Are mandatory terms used consistently and naturally?"
        }

        html_chunks.append('<div class="card"><h2>Quality profile</h2>')
        dims = data.get("dimensions", {}) or {}
        keymap = {k.lower(): k for k in dims.keys()}
        for label in self.STANDARD_HEADINGS:
            kk = keymap.get(label.lower())
            if not kk:
                continue
            info = dims[kk]

            score = float(info.get("score", 0.0))
            html_chunks += [
                f"<h3>• {html.escape(label)}: {score:.2f}/5.00</h3>",
                f"<p style='font-style:italic;color:#AAB; margin-top:-4px;'> {html.escape(BLURB.get(label,''))}</p>"
            ]
            if label == "Terminology Adherence":
                tp = term_payload
                html_chunks += [
                    "<div class='term-grid'>",
                        f"<div class='term-stat'><div class='k'>Terms</div><div class='v'>{int(tp.get('total_obligations', 0) or 0)}</div></div>",
                        f"<div class='term-stat'><div class='k'>Fulfilled</div><div class='v'>{int(tp.get('fulfilled', 0) or 0)}</div></div>",
                        f"<div class='term-stat'><div class='k'>Missed</div><div class='v'>{int(tp.get('missed', 0) or 0)}</div></div>",
                    "</div>",
                    "<div class='term-box'>",
                        "<h4>Key findings</h4>",
                        f"<ul class='bad'>{_term_findings_html(tp)}</ul>",
                        "<div class='term-note'>Detailed pair-by-pair terminology audit is available in the XLSX sheet <b>Terminology_Audit</b>.</div>",
                        (
                            f"<div class='term-note'><b>Term check status:</b> "
                            f"{html.escape(str(tp.get('adjudication_status', 'ok')))}</div>"
                            if str(tp.get("adjudication_status", "ok")) != "ok" else ""
                        ),
                    "</div>",
                    "<div style='height:8px'></div>",
                ]
                continue

            raw_reasons = list((info.get("scored_reasons") or info.get("reasons") or info.get("justification") or [])[:])

            reasons = self._sanitize_report_reasons(
                label,
                score,
                raw_reasons[:]
            )

            pos_items, neg_major_items, neg_minor_items = [], [], []
            for reason in reasons:
                m = re.search(r'“([^”]+)”|\"([^"]+)\"|‘([^’]+)’|\'([^\'\\n]+)\'|«([^»]+)»|‹([^›]+)›', reason)
                snippet = next((g for g in (
                    m.group(1) if m else None, m.group(2) if m else None,
                    m.group(3) if m else None, m.group(4) if m else None,
                    m.group(5) if m else None, m.group(6) if m else None
                ) if g), "")
                ref_id = _find_id(snippet)

                raw_reason = str(reason).lstrip()
                is_strength = raw_reason.startswith("✓")
                is_major = _is_major_concern_text(raw_reason)

                shown_reason = raw_reason
                if is_strength:
                    shown_reason = shown_reason[1:].lstrip()
                elif is_major:
                    shown_reason = shown_reason[1:].lstrip()

                link = f'<a class="bullet" data-ref="{ref_id}">{html.escape(shown_reason)}</a>'

                if is_strength:
                    pos_items.append(f"<li>{link}</li>")
                elif is_major:
                    neg_major_items.append(f"<li>{link}</li>")
                else:
                    neg_minor_items.append(f"<li>{link}</li>")

            show_strengths = _show_strengths_for_dimension(label, score)
            has_pos = bool(pos_items)
            has_major = bool(neg_major_items)
            has_minor = bool(neg_minor_items)

            if show_strengths and has_pos and has_major:
                html_chunks += [
                    "<div class='col2'>",
                        "<div class='box'><h4>Strengths</h4>",
                            f"<ul class='good'>{''.join(pos_items)}</ul>",
                        "</div>",
                        "<div class='box'><h4>Concerns</h4>",
                            f"<ul class='bad'>{''.join(neg_major_items)}</ul>",
                        "</div>",
                    "</div>",
                    "<div style='height:8px'></div>"
                ]

            elif show_strengths and has_pos and (not has_major) and has_minor:
                html_chunks += [
                    "<div class='col2'>",
                        "<div class='box'><h4>Strengths</h4>",
                            f"<ul class='good'>{''.join(pos_items)}</ul>",
                        "</div>",
                        "<div class='box'><h4>Notes</h4>",
                            f"<ul class='bad'>{''.join(neg_minor_items)}</ul>",
                        "</div>",
                    "</div>",
                    "<div style='height:8px'></div>"
                ]

            else:
                title = _single_column_title(label, score, has_pos, has_major, has_minor)

                if not title:
                    pass
                else:
                    if title == "Strengths":
                        items_html = ''.join(pos_items) or '<li>—</li>'
                        list_class = "good"
                    elif title == "Notes":
                        items_html = ''.join(neg_minor_items) or '<li>—</li>'
                        list_class = "bad"
                    else:
                        items_html = ''.join(neg_major_items + ([] if show_strengths else neg_minor_items)) or '<li>—</li>'
                        list_class = "bad"

                    html_chunks += [
                        "<div class='col1'>",
                            f"<div class='box'><h4>{title}</h4>",
                                f"<ul class='{list_class}'>{items_html}</ul>",
                            "</div>",
                        "</div>",
                        "<div style='height:8px'></div>"
                    ]
        html_chunks.append('</div>')

        best_est = float(data.get("best_estimate", data.get("overall_weighted", data.get("overall", 0.0))))
        exp_low = float(data.get("expected_range_low", best_est - float(data.get("ci", 0.0))))
        exp_high = float(data.get("expected_range_high", best_est + float(data.get("ci", 0.0))))
        exp_range_label = str(data.get("expected_range_label", f"{exp_low:.2f}–{exp_high:.2f}"))
        checks = int(data.get("checks", data.get("n_kept", data.get("n_runs", 0))))

        html_chunks += [
            f'<div class="card"><h2>Best estimate across kept runs: {best_est:.2f}</h2>',
            f"<p>Expected range: {html.escape(exp_range_label)}. Checks: {checks}.</p></div>"
        ]

        if getattr(self, "_audit_log", []):
            html_chunks.append('<div class="card"><h2>Audit — Discarded Bullets</h2>')
            html_chunks.append('<table><tr><th>Run</th><th>Dimension</th><th>Bullet</th><th>Why</th></tr>')
            for run_no, dim, just, why in self._audit_log:
                html_chunks.append(
                    "<tr><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>".format(
                        run_no, html.escape(dim or "?"), html.escape(just or ""), html.escape(why or "")
                    )
                )
            html_chunks.append('</table></div>')

        html_chunks += [js, '</body></html>']
        return "".join(html_chunks)

    def extract_text_with_formatting(self, doc_path):
        """Minimal .docx → HTML conversion with safe escaping."""
        from html import escape
        doc = Document(doc_path)
        html = []
        open_list = None

        def close_list():
            nonlocal open_list
            if open_list:
                html.append(f"</{open_list}>")
                open_list = None

        for p in doc.paragraphs:
            style = (p.style.name or "").strip()

            if style.startswith("Heading"):
                close_list()
                m = re.search(r"(\d+)$", style)
                lvl = int(m.group(1)) if m else 1
                lvl = min(6, max(1, lvl))
                html.append(f"<h{lvl}>{escape(p.text)}</h{lvl}>")
                continue

            if style in ("List Bullet", "List Number"):
                tag = "ul" if "Bullet" in style else "ol"
                if open_list != tag:
                    close_list()
                    html.append(f"<{tag}>")
                    open_list = tag
                txt = p.text or ""
                if style == "List Number" and len(txt) > 2 and txt[1] == ")":
                    txt = txt[2:].strip()
                html.append(f"<li>{escape(txt)}</li>")
                continue

            close_list()
            if p.text.strip():
                parts = []
                for run in p.runs:
                    t = escape(run.text or "")
                    if not t:
                        continue
                    if run.bold:      t = f"<b>{t}</b>"
                    if run.italic:    t = f"<i>{t}</i>"
                    if run.underline: t = f"<u>{t}</u>"
                    parts.append(t)
                if parts:
                    html.append("<p>" + "".join(parts) + "</p>")

        close_list()
        return "\n".join(html).strip()

    def clear_layout(self, layout=None, keep_header=False):
        """Recursively delete widgets and layouts."""
        if layout is None: layout = self.layout
        start_idx = 1 if keep_header and layout.count() else 0
        while layout.count() > start_idx:
            c = layout.takeAt(start_idx)
            w = c.widget()
            if w: w.deleteLater()
            elif c.layout(): self.clear_layout(c.layout())

    def closeEvent(self, event):
        """Restore stdout/stderr and stop active threads on exit."""
        try:
            for attr in ("_eval_thread", "_qf_thread", "_purp_thread", "_report_thread"):
                th = getattr(self, attr, None)
                if th and hasattr(th, "isRunning") and th.isRunning():
                    try:
                        th.quit()
                        th.wait(1500)
                    except Exception:
                        pass
        except Exception:
            pass

        try:
            sys.stdout = self._stdout_orig
            sys.stderr = self._stderr_orig
        except Exception:
            pass
        super().closeEvent(event)
        
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(resource_path("paemcmtlogo.ico")))

    import os
    k = os.getenv("OPENAI_API_KEY", "").strip()
    if k:
        set_api_key(k)
        client = new_client()

    splash = SplashScreen()
    splash.show()
    splash.raise_()
    app.processEvents()

    start = time.time()

    ex = TranslationTool()
    ex.showMaximized()

    while time.time() - start < 3.4:
        app.processEvents()
        time.sleep(0.01)

    splash.fade_finish(ex)

    sys.exit(app.exec_())